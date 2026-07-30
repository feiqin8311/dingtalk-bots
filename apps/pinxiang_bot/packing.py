#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""不分仓拼箱算法（与 dingtalk-lcl-bot 独立）。

箱规硬约束：
  A. 单箱毛重 1~18 kg
  B. 非原箱规：长宽高单边 15~60 cm
  C. 单箱体积 < 10000 cm³ → 长=宽=高=20

箱数比 r = 发货数量 / 单箱数量：
  r 整数 ≥1 → 原箱（不进拼箱结果表）
  0 < r < 1 → 不足一箱：重 1~18 单装(B)；重 <1 与同仓轻货合箱(A)
  r > 1 非整数 → 优先整票拼 1 箱(C，毛重 1~18)；否则整数原箱(D)+余数；余数 <1kg 借 1 箱合并
"""

from __future__ import annotations

import math
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Mapping, Optional, Sequence, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

PathLike = Union[str, Path]

MIN_BOX_WEIGHT_KG = 1.0
MAX_BOX_WEIGHT_KG = 18.0
MIN_SIDE_CM = 15.0
MAX_SIDE_CM = 60.0
SMALL_VOLUME_CM3 = 10000.0
DEFAULT_SIDE_CM = 20.0

PACKING_HEADERS = (
    "发货仓库",
    "SKU",
    "发货数量",
    "单箱数量",
    "箱数",
    "单箱毛重",
    "长",
    "宽",
    "高",
)


@dataclass(frozen=True)
class ProductSpec:
    sku: str
    units_per_box: Optional[float] = None
    box_weight_kg: Optional[float] = None
    length_cm: Optional[float] = None
    width_cm: Optional[float] = None
    height_cm: Optional[float] = None


@dataclass
class LineItem:
    shipment_sn: str
    warehouse: str
    sku: str
    msku: str
    product_name: str
    qty: float
    units_per_box: float
    box_weight_kg: float
    length_cm: float
    width_cm: float
    height_cm: float


@dataclass
class PackingRow:
    warehouse: str
    sku: str
    msku: str
    qty: float
    units_per_box: float
    box_count: float
    box_weight_kg: float
    length_cm: float
    width_cm: float
    height_cm: float
    remark: str = ""
    # 同一 box_group_id = 同一物理箱（多 SKU 合箱时多行共享）
    box_group_id: str = ""

    def as_dict(self) -> dict[str, Any]:
        return {
            "发货仓库": self.warehouse,
            "SKU": self.sku,
            "发货数量": _num(self.qty),
            "单箱数量": _num(self.units_per_box),
            "箱数": _num(self.box_count),
            "单箱毛重": _round2(self.box_weight_kg),
            "长": _round1(self.length_cm),
            "宽": _round1(self.width_cm),
            "高": _round1(self.height_cm),
        }


@dataclass
class PackingResult:
    rows: list[PackingRow] = field(default_factory=list)
    all_rows: list[PackingRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    shipment_sns: list[str] = field(default_factory=list)

    @property
    def amazon_rows(self) -> list[PackingRow]:
        return list(self.all_rows or self.rows)

    def result_basename(self, fallback: str = "未知") -> str:
        """拼箱结果文件名（无扩展名），如：SP1 SP2 SP3 拼箱数据。"""
        sns = [s.strip() for s in self.shipment_sns if s and str(s).strip()]
        if sns:
            return f"{' '.join(sns)} 拼箱数据"
        return f"{fallback} 拼箱数据"


@dataclass
class _Partial:
    """待改箱/合箱的零头或不足一箱。"""

    warehouse: str
    sku: str
    msku: str
    qty: float
    unit_w: float
    unit_vol: float
    orig_l: float
    orig_w: float
    orig_h: float
    orig_units_per_box: float
    orig_box_weight: float
    source: str  # under_one | remainder | all_in_one

    @property
    def weight(self) -> float:
        return self.unit_w * self.qty

    @property
    def volume(self) -> float:
        return self.unit_vol * self.qty


def process_shipment_file(
    input_path: PathLike,
    *,
    product_specs: Optional[Mapping[str, ProductSpec]] = None,
) -> PackingResult:
    wb = load_workbook(filename=str(input_path), data_only=True, read_only=True)
    try:
        pack_sheet = _require_sheet(wb, "装箱信息")
        detail_sheet = _require_sheet(wb, "发货单详情")
        pack_rows = list(pack_sheet.iter_rows(values_only=True))
        detail_rows = list(detail_sheet.iter_rows(values_only=True))
    finally:
        wb.close()
    items = _build_line_items(pack_rows, detail_rows, product_specs or {})
    return compute_packing(items)


def compute_packing(items: Sequence[LineItem]) -> PackingResult:
    result = PackingResult()
    # 原箱整票（仅 all_rows）
    pure_originals: list[PackingRow] = []
    # D 的整数原箱部分（进结果表，可被借箱）
    integer_parts: list[PackingRow] = []
    partials: list[_Partial] = []
    seen_sn: set[str] = set()

    for item in items:
        sn = (item.shipment_sn or "").strip()
        if sn and sn not in seen_sn:
            seen_sn.add(sn)
            result.shipment_sns.append(sn)
        if item.units_per_box <= 0:
            result.warnings.append(f"{item.sku}: 单箱数量无效，已跳过")
            continue
        ratio = item.qty / item.units_per_box
        unit_w = item.box_weight_kg / item.units_per_box
        unit_vol = (item.length_cm * item.width_cm * item.height_cm) / item.units_per_box

        # —— 整箱原箱 ——
        if _is_integer(ratio) and ratio >= 1:
            pure_originals.append(
                PackingRow(
                    warehouse=item.warehouse,
                    sku=item.sku,
                    msku=item.msku or item.sku,
                    qty=item.qty,
                    units_per_box=item.units_per_box,
                    box_count=ratio,
                    box_weight_kg=item.box_weight_kg,
                    length_cm=item.length_cm,
                    width_cm=item.width_cm,
                    height_cm=item.height_cm,
                    remark="原箱",
                )
            )
            continue

        # —— >1 非整数：优先 C 整票拼 1 箱（毛重 1~18），否则 D 拆整数+余数 ——
        if ratio > 1:
            all_weight = unit_w * item.qty
            if MIN_BOX_WEIGHT_KG <= all_weight <= MAX_BOX_WEIGHT_KG:
                partials.append(
                    _Partial(
                        warehouse=item.warehouse,
                        sku=item.sku,
                        msku=item.msku or item.sku,
                        qty=item.qty,
                        unit_w=unit_w,
                        unit_vol=unit_vol,
                        orig_l=item.length_cm,
                        orig_w=item.width_cm,
                        orig_h=item.height_cm,
                        orig_units_per_box=item.units_per_box,
                        orig_box_weight=item.box_weight_kg,
                        source="all_in_one",
                    )
                )
                continue

            full_boxes = int(math.floor(ratio))
            full_qty = full_boxes * item.units_per_box
            rem_qty = item.qty - full_qty
            if full_boxes > 0:
                integer_parts.append(
                    PackingRow(
                        warehouse=item.warehouse,
                        sku=item.sku,
                        msku=item.msku or item.sku,
                        qty=full_qty,
                        units_per_box=item.units_per_box,
                        box_count=float(full_boxes),
                        box_weight_kg=item.box_weight_kg,
                        length_cm=item.length_cm,
                        width_cm=item.width_cm,
                        height_cm=item.height_cm,
                        remark="原箱-整数部分",
                    )
                )
            if rem_qty > 0:
                partials.append(
                    _Partial(
                        warehouse=item.warehouse,
                        sku=item.sku,
                        msku=item.msku or item.sku,
                        qty=rem_qty,
                        unit_w=unit_w,
                        unit_vol=unit_vol,
                        orig_l=item.length_cm,
                        orig_w=item.width_cm,
                        orig_h=item.height_cm,
                        orig_units_per_box=item.units_per_box,
                        orig_box_weight=item.box_weight_kg,
                        source="remainder",
                    )
                )
            continue

        # —— <1 不足一箱 ——
        if 0 < ratio < 1:
            partials.append(
                _Partial(
                    warehouse=item.warehouse,
                    sku=item.sku,
                    msku=item.msku or item.sku,
                    qty=item.qty,
                    unit_w=unit_w,
                    unit_vol=unit_vol,
                    orig_l=item.length_cm,
                    orig_w=item.width_cm,
                    orig_h=item.height_cm,
                    orig_units_per_box=item.units_per_box,
                    orig_box_weight=item.box_weight_kg,
                    source="under_one",
                )
            )
            continue

        result.warnings.append(f"{item.sku}: 发货数量异常 qty={item.qty}")

    # D 备注：余数毛重 <1 → 从同 SKU 整数箱借 1 箱
    integer_parts, partials, borrow_rows = _apply_borrow_from_integer(integer_parts, partials)

    # 分类处理 partials：B 单装 / A 合箱 / 超重告警
    changed_rows: list[PackingRow] = list(integer_parts) + list(borrow_rows)
    changed_rows.extend(_pack_partials_by_warehouse(partials, result.warnings))

    result.rows = changed_rows
    result.all_rows = pure_originals + list(changed_rows)
    return result


def write_packing_workbook(result: PackingResult, output_path: PathLike) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "拼箱结果"
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    for col, header in enumerate(PACKING_HEADERS, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.border = thin
        cell.alignment = center
    for r_idx, row in enumerate(result.rows, start=2):
        data = row.as_dict()
        for c_idx, key in enumerate(PACKING_HEADERS, start=1):
            cell = ws.cell(r_idx, c_idx, data.get(key))
            cell.border = thin
            cell.alignment = center
    if result.warnings:
        ws2 = wb.create_sheet("警告")
        ws2.append(["警告"])
        for msg in result.warnings:
            ws2.append([msg])
    wb.save(path)
    return path


# --- 合箱 / 借箱 ---


def _apply_borrow_from_integer(
    integer_parts: list[PackingRow],
    partials: list[_Partial],
) -> tuple[list[PackingRow], list[_Partial], list[PackingRow]]:
    """余数/改箱毛重 <1kg：同 SKU 从整数原箱借 1 箱合并成一箱非原箱规。"""
    parts = list(integer_parts)
    left: list[_Partial] = []
    borrowed: list[PackingRow] = []

    for p in partials:
        if p.weight >= MIN_BOX_WEIGHT_KG or p.source != "remainder":
            left.append(p)
            continue
        # 找同 SKU 整数箱
        idx = next((i for i, row in enumerate(parts) if row.sku == p.sku and row.box_count >= 1), None)
        if idx is None:
            left.append(p)
            continue
        full = parts[idx]
        unit = full.units_per_box
        new_count = full.box_count - 1
        merged_qty = p.qty + unit
        unit_w = full.box_weight_kg / unit if unit else p.unit_w
        unit_vol = (full.length_cm * full.width_cm * full.height_cm) / unit if unit else p.unit_vol
        weight = unit_w * merged_qty
        vol = unit_vol * merged_qty
        length, width, height = _dims_non_original(vol)
        if new_count <= 0:
            parts.pop(idx)
        else:
            parts[idx] = PackingRow(
                warehouse=full.warehouse,
                sku=full.sku,
                msku=full.msku,
                qty=new_count * unit,
                units_per_box=unit,
                box_count=new_count,
                box_weight_kg=full.box_weight_kg,
                length_cm=full.length_cm,
                width_cm=full.width_cm,
                height_cm=full.height_cm,
                remark=full.remark,
            )
        gid = uuid.uuid4().hex[:8]
        borrowed.append(
            PackingRow(
                warehouse=p.warehouse,
                sku=p.sku,
                msku=p.msku,
                qty=merged_qty,
                units_per_box=merged_qty,
                box_count=1.0,
                box_weight_kg=weight,
                length_cm=length,
                width_cm=width,
                height_cm=height,
                remark="借箱合并",
                box_group_id=gid,
            )
        )
    return parts, left, borrowed


def _pack_partials_by_warehouse(partials: list[_Partial], warnings: list[str]) -> list[PackingRow]:
    """同仓库：B 单装(1~18kg)；A 轻货合箱(<1kg)；>18 告警仍单装。"""
    by_wh: dict[str, list[_Partial]] = {}
    for p in partials:
        by_wh.setdefault(p.warehouse, []).append(p)

    out: list[PackingRow] = []
    for warehouse, group in by_wh.items():
        singles: list[_Partial] = []
        lights: list[_Partial] = []
        for p in group:
            w = p.weight
            if w < MIN_BOX_WEIGHT_KG:
                lights.append(p)
            elif w > MAX_BOX_WEIGHT_KG:
                warnings.append(
                    f"{p.sku}: 改箱毛重 {w:.2f}kg 超过 {MAX_BOX_WEIGHT_KG}kg，仍按一箱输出请人工核对"
                )
                singles.append(p)
            else:
                singles.append(p)

        # B：单装
        for p in singles:
            out.append(_partial_to_solo_row(p, remark=_remark_for_solo(p)))

        # A：轻货贪心合箱（同仓，目标合计 1~18kg）
        lights.sort(key=lambda x: x.weight)
        bins: list[list[_Partial]] = []
        for p in lights:
            placed = False
            for b in bins:
                total = sum(x.weight for x in b) + p.weight
                if total <= MAX_BOX_WEIGHT_KG:
                    b.append(p)
                    placed = True
                    break
            if not placed:
                bins.append([p])

        for b in bins:
            total_w = sum(x.weight for x in b)
            # 合完仍 <1：告警，仍输出（人工处理）
            if total_w < MIN_BOX_WEIGHT_KG:
                warnings.append(
                    f"仓库[{warehouse}] 合箱后毛重仍 <{MIN_BOX_WEIGHT_KG}kg："
                    + ",".join(x.sku for x in b)
                )
            if len(b) == 1:
                out.append(_partial_to_solo_row(b[0], remark=_remark_for_solo(b[0])))
            else:
                out.extend(_partials_to_mixed_rows(b))
    return out


def _remark_for_solo(p: _Partial) -> str:
    if p.source == "all_in_one":
        return "整票拼1箱"
    if p.source == "remainder":
        return "余数改箱"
    return "不足一箱改箱"


def _partial_to_solo_row(p: _Partial, *, remark: str) -> PackingRow:
    length, width, height = _dims_non_original(p.volume)
    gid = uuid.uuid4().hex[:8]
    return PackingRow(
        warehouse=p.warehouse,
        sku=p.sku,
        msku=p.msku,
        qty=p.qty,
        units_per_box=p.qty,
        box_count=1.0,
        box_weight_kg=p.weight,
        length_cm=length,
        width_cm=width,
        height_cm=height,
        remark=remark,
        box_group_id=gid,
    )


def _partials_to_mixed_rows(group: list[_Partial]) -> list[PackingRow]:
    """多 SKU 同一物理箱：多行共享 box_group_id / 箱规 / 总毛重。"""
    total_w = sum(p.weight for p in group)
    total_v = sum(p.volume for p in group)
    length, width, height = _dims_non_original(total_v)
    gid = uuid.uuid4().hex[:8]
    rows: list[PackingRow] = []
    for p in group:
        rows.append(
            PackingRow(
                warehouse=p.warehouse,
                sku=p.sku,
                msku=p.msku,
                qty=p.qty,
                units_per_box=p.qty,
                box_count=1.0,
                box_weight_kg=total_w,  # 同箱总重写在每行，便于结果表阅读
                length_cm=length,
                width_cm=width,
                height_cm=height,
                remark="多SKU合箱",
                box_group_id=gid,
            )
        )
    return rows


def _dims_non_original(volume_cm3: float) -> tuple[float, float, float]:
    """非原箱规尺寸：体积<10000→20³；否则长=宽=20 推高，越界再调长宽，夹在 15~60。"""
    if volume_cm3 <= 0:
        return DEFAULT_SIDE_CM, DEFAULT_SIDE_CM, DEFAULT_SIDE_CM
    if volume_cm3 < SMALL_VOLUME_CM3:
        return DEFAULT_SIDE_CM, DEFAULT_SIDE_CM, DEFAULT_SIDE_CM

    length = width = DEFAULT_SIDE_CM
    height = volume_cm3 / (length * width)
    if MIN_SIDE_CM <= height <= MAX_SIDE_CM:
        return _round1(length), _round1(width), _round1(height)

    if height > MAX_SIDE_CM:
        # 放大底面积使高=60
        area = volume_cm3 / MAX_SIDE_CM
        side = math.sqrt(max(area, MIN_SIDE_CM * MIN_SIDE_CM))
        side = _clamp_side(side)
        h2 = volume_cm3 / (side * side)
        return _round1(side), _round1(side), _round1(_clamp_side(h2))

    # height < MIN：略缩小底边（不低于 15）
    area = volume_cm3 / MIN_SIDE_CM
    side = math.sqrt(max(area, MIN_SIDE_CM * MIN_SIDE_CM))
    side = _clamp_side(side)
    h2 = volume_cm3 / (side * side)
    return _round1(side), _round1(side), _round1(_clamp_side(h2))


# --- 读表 ---


def _build_line_items(
    pack_rows: Sequence[Sequence[Any]],
    detail_rows: Sequence[Sequence[Any]],
    product_specs: Mapping[str, ProductSpec],
) -> list[LineItem]:
    pack_header = [_cell_str(c) for c in pack_rows[0]] if pack_rows else []
    detail_header = [_cell_str(c) for c in detail_rows[0]] if detail_rows else []
    pack_idx = _header_index(pack_header)
    detail_idx = _header_index(detail_header)

    detail_by_sku: dict[str, dict[str, Any]] = {}
    last_wh = ""
    last_sn = ""
    for raw in detail_rows[1:]:
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            continue
        sku = _cell_str(_get(raw, detail_idx, "SKU"))
        if not sku:
            continue
        wh = _cell_str(_get(raw, detail_idx, "发货仓库（单据）")) or _cell_str(
            _get(raw, detail_idx, "发货仓库（单据明细）")
        )
        sn = _cell_str(_get(raw, detail_idx, "发货单号"))
        if wh:
            last_wh = wh
        else:
            wh = last_wh
        if sn:
            last_sn = sn
        else:
            sn = last_sn
        detail_by_sku[sku] = {
            "warehouse": wh,
            "shipment_sn": sn,
            "msku": _cell_str(_get(raw, detail_idx, "MSKU")) or sku,
            "product_name": _cell_str(_get(raw, detail_idx, "品名")),
            "qty": _to_float(_get(raw, detail_idx, "发货量")),
        }

    items: list[LineItem] = []
    last_wh = ""
    last_sn = ""
    for raw in pack_rows[1:]:
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            continue
        sku = _cell_str(_get(raw, pack_idx, "SKU"))
        if not sku:
            continue
        wh = _cell_str(_get(raw, pack_idx, "发货仓库（单据）"))
        sn = _cell_str(_get(raw, pack_idx, "发货单号"))
        if wh:
            last_wh = wh
        else:
            wh = last_wh
        if sn:
            last_sn = sn
        else:
            sn = last_sn

        detail = detail_by_sku.get(sku, {})
        warehouse = wh or str(detail.get("warehouse") or "")
        shipment_sn = sn or str(detail.get("shipment_sn") or "")
        msku = str(detail.get("msku") or sku)
        product_name = _cell_str(_get(raw, pack_idx, "品名")) or str(detail.get("product_name") or "")

        qty = _to_float(_get(raw, pack_idx, "发货数量"))
        if qty is None:
            qty = _to_float(detail.get("qty"))
        units = _to_float(_get(raw, pack_idx, "单箱数量"))
        box_w = _to_float(_get(raw, pack_idx, "箱子毛重（kg）"))
        length = _to_float(_get(raw, pack_idx, "箱子长度（cm）"))
        width = _to_float(_get(raw, pack_idx, "箱子宽度（cm）"))
        height = _to_float(_get(raw, pack_idx, "箱子高度（cm）"))

        spec = product_specs.get(sku) or product_specs.get(msku)
        if spec:
            units = spec.units_per_box if spec.units_per_box is not None else units
            box_w = spec.box_weight_kg if spec.box_weight_kg is not None else box_w
            length = spec.length_cm if spec.length_cm is not None else length
            width = spec.width_cm if spec.width_cm is not None else width
            height = spec.height_cm if spec.height_cm is not None else height

        if qty is None or units is None or units <= 0:
            continue
        items.append(
            LineItem(
                shipment_sn=shipment_sn,
                warehouse=warehouse or "未知仓库",
                sku=sku,
                msku=msku,
                product_name=product_name,
                qty=float(qty),
                units_per_box=float(units),
                box_weight_kg=float(box_w or 0.0),
                length_cm=float(length or 0.0),
                width_cm=float(width or 0.0),
                height_cm=float(height or 0.0),
            )
        )
    return items


def _require_sheet(wb: Any, name: str) -> Any:
    if name not in wb.sheetnames:
        raise ValueError(f"发货单缺少工作表：{name}（现有：{', '.join(wb.sheetnames)}）")
    return wb[name]


def _header_index(headers: Sequence[str]) -> dict[str, int]:
    return {h: i for i, h in enumerate(headers) if h}


def _get(row: Sequence[Any], idx: Mapping[str, int], key: str) -> Any:
    i = idx.get(key)
    if i is None or i >= len(row):
        return None
    return row[i]


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _is_integer(value: float, tol: float = 1e-9) -> bool:
    return abs(value - round(value)) <= tol


def _clamp_side(value: float) -> float:
    return max(MIN_SIDE_CM, min(MAX_SIDE_CM, float(value)))


def _round1(value: float) -> float:
    return round(float(value), 1)


def _round2(value: float) -> float:
    return round(float(value), 2)


def _num(value: float) -> float | int:
    if _is_integer(value):
        return int(round(value))
    return _round1(value)
