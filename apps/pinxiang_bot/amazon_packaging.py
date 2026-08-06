#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""把拼箱结果写入亚马逊「包装箱包装信息」表（运营上传的 STA 装箱表）。

国家：美国 → cm/kg 换 in/lb；加拿大 → 公制直填；其它默认美国。
"""

from __future__ import annotations

import shutil
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence, Union

from openpyxl import load_workbook

from amazon_export import _use_imperial  # noqa: E402
from packing import PackingRow

PathLike = Union[str, Path]

CM_TO_INCH = 0.393701
KG_TO_LB = 2.20462
SHEET_NAME = "包装箱包装信息"
START_COL = 13
SKU_HEADER_ROW = 5
SKU_DATA_START_ROW = 6


@dataclass
class BoxSkuLine:
    keys: set[str]
    units: int


@dataclass
class BoxPlan:
    """一个物理包装箱，可含多 SKU（合箱）。"""

    lines: list[BoxSkuLine] = field(default_factory=list)
    # 字段名历史遗留；加拿大时存公制 cm/kg
    weight_lb: float = 0.0
    width_in: float = 0.0
    length_in: float = 0.0
    height_in: float = 0.0


def expand_packing_rows_to_boxes(
    rows: Sequence[PackingRow],
    *,
    imperial: bool = True,
) -> list[BoxPlan]:
    """拼箱行 → 物理箱列表。同 box_group_id 合并为一箱；原箱-整数部分按箱数展开。"""
    boxes: list[BoxPlan] = []
    grouped: "OrderedDict[str, list[PackingRow]]" = OrderedDict()
    solo_order: list[PackingRow] = []

    for row in rows:
        gid = (row.box_group_id or "").strip()
        if gid:
            grouped.setdefault(gid, []).append(row)
        else:
            solo_order.append(row)

    # 先输出无 group 的（整数原箱多箱展开）
    for row in solo_order:
        n_boxes = int(round(row.box_count)) if row.box_count else 1
        n_boxes = max(n_boxes, 1)
        units = int(round(row.units_per_box))
        keys = {str(row.sku).strip(), str(row.msku or "").strip()}
        keys.discard("")
        weight, width, length, height = _box_dims(
            row.box_weight_kg, row.width_cm, row.length_cm, row.height_cm, imperial=imperial
        )
        for _ in range(n_boxes):
            boxes.append(
                BoxPlan(
                    lines=[BoxSkuLine(keys=set(keys), units=units)],
                    weight_lb=weight,
                    width_in=width,
                    length_in=length,
                    height_in=height,
                )
            )

    # 合箱：同 group 一物理箱
    for _gid, members in grouped.items():
        lines: list[BoxSkuLine] = []
        for row in members:
            keys = {str(row.sku).strip(), str(row.msku or "").strip()}
            keys.discard("")
            units = int(round(row.qty if row.box_count == 1 else row.units_per_box))
            lines.append(BoxSkuLine(keys=keys, units=max(units, 0)))
        # 同箱总重/尺寸取第一行（合箱时各行已写相同值）
        head = members[0]
        weight, width, length, height = _box_dims(
            head.box_weight_kg,
            head.width_cm,
            head.length_cm,
            head.height_cm,
            imperial=imperial,
        )
        boxes.append(
            BoxPlan(
                lines=lines,
                weight_lb=weight,
                width_in=width,
                length_in=length,
                height_in=height,
            )
        )
    return boxes


def fill_amazon_packaging_file(
    amazon_path: PathLike,
    packing_rows: Sequence[PackingRow],
    output_path: PathLike,
    *,
    country: str = "",
) -> Path:
    src = Path(amazon_path)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    if src.resolve() != out.resolve():
        shutil.copy2(src, out)

    imperial = _use_imperial(country)
    boxes = expand_packing_rows_to_boxes(packing_rows, imperial=imperial)
    if not boxes:
        raise ValueError("拼箱结果为空，无法填写包装箱包装信息")

    wb = load_workbook(out)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"文件缺少工作表「{SHEET_NAME}」，请上传亚马逊装箱信息表")
    ws = wb[SHEET_NAME]

    total = len(boxes)
    ws.cell(3, START_COL, total)

    for idx in range(total):
        col = START_COL + idx
        ws.cell(SKU_HEADER_ROW, col, f"包装箱 {idx + 1} 数量")

    name_row = _find_row_by_label(ws, "包装箱名称")
    weight_row = _find_row_by_label(ws, "包装箱重量")
    width_row = _find_row_by_label(ws, "包装箱宽度")
    length_row = _find_row_by_label(ws, "包装箱长度")
    height_row = _find_row_by_label(ws, "包装箱高度")
    if name_row is None:
        raise ValueError("未找到「包装箱名称」行")

    for idx in range(total):
        col = START_COL + idx
        ws.cell(name_row, col, f"P1 - B{idx + 1}")
        box = boxes[idx]
        if weight_row:
            ws.cell(weight_row, col, box.weight_lb)
        if width_row:
            ws.cell(width_row, col, box.width_in)
        if length_row:
            ws.cell(length_row, col, box.length_in)
        if height_row:
            ws.cell(height_row, col, box.height_in)

    sku_end = name_row - 1 if name_row else ws.max_row
    data_rows: list[tuple[int, str]] = []
    for r in range(SKU_DATA_START_ROW, sku_end + 1):
        sku = ws.cell(r, 1).value
        if sku is None or str(sku).strip() == "":
            break
        if str(sku).strip() in {"包装箱名称", "STOP"}:
            break
        data_rows.append((r, str(sku).strip()))

    for r, _ in data_rows:
        for idx in range(max(total, 20)):
            ws.cell(r, START_COL + idx, None)

    for r, amazon_sku in data_rows:
        packed = 0
        for idx, box in enumerate(boxes):
            units = _units_for_sku(amazon_sku, box)
            if units:
                ws.cell(r, START_COL + idx, units)
                packed += units
        expected = ws.cell(r, 10).value
        ws.cell(r, 11, packed if packed else expected)

    wb.save(out)
    return out


def is_amazon_packaging_workbook(path: PathLike) -> bool:
    try:
        wb = load_workbook(path, read_only=True)
        ok = SHEET_NAME in wb.sheetnames
        wb.close()
        return ok
    except Exception:
        return False


def _box_dims(
    weight_kg: float,
    width_cm: float,
    length_cm: float,
    height_cm: float,
    *,
    imperial: bool,
) -> tuple[float, float, float, float]:
    """返回 (weight, width, length, height)；imperial=True 时为 lb/in。统一两位小数。"""
    if imperial:
        return (
            round(float(weight_kg) * KG_TO_LB, 2),
            round(float(width_cm) * CM_TO_INCH, 2),
            round(float(length_cm) * CM_TO_INCH, 2),
            round(float(height_cm) * CM_TO_INCH, 2),
        )
    return (
        round(float(weight_kg), 2),
        round(float(width_cm), 2),
        round(float(length_cm), 2),
        round(float(height_cm), 2),
    )


def _units_for_sku(amazon_sku: str, box: BoxPlan) -> int:
    for line in box.lines:
        if _sku_matches(amazon_sku, line.keys):
            return line.units
    return 0


def _find_row_by_label(ws, prefix: str) -> int | None:
    for r in range(1, min((ws.max_row or 1) + 1, 80)):
        val = ws.cell(r, 1).value
        if val is not None and str(val).strip().startswith(prefix):
            return r
    return None


def _sku_matches(amazon_sku: str, keys: set[str]) -> bool:
    a = amazon_sku.strip()
    if a in keys:
        return True
    for k in keys:
        if not k:
            continue
        if a == k or a.startswith(k) or k.startswith(a):
            return True
    return False
