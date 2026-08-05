from __future__ import annotations

import asyncio
import json
import logging
import sys
import threading
import time
import zipfile
from dataclasses import dataclass
from functools import partial
from io import BytesIO
from pathlib import Path
from typing import Optional, Tuple
from urllib.parse import urlparse

import dingtalk_stream
import requests
from dingtalk_stream import AckMessage

APP_DIR = Path(__file__).resolve().parent
ROOT_DIR = APP_DIR.parent.parent
SPLIT_DIR = ROOT_DIR / "apps" / "split_bot"
# pinxiang app dir first so packing/product_info resolve here; never import bare "config"
for path in (str(SPLIT_DIR), str(ROOT_DIR)):
    if path not in sys.path:
        sys.path.insert(0, path)
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

import pinxiang_config  # noqa: E402
from amazon_export import create_amazon_workbook  # noqa: E402
from amazon_packaging import fill_amazon_packaging_file, is_amazon_packaging_workbook  # noqa: E402
from packing import (  # noqa: E402
    load_packing_result_workbook,
    process_shipment_file,
    write_packing_workbook,
)
from product_info_source import load_product_specs  # noqa: E402
from runtime import (  # noqa: E402
    MessageDeduplicator,
    collect_download_codes,
    collect_file_names_by_download_code,
)
from Utils.dingtalk_api import (  # noqa: E402
    get_token,
    send_robot_private_file_message,
    send_robot_private_text_message,
)

# survive process restart / deploy recreate (workspace is bind-mounted)
_PENDING_STATE_NAME = "pending_state.json"


class MessageFormatError(ValueError):
    pass


@dataclass
class DownloadedFile:
    path: Path
    file_name: str
    content_type: str


@dataclass
class PendingLogistics:
    packing_result_path: Path
    merge_dir: Path
    shipment_path: Path
    stage: str  # confirm | select_ops
    updated_at: float
    logistics_channel: str = ""
    store_name: str = ""
    country: str = ""
    amazon_template_path: Optional[Path] = None


@dataclass
class PendingOpsJob:
    logistics_user_id: str
    logistics_name_hint: str
    packing_result_path: Path
    merge_dir: Path
    shipment_path: Path
    updated_at: float
    logistics_channel: str = ""
    store_name: str = ""
    country: str = ""
    amazon_template_path: Optional[Path] = None


class PinxiangBotHandler(dingtalk_stream.ChatbotHandler):
    """不分仓拼箱：发货单→拼箱结果→物流选运营转发→运营上传装箱表→回填。"""

    def __init__(self, logger: Optional[logging.Logger] = None, config: Optional[object] = None):
        super().__init__()
        self.logger = logger or logging.getLogger(__name__)
        self.config = config
        self.workspace = Path(
            getattr(config, "pinxiang_workspace", None)
            or getattr(config, "workspace", None)
            or pinxiang_config.WORKSPACE
        )
        self.workspace.mkdir(parents=True, exist_ok=True)
        self._deduplicator = MessageDeduplicator()
        self._job_semaphore = asyncio.Semaphore(1)
        self._pending_lock = threading.Lock()
        self._pending_state_path = self.workspace / _PENDING_STATE_NAME
        self._pending_ttl_seconds = int(
            getattr(config, "pinxiang_pending_ttl_sec", pinxiang_config.PENDING_TTL_SEC)
        )
        self._pending_logistics: dict[str, PendingLogistics] = {}
        self._pending_ops: dict[str, PendingOpsJob] = {}
        # 运营队列：忙碌时入队，做完一单自动推下一单；物流转发后立即释放
        self._ops_queue: dict[str, list[PendingOpsJob]] = {}
        self._load_pending_state()
        self.ops_users = list(pinxiang_config.OPS_USERS)

    async def process(self, callback: dingtalk_stream.CallbackMessage) -> Tuple[str, str]:
        incoming_message = dingtalk_stream.ChatbotMessage.from_dict(callback.data)
        user_id = str(incoming_message.sender_staff_id or "")
        message_key = incoming_message.message_id or f"{user_id}:{incoming_message.create_at}"
        self.logger.info("pinxiang received message: user_id=%s message_id=%s", user_id, incoming_message.message_id)
        if self._deduplicator.seen(message_key):
            self.logger.info("pinxiang skip duplicate: %s", message_key)
            return AckMessage.STATUS_OK, "OK"
        try:
            async with self._job_semaphore:
                await self._handle_message(incoming_message, user_id, callback.data)
        except (FileNotFoundError, MessageFormatError, ValueError) as exc:
            await self._send_text(user_id, str(exc))
        except Exception as exc:
            self.logger.exception("pinxiang task failed")
            await self._send_text(user_id, f"不分仓拼箱失败：{exc}")
        return AckMessage.STATUS_OK, "OK"

    async def _handle_message(self, incoming_message, user_id: str, raw_payload: dict) -> None:
        message_text = self._extract_message_text(incoming_message, raw_payload)
        self._cleanup_pending()
        has_files = bool(collect_download_codes(raw_payload))

        if has_files and user_id in self._pending_ops:
            await self._handle_ops_amazon_upload(incoming_message, user_id, raw_payload)
            return

        if not has_files:
            if await self._handle_text_commands(user_id, message_text):
                return
            if user_id in self._pending_ops:
                raise MessageFormatError(
                    "请上传亚马逊「包装箱包装信息」Excel（.xlsx）。\n回复【取消】可放弃当前任务。"
                )
            raise MessageFormatError(
                "请上传发货单 Excel（.xlsx）。\n"
                "拼箱结果出来后：回复【确认】并选择运营转发；运营上传装箱表后机器人填写回传。"
            )

        download_codes = collect_download_codes(raw_payload)
        file_names = collect_file_names_by_download_code(raw_payload)
        downloaded = await asyncio.get_running_loop().run_in_executor(
            None,
            self._download_excel,
            download_codes,
            incoming_message.message_id,
            file_names,
        )
        if downloaded is None:
            raise MessageFormatError("未识别到 Excel（.xlsx），请重试。")

        if is_amazon_packaging_workbook(downloaded.path):
            if user_id in self._pending_ops:
                await self._fill_and_reply_amazon(user_id, downloaded)
                return
            raise MessageFormatError(
                "这是亚马逊装箱表。请先由物流完成拼箱并转发任务后，再由运营上传此文件。"
            )

        pending = self._pending_logistics.get(user_id)
        if pending is not None and pending.stage in {"confirm", "select_ops"}:
            if self._looks_like_packing_result(downloaded.path):
                await self._handle_modified_packing_result(user_id, downloaded, pending)
                return

        await self._handle_shipment_upload(user_id, downloaded, incoming_message)

    async def _handle_shipment_upload(self, user_id: str, shipment: DownloadedFile, incoming_message) -> None:
        await self._send_text(user_id, "已收到发货单，正在计算拼箱，请稍等…")
        job_dir = self.workspace / str(incoming_message.message_id or time.time_ns())
        job_dir.mkdir(parents=True, exist_ok=True)

        product_specs = await asyncio.get_running_loop().run_in_executor(
            None, self._load_product_specs_safe
        )
        result = await asyncio.get_running_loop().run_in_executor(
            None,
            partial(process_shipment_file, shipment.path, product_specs=product_specs),
        )
        packing_name = f"{result.result_basename(fallback=Path(shipment.file_name).stem)}.xlsx"
        packing_path = job_dir / packing_name
        await asyncio.get_running_loop().run_in_executor(
            None, partial(write_packing_workbook, result, packing_path)
        )

        amazon_template_path = await asyncio.get_running_loop().run_in_executor(
            None,
            partial(self._build_amazon_template, job_dir, result, pinxiang_config.AMAZON_TEMPLATE_MPL),
        )

        shipment_copy = job_dir / shipment.file_name
        if shipment.path.resolve() != shipment_copy.resolve():
            shipment_copy.write_bytes(shipment.path.read_bytes())
        else:
            shipment_copy = shipment.path

        self._set_pending_logistics(
            user_id,
            PendingLogistics(
                packing_result_path=packing_path,
                merge_dir=job_dir,
                shipment_path=shipment_copy,
                stage="confirm",
                updated_at=time.time(),
                logistics_channel=result.logistics_channel or "",
                store_name=result.store_name or "",
                country=result.country or "",
                amazon_template_path=amazon_template_path,
            ),
        )

        await self._send_file(
            user_id,
            packing_name,
            packing_path.read_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        warn_text = ""
        if result.warnings:
            warn_text = "\n⚠️ " + "；".join(result.warnings[:5])
        if not result.rows:
            warn_text += "\n（本票无非整箱行，拼箱结果表为空。）"
        await self._send_text(
            user_id,
            "拼箱结果已发送。\n"
            f"物流渠道：{result.logistics_channel or ''}\n"
            f"店铺：{result.store_name or ''}\n"
            f"国家：{result.country or ''}"
            f"{warn_text}\n\n"
            "回复【确认】➡️ 选择运营并转发\n"
            "上传修正版拼箱结果Excel ➡️ 替换当前结果并重新生成Amazon模板\n"
            "回复【取消】➡️ 放弃本次任务",
        )

    async def _handle_modified_packing_result(
        self, user_id: str, uploaded: DownloadedFile, pending: PendingLogistics
    ) -> None:
        await self._send_text(user_id, "已收到修正版拼箱结果，正在校验并重新生成 Amazon 模板…")
        result = await asyncio.get_running_loop().run_in_executor(
            None, partial(load_packing_result_workbook, uploaded.path)
        )
        # 保留发货单里的渠道/店铺/国家与单号（修正表通常不含）
        result.logistics_channel = pending.logistics_channel or result.logistics_channel
        result.store_name = pending.store_name or result.store_name
        result.country = pending.country or result.country
        if not result.shipment_sns and pending.packing_result_path.is_file():
            # 尽量从旧文件名还原单号展示
            stem = pending.packing_result_path.stem.replace(" 拼箱数据", "").strip()
            if stem:
                result.shipment_sns = [p for p in stem.split() if p.startswith("SP")]

        packing_name = f"{result.result_basename(fallback=Path(uploaded.file_name).stem)}.xlsx"
        packing_path = pending.merge_dir / packing_name
        await asyncio.get_running_loop().run_in_executor(
            None, partial(write_packing_workbook, result, packing_path)
        )
        amazon_template_path = await asyncio.get_running_loop().run_in_executor(
            None,
            partial(self._build_amazon_template, pending.merge_dir, result, pinxiang_config.AMAZON_TEMPLATE_MPL),
        )

        pending.packing_result_path = packing_path
        pending.amazon_template_path = amazon_template_path
        pending.stage = "confirm"
        pending.updated_at = time.time()
        self._persist_pending_state()

        await self._send_file(
            user_id,
            packing_name,
            packing_path.read_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        if amazon_template_path and Path(amazon_template_path).is_file():
            tpl = Path(amazon_template_path)
            await self._send_file(
                user_id,
                tpl.name,
                tpl.read_bytes(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        await self._send_text(
            user_id,
            "✅ 已使用您上传的修正版拼箱结果，并重新生成 Amazon 模板。\n"
            f"物流渠道：{pending.logistics_channel or ''}\n"
            f"店铺：{pending.store_name or ''}\n"
            f"国家：{pending.country or ''}\n\n"
            "请检查后回复【确认】选择运营并转发，或继续上传新的修正版。\n"
            "回复【取消】➡️ 放弃本次任务",
        )

    async def _handle_text_commands(self, user_id: str, message_text: str) -> bool:
        normalized = _normalize(message_text)
        if not normalized:
            return False

        if normalized in {"取消", "cancel", "取消拼箱"}:
            # 运营取消：只清自己当前单/队列；物流取消：只清物流待办，不动运营
            if user_id in self._pending_ops or self._ops_queue.get(user_id):
                self._drop_ops_pending_only(user_id)
                await self._send_text(user_id, "已取消您当前的运营任务（含队列）。")
            else:
                self._drop_logistics_pending_only(user_id)
                await self._send_text(user_id, "已取消本次拼箱任务（不影响运营侧进行中的单）。")
            return True

        if normalized in {"模板2", "模板 2", "生成模板2", "重新生成模板2", "template2", "template 2"}:
            await self._handle_template_v2(user_id)
            return True

        pending = self._pending_logistics.get(user_id)

        if pending and pending.stage == "select_ops":
            ops = self._match_ops_choice(normalized)
            if ops is None:
                await self._send_text(user_id, self._ops_menu_text(prefix="请选择运营人员：\n"))
                return True
            await self._forward_to_ops(user_id, pending, ops)
            return True

        if normalized in {"确认", "confirm", "确认拼箱"}:
            if pending is None:
                raise MessageFormatError("当前没有待确认的拼箱任务，请先上传发货单。")
            if not pending.packing_result_path.is_file():
                raise MessageFormatError("拼箱结果文件丢失，请重新上传发货单。")
            pending.stage = "select_ops"
            pending.updated_at = time.time()
            self._persist_pending_state()
            await self._send_text(
                user_id, self._ops_menu_text(prefix="已确认拼箱结果。\n请选择要转发的运营：\n")
            )
            return True

        return False

    async def _ensure_template_for_pending(self, pending: PendingLogistics) -> Path:
        template_path = pending.amazon_template_path
        if template_path is not None and Path(template_path).is_file():
            return Path(template_path)
        packing_result = await asyncio.get_running_loop().run_in_executor(
            None,
            partial(
                self._load_packing_result_for_template,
                pending.packing_result_path,
                pending.shipment_path,
            ),
        )
        template_path = await asyncio.get_running_loop().run_in_executor(
            None,
            partial(
                self._build_amazon_template,
                pending.merge_dir,
                packing_result,
                pinxiang_config.AMAZON_TEMPLATE_MPL,
            ),
        )
        pending.amazon_template_path = template_path
        return Path(template_path)

    def _job_from_logistics(
        self, logistics_user_id: str, pending: PendingLogistics, template_path: Path
    ) -> PendingOpsJob:
        return PendingOpsJob(
            logistics_user_id=logistics_user_id,
            logistics_name_hint=logistics_user_id,
            packing_result_path=pending.packing_result_path,
            merge_dir=pending.merge_dir,
            shipment_path=pending.shipment_path,
            updated_at=time.time(),
            logistics_channel=pending.logistics_channel or "",
            store_name=pending.store_name or "",
            country=pending.country or "",
            amazon_template_path=template_path if template_path else None,
        )

    async def _push_ops_job_files(self, ops_id: str, ops_name: str, job: PendingOpsJob) -> None:
        packing_path = job.packing_result_path
        packing_name = packing_path.name
        await self._send_text(
            ops_id,
            "【不分仓拼箱】物流已审核通过，请处理。\n"
            "请下载拼箱结果与 Amazon 模板；已默认发送模板1，如需新版格式回复【模板2】。\n"
            "也可上传卖家后台「包装箱包装信息」表，机器人将自动填写并回传。\n"
            f"物流渠道：{job.logistics_channel or ''}\n"
            f"店铺：{job.store_name or ''}\n"
            f"国家：{job.country or ''}\n"
            f"拼箱结果：{packing_name}",
        )
        await self._send_file(
            ops_id,
            packing_name,
            packing_path.read_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        if job.amazon_template_path and Path(job.amazon_template_path).is_file():
            tpl = Path(job.amazon_template_path)
            await self._send_file(
                ops_id,
                tpl.name,
                tpl.read_bytes(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    async def _activate_ops_job(self, ops_id: str, ops_name: str, job: PendingOpsJob) -> None:
        """设为运营当前单并推送文件。"""
        job.updated_at = time.time()
        await self._push_ops_job_files(ops_id, ops_name, job)
        with self._pending_lock:
            self._pending_ops[ops_id] = job
            self._save_pending_state_unlocked()

    async def _forward_to_ops(self, logistics_user_id: str, pending: PendingLogistics, ops: dict) -> None:
        self._cleanup_pending()
        ops_id = ops["user_id"]
        ops_name = ops["name"]
        packing_path = pending.packing_result_path
        packing_name = packing_path.name

        template_path = await self._ensure_template_for_pending(pending)
        job = self._job_from_logistics(logistics_user_id, pending, template_path)

        existing = self._pending_ops.get(ops_id)
        if existing is not None:
            # 运营忙碌：入队，物流立即释放可接下一单
            with self._pending_lock:
                q = self._ops_queue.setdefault(ops_id, [])
                q.append(job)
                queue_pos = len(q)
                self._pending_logistics.pop(logistics_user_id, None)
                self._save_pending_state_unlocked()
            busy_label = (
                existing.packing_result_path.name
                if existing.packing_result_path
                else "进行中任务"
            )
            await self._send_text(
                logistics_user_id,
                f"已转发给运营【{ops_name}】（排队中）。\n"
                f"对方正在处理：{busy_label}\n"
                f"本单排队第 {queue_pos} 位，完成后会自动推送给运营。\n"
                f"拼箱结果：{packing_name}\n\n"
                "您可继续处理下一单。",
            )
            await self._send_text(
                ops_id,
                f"【不分仓拼箱】新任务已加入队列（第 {queue_pos} 位）。\n"
                f"文件：{packing_name}\n"
                "请先完成当前单，完成后系统会自动推送下一单。",
            )
            self.logger.info(
                "pinxiang queued for ops=%s(%s) pos=%s packing=%s from logistics=%s",
                ops_name,
                ops_id,
                queue_pos,
                packing_name,
                logistics_user_id,
            )
            return

        await self._activate_ops_job(ops_id, ops_name, job)
        with self._pending_lock:
            self._pending_logistics.pop(logistics_user_id, None)
            self._save_pending_state_unlocked()

        await self._send_text(
            logistics_user_id,
            f"已转发给运营【{ops_name}】。\n"
            "已发送拼箱结果与 Amazon 模板1；运营可回复【模板2】换新版。\n"
            "您可继续处理下一单。",
        )
        self.logger.info(
            "pinxiang forwarded to ops=%s(%s) from logistics=%s",
            ops_name,
            ops_id,
            logistics_user_id,
        )

    async def _promote_ops_queue(self, ops_user_id: str) -> None:
        """当前单结束后弹出队列下一单并推送。"""
        ops_name = next(
            (o["name"] for o in self.ops_users if o.get("user_id") == ops_user_id),
            ops_user_id,
        )
        next_job: Optional[PendingOpsJob] = None
        remaining = 0
        with self._pending_lock:
            q = self._ops_queue.get(ops_user_id) or []
            if q:
                next_job = q.pop(0)
                remaining = len(q)
                if not q:
                    self._ops_queue.pop(ops_user_id, None)
                else:
                    self._ops_queue[ops_user_id] = q
            self._save_pending_state_unlocked()
        if next_job is None:
            return
        if not next_job.packing_result_path.is_file():
            self.logger.warning(
                "pinxiang queue job missing file ops=%s path=%s",
                ops_user_id,
                next_job.packing_result_path,
            )
            await self._promote_ops_queue(ops_user_id)
            return
        await self._send_text(
            ops_user_id,
            f"上一单已完成。开始处理排队下一单"
            + (f"（队列仍剩 {remaining} 单）" if remaining else "")
            + "：",
        )
        await self._activate_ops_job(ops_user_id, ops_name, next_job)
        self.logger.info(
            "pinxiang promoted queue ops=%s remaining=%s packing=%s",
            ops_user_id,
            remaining,
            next_job.packing_result_path.name,
        )

    async def _handle_ops_amazon_upload(self, incoming_message, user_id: str, raw_payload: dict) -> None:
        download_codes = collect_download_codes(raw_payload)
        file_names = collect_file_names_by_download_code(raw_payload)
        downloaded = await asyncio.get_running_loop().run_in_executor(
            None,
            self._download_excel,
            download_codes,
            incoming_message.message_id,
            file_names,
        )
        if downloaded is None:
            raise MessageFormatError("未识别到 Excel 文件。")
        if not is_amazon_packaging_workbook(downloaded.path):
            raise MessageFormatError(
                "请上传含「包装箱包装信息」工作表的亚马逊装箱 Excel（不是发货单、也不是 Manifest 模版）。"
            )
        await self._fill_and_reply_amazon(user_id, downloaded)

    async def _fill_and_reply_amazon(self, ops_user_id: str, amazon_file: DownloadedFile) -> None:
        job = self._pending_ops.get(ops_user_id)
        if job is None:
            raise MessageFormatError("当前没有待处理的拼箱任务。")

        await self._send_text(ops_user_id, "已收到装箱表，正在按拼箱结果填写，请稍等…")

        rows = await asyncio.get_running_loop().run_in_executor(
            None, partial(self._load_rows_for_job, job.packing_result_path, job.shipment_path)
        )
        if not rows:
            raise MessageFormatError("拼箱结果无改箱行，无法填写装箱表。")

        # 回传文件名与运营上传的亚马逊表保持完全一致
        out_name = amazon_file.file_name or "包装箱包装信息.xlsx"
        if not out_name.lower().endswith((".xlsx", ".xlsm")):
            out_name = f"{out_name}.xlsx"
        out_path = job.merge_dir / f"filled-{out_name}"

        await asyncio.get_running_loop().run_in_executor(
            None,
            partial(fill_amazon_packaging_file, amazon_file.path, rows, out_path),
        )

        await self._send_file(
            ops_user_id,
            out_name,
            out_path.read_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        await self._send_text(ops_user_id, "装箱表已填写完成，请查收。")
        # 装箱表文件只回运营；物流仅收完成通知
        logistics_uid = (job.logistics_user_id or "").strip()
        packing_label = job.packing_result_path.name if job.packing_result_path else ""
        if logistics_uid:
            ops_name = next(
                (o["name"] for o in self.ops_users if o.get("user_id") == ops_user_id),
                ops_user_id,
            )
            lines = [
                "【不分仓拼箱】运营已完成装箱表填写。",
                f"运营：{ops_name}",
            ]
            if packing_label:
                lines.append(f"拼箱结果：{packing_label}")
            lines.append("装箱表已回传运营（未抄送文件）。")
            await self._send_text(logistics_uid, "\n".join(lines))
        with self._pending_lock:
            self._pending_ops.pop(ops_user_id, None)
            self._save_pending_state_unlocked()
        # 一单一单：完成后自动推送排队中的下一单
        await self._promote_ops_queue(ops_user_id)

    async def _handle_template_v2(self, user_id: str) -> None:
        """运营或物流：按当前拼箱结果用 MPL2 模板重新生成 Amazon 文件。"""
        self._cleanup_pending()
        pending_log = self._pending_logistics.get(user_id)
        pending_ops = self._pending_ops.get(user_id)
        if pending_log is not None:
            packing_path = pending_log.packing_result_path
            shipment_path = pending_log.shipment_path
            merge_dir = pending_log.merge_dir
            target = "logistics"
        elif pending_ops is not None:
            packing_path = pending_ops.packing_result_path
            shipment_path = pending_ops.shipment_path
            merge_dir = pending_ops.merge_dir
            target = "ops"
        else:
            raise MessageFormatError("当前没有进行中的拼箱任务，无法生成模板2。")

        await self._send_text(user_id, "正在按当前拼箱结果生成 Amazon 模板2，请稍等…")
        packing_result = await asyncio.get_running_loop().run_in_executor(
            None, partial(self._load_packing_result_for_template, packing_path, shipment_path)
        )
        template_path = await asyncio.get_running_loop().run_in_executor(
            None,
            partial(
                self._build_amazon_template,
                merge_dir,
                packing_result,
                pinxiang_config.AMAZON_TEMPLATE_MPL2,
            ),
        )
        if target == "logistics" and pending_log is not None:
            pending_log.amazon_template_path = template_path
            pending_log.updated_at = time.time()
            self._persist_pending_state()
        elif pending_ops is not None:
            pending_ops.amazon_template_path = template_path
            pending_ops.updated_at = time.time()
            self._persist_pending_state()

        await self._send_file(
            user_id,
            template_path.name,
            template_path.read_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        await self._send_text(user_id, f"模板2已生成并发送：{template_path.name}")

    def _build_amazon_template(self, job_dir: Path, result, template_source: Path) -> Path:
        source = Path(template_source)
        if not source.is_file():
            raise MessageFormatError(f"Amazon 模板不存在：{source}")
        out_path = Path(job_dir) / source.name
        create_amazon_workbook(
            template_source=source,
            output_path=out_path,
            result=result,
        )
        return out_path

    @staticmethod
    def _looks_like_packing_result(path: Path) -> bool:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            ok = "拼箱结果" in wb.sheetnames
            wb.close()
            return ok
        except Exception:
            return False

    def _load_rows_for_job(self, packing_path: Path, shipment_path: Path):
        """优先用当前拼箱结果表（含修正版），否则回退发货单重算。"""
        if packing_path and Path(packing_path).is_file() and self._looks_like_packing_result(packing_path):
            return load_packing_result_workbook(packing_path).rows
        if shipment_path and Path(shipment_path).is_file():
            return process_shipment_file(
                shipment_path, product_specs=self._load_product_specs_safe()
            ).rows
        return []

    def _load_packing_result_for_template(self, packing_path: Path, shipment_path: Path):
        if packing_path and Path(packing_path).is_file() and self._looks_like_packing_result(packing_path):
            return load_packing_result_workbook(packing_path)
        if shipment_path and Path(shipment_path).is_file():
            return process_shipment_file(
                shipment_path, product_specs=self._load_product_specs_safe()
            )
        raise MessageFormatError("拼箱结果/发货单文件丢失，无法生成模板。")

    def _ops_menu_text(self, *, prefix: str = "") -> str:
        lines = [prefix.rstrip(), ""]
        for i, ops in enumerate(self.ops_users, start=1):
            lines.append(f"{i}. {ops['name']}")
        lines.append("")
        lines.append("回复序号或姓名选择；回复【取消】放弃。")
        return "\n".join(lines).strip()

    def _match_ops_choice(self, normalized: str) -> Optional[dict]:
        for i, ops in enumerate(self.ops_users, start=1):
            if normalized == str(i) or normalized.startswith(f"{i}.") or normalized.startswith(f"{i}、"):
                return ops
            name = _normalize(ops["name"])
            if name and (normalized == name or name in normalized or normalized in name):
                return ops
            if normalized == ops["user_id"]:
                return ops
        return None

    def _download_excel(
        self,
        download_codes: list[str],
        message_id: str,
        file_names: Optional[dict[str, str]] = None,
    ) -> Optional[DownloadedFile]:
        job_dir = self.workspace / str(message_id) / "downloads"
        job_dir.mkdir(parents=True, exist_ok=True)
        file_names = file_names or {}
        for index, code in enumerate(download_codes, start=1):
            download_url = self.get_image_download_url(code)
            if not download_url:
                continue
            response = requests.get(download_url, timeout=120)
            response.raise_for_status()
            file_name = file_names.get(code) or _infer_filename(response.headers, f"file-{index}.xlsx")
            suffix = Path(file_name).suffix.lower()
            file_path = job_dir / file_name
            file_path.write_bytes(response.content)
            downloaded = DownloadedFile(file_path, file_name, response.headers.get("Content-Type", ""))
            if suffix in {".xlsx", ".xlsm"} or _looks_like_xlsx(response.content):
                return downloaded
            if suffix == ".xls":
                raise MessageFormatError("暂不支持 .xls，请另存为 .xlsx 后重试。")
        return None

    def _cleanup_pending(self) -> None:
        cutoff = time.time() - self._pending_ttl_seconds
        with self._pending_lock:
            before_l = len(self._pending_logistics)
            before_o = len(self._pending_ops)
            before_q = sum(len(v) for v in self._ops_queue.values())
            self._pending_logistics = {
                k: v for k, v in self._pending_logistics.items() if v.updated_at >= cutoff
            }
            self._pending_ops = {k: v for k, v in self._pending_ops.items() if v.updated_at >= cutoff}
            cleaned_q: dict[str, list[PendingOpsJob]] = {}
            for oid, jobs in self._ops_queue.items():
                kept = [
                    j
                    for j in jobs
                    if j.updated_at >= cutoff
                    and j.packing_result_path.is_file()
                    and j.shipment_path.is_file()
                ]
                if kept:
                    cleaned_q[oid] = kept
            self._ops_queue = cleaned_q
            after_q = sum(len(v) for v in self._ops_queue.values())
            if (
                len(self._pending_logistics) != before_l
                or len(self._pending_ops) != before_o
                or after_q != before_q
            ):
                self._save_pending_state_unlocked()

    def clear_user(self, user_id: str) -> None:
        """路由【重置】：只清该用户物流侧待办，绝不碰运营 active/queue。"""
        self._drop_logistics_pending_only(user_id)

    def has_pending(self, user_id: str) -> bool:
        if not user_id:
            return False
        self._cleanup_pending()
        return (
            user_id in self._pending_logistics
            or user_id in self._pending_ops
            or bool(self._ops_queue.get(user_id))
        )

    def _set_pending_logistics(self, user_id: str, pending: PendingLogistics) -> None:
        with self._pending_lock:
            self._pending_logistics[user_id] = pending
            self._save_pending_state_unlocked()

    def _drop_logistics_pending_only(self, user_id: str) -> None:
        if not user_id:
            return
        with self._pending_lock:
            self._pending_logistics.pop(user_id, None)
            self._save_pending_state_unlocked()

    def _drop_ops_pending_only(self, user_id: str) -> None:
        if not user_id:
            return
        with self._pending_lock:
            self._pending_ops.pop(user_id, None)
            self._ops_queue.pop(user_id, None)
            self._save_pending_state_unlocked()

    def _drop_user_pending(self, user_id: str) -> None:
        """兼容旧调用：按角色拆清（优先运营键，否则物流）。"""
        if not user_id:
            return
        if user_id in self._pending_ops or self._ops_queue.get(user_id):
            self._drop_ops_pending_only(user_id)
            return
        self._drop_logistics_pending_only(user_id)

    def _persist_pending_state(self) -> None:
        with self._pending_lock:
            self._save_pending_state_unlocked()

    def _pending_state_path_resolved(self) -> Path:
        return self._pending_state_path

    def _load_pending_state(self) -> None:
        path = self._pending_state_path_resolved()
        try:
            if not path.is_file():
                return
            raw = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(raw, dict):
                return
            cutoff = time.time() - self._pending_ttl_seconds
            logistics: dict[str, PendingLogistics] = {}
            for uid, item in (raw.get("logistics") or {}).items():
                if not uid or not isinstance(item, dict):
                    continue
                pending = self._logistics_from_dict(item)
                if pending is None or pending.updated_at < cutoff:
                    continue
                if not pending.packing_result_path.is_file() or not pending.shipment_path.is_file():
                    continue
                logistics[str(uid)] = pending
            ops: dict[str, PendingOpsJob] = {}
            for uid, item in (raw.get("ops") or {}).items():
                if not uid or not isinstance(item, dict):
                    continue
                job = self._ops_from_dict(item)
                if job is None or job.updated_at < cutoff:
                    continue
                if not job.packing_result_path.is_file() or not job.shipment_path.is_file():
                    continue
                ops[str(uid)] = job
            ops_queue: dict[str, list[PendingOpsJob]] = {}
            for uid, items in (raw.get("ops_queue") or {}).items():
                if not uid or not isinstance(items, list):
                    continue
                kept: list[PendingOpsJob] = []
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    job = self._ops_from_dict(item)
                    if job is None or job.updated_at < cutoff:
                        continue
                    if not job.packing_result_path.is_file() or not job.shipment_path.is_file():
                        continue
                    kept.append(job)
                if kept:
                    ops_queue[str(uid)] = kept
            self._pending_logistics = logistics
            self._pending_ops = ops
            self._ops_queue = ops_queue
            if logistics or ops or ops_queue:
                self.logger.info(
                    "pinxiang loaded pending state: logistics=%s ops=%s queue=%s",
                    len(logistics),
                    len(ops),
                    sum(len(v) for v in ops_queue.values()),
                )
        except Exception as exc:
            self.logger.warning("load pinxiang pending state failed: %s", exc)
            self._pending_logistics = {}
            self._pending_ops = {}
            self._ops_queue = {}

    def _save_pending_state_unlocked(self) -> None:
        path = self._pending_state_path_resolved()
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            payload = {
                "logistics": {
                    uid: {
                        "packing_result_path": str(p.packing_result_path),
                        "merge_dir": str(p.merge_dir),
                        "shipment_path": str(p.shipment_path),
                        "stage": p.stage,
                        "updated_at": p.updated_at,
                        "logistics_channel": p.logistics_channel,
                        "store_name": p.store_name,
                        "country": p.country,
                        "amazon_template_path": str(p.amazon_template_path)
                        if p.amazon_template_path
                        else "",
                    }
                    for uid, p in self._pending_logistics.items()
                },
                "ops": {
                    uid: {
                        "logistics_user_id": j.logistics_user_id,
                        "logistics_name_hint": j.logistics_name_hint,
                        "packing_result_path": str(j.packing_result_path),
                        "merge_dir": str(j.merge_dir),
                        "shipment_path": str(j.shipment_path),
                        "updated_at": j.updated_at,
                        "logistics_channel": j.logistics_channel,
                        "store_name": j.store_name,
                        "country": j.country,
                        "amazon_template_path": str(j.amazon_template_path)
                        if j.amazon_template_path
                        else "",
                    }
                    for uid, j in self._pending_ops.items()
                },
                "ops_queue": {
                    uid: [
                        {
                            "logistics_user_id": j.logistics_user_id,
                            "logistics_name_hint": j.logistics_name_hint,
                            "packing_result_path": str(j.packing_result_path),
                            "merge_dir": str(j.merge_dir),
                            "shipment_path": str(j.shipment_path),
                            "updated_at": j.updated_at,
                            "logistics_channel": j.logistics_channel,
                            "store_name": j.store_name,
                            "country": j.country,
                            "amazon_template_path": str(j.amazon_template_path)
                            if j.amazon_template_path
                            else "",
                        }
                        for j in jobs
                    ]
                    for uid, jobs in self._ops_queue.items()
                },
            }
            tmp = path.with_suffix(".tmp")
            tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=0), encoding="utf-8")
            tmp.replace(path)
        except Exception as exc:
            self.logger.warning("save pinxiang pending state failed: %s", exc)

    @staticmethod
    def _logistics_from_dict(item: dict) -> Optional[PendingLogistics]:
        try:
            stage = str(item.get("stage") or "confirm")
            if stage not in {"confirm", "select_ops"}:
                stage = "confirm"
            tpl = str(item.get("amazon_template_path") or "").strip()
            return PendingLogistics(
                packing_result_path=Path(str(item["packing_result_path"])),
                merge_dir=Path(str(item["merge_dir"])),
                shipment_path=Path(str(item["shipment_path"])),
                stage=stage,
                updated_at=float(item.get("updated_at") or 0),
                logistics_channel=str(item.get("logistics_channel") or ""),
                store_name=str(item.get("store_name") or ""),
                country=str(item.get("country") or ""),
                amazon_template_path=Path(tpl) if tpl else None,
            )
        except Exception:
            return None

    @staticmethod
    def _ops_from_dict(item: dict) -> Optional[PendingOpsJob]:
        try:
            tpl = str(item.get("amazon_template_path") or "").strip()
            return PendingOpsJob(
                logistics_user_id=str(item.get("logistics_user_id") or ""),
                logistics_name_hint=str(item.get("logistics_name_hint") or ""),
                packing_result_path=Path(str(item["packing_result_path"])),
                merge_dir=Path(str(item["merge_dir"])),
                shipment_path=Path(str(item["shipment_path"])),
                updated_at=float(item.get("updated_at") or 0),
                logistics_channel=str(item.get("logistics_channel") or ""),
                store_name=str(item.get("store_name") or ""),
                country=str(item.get("country") or ""),
                amazon_template_path=Path(tpl) if tpl else None,
            )
        except Exception:
            return None

    def _load_product_specs_safe(self) -> dict:
        path = (pinxiang_config.PRODUCT_INFO_PATH or "").strip()
        if not path:
            return {}
        if not path.lower().startswith("smb://") and not path.startswith("\\\\") and not path.startswith("//"):
            if not Path(path).is_file():
                self.logger.warning("product info file missing, fallback to shipment pack fields: %s", path)
                return {}
        try:
            return load_product_specs(
                path,
                smb_username=pinxiang_config.SMB_USERNAME,
                smb_password=pinxiang_config.SMB_PASSWORD,
                smb_port=pinxiang_config.SMB_PORT,
                smb_timeout_sec=pinxiang_config.SMB_TIMEOUT_SEC,
                smb_client_name=pinxiang_config.SMB_CLIENT_NAME,
            )
        except Exception as exc:
            self.logger.warning("load product info failed, fallback to shipment pack fields: %s", exc)
            return {}

    async def _send_text(self, user_id: str, message: str) -> None:
        if not user_id:
            self.logger.warning("pinxiang cannot send text: missing user id")
            return
        token = await get_token(self.config)
        if token:
            loop = asyncio.get_running_loop()
            await loop.run_in_executor(
                None, send_robot_private_text_message, token, self.config, [user_id], message
            )

    async def _send_file(
        self,
        user_id: str,
        file_name: str,
        content: bytes,
        *,
        content_type: str = "application/octet-stream",
    ) -> None:
        token = await get_token(self.config)
        if not token:
            raise RuntimeError("无法获取钉钉 access token")
        loop = asyncio.get_running_loop()
        media_id = await loop.run_in_executor(
            None,
            self.dingtalk_client.upload_to_dingtalk,
            content,
            "file",
            file_name,
            content_type,
        )
        if not media_id:
            raise RuntimeError("文件上传到钉钉失败")
        result = await loop.run_in_executor(
            None,
            send_robot_private_file_message,
            token,
            self.config,
            [user_id],
            media_id,
            file_name,
        )
        if not result:
            raise RuntimeError("文件消息发送失败")

    def _extract_message_text(self, incoming_message, raw_payload: dict) -> str:
        try:
            return "\n".join(self.extract_text_from_incoming_message(incoming_message) or []).strip()
        except Exception:
            pass
        content = raw_payload.get("text", {}).get("content")
        if isinstance(content, str):
            return content.strip()
        return ""


def _normalize(text: str) -> str:
    return (text or "").strip().lower().replace(" ", "").replace("\u3000", "")


def _infer_filename(headers, fallback_name: str) -> str:
    content_disposition = headers.get("Content-Disposition", "")
    if "filename=" in content_disposition:
        raw = content_disposition.split("filename=", 1)[1].strip().strip('"')
        return Path(urlparse(raw).path).name or fallback_name
    return fallback_name


def _looks_like_xlsx(content: bytes) -> bool:
    if not content.startswith(b"PK"):
        return False
    try:
        with zipfile.ZipFile(BytesIO(content)) as zf:
            names = set(zf.namelist())
    except zipfile.BadZipFile:
        return False
    return "[Content_Types].xml" in names and "xl/workbook.xml" in names
