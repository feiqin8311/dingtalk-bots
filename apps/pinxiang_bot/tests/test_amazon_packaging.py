#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""运营装箱表填写单测。"""

from __future__ import annotations

import shutil
import sys
import tempfile
import unittest
from pathlib import Path

APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from amazon_packaging import expand_packing_rows_to_boxes, fill_amazon_packaging_file  # noqa: E402
from amazon_export import CM_TO_INCH, KG_TO_LB  # noqa: E402
from packing import PackingRow  # noqa: E402

REF = Path("/Users/kerden/Desktop/不分仓拼箱/参考表格/EZARC 美国 AMP 2SKU拼箱数据.xlsx")


class AmazonPackagingTests(unittest.TestCase):
    def test_expand_boxes(self):
        rows = [
            PackingRow("WH", "80376351F1", "80376351F1", 180, 60, 3, 17.5, 54, 31, 18, "原箱-整数部分"),
            PackingRow("WH", "80376351F1", "80376351F1", 55, 55, 1, 17.0, 54, 30, 17, "余数改箱", box_group_id="r1"),
            PackingRow("WH", "801905N", "801905", 20, 20, 1, 9.0, 25, 34, 24, "不足一箱改箱", box_group_id="r2"),
        ]
        boxes = expand_packing_rows_to_boxes(rows)
        self.assertEqual(len(boxes), 5)
        self.assertEqual(boxes[0].lines[0].units, 60)
        self.assertEqual(boxes[3].lines[0].units, 55)
        self.assertEqual(boxes[4].lines[0].units, 20)
        self.assertIn("801905", boxes[4].lines[0].keys)
        # 默认美国：英制
        self.assertAlmostEqual(boxes[0].weight_lb, 17.5 * KG_TO_LB, places=5)
        self.assertAlmostEqual(boxes[0].length_in, 54 * CM_TO_INCH, places=5)

    def test_expand_boxes_canada_metric(self):
        rows = [
            PackingRow("WH", "A", "A", 100, 50, 2, 10.0, 40, 30, 20, "原箱"),
        ]
        boxes = expand_packing_rows_to_boxes(rows, imperial=False)
        self.assertEqual(len(boxes), 2)
        self.assertEqual(boxes[0].weight_lb, 10.0)
        self.assertEqual(boxes[0].length_in, 40.0)
        self.assertEqual(boxes[0].width_in, 30.0)
        self.assertEqual(boxes[0].height_in, 20.0)

    @unittest.skipUnless(REF.is_file(), "reference amazon packaging file missing")
    def test_fill_reference_shell(self):
        # 用参考文件当壳（已有 SKU 行），用标准拼箱行重填
        rows = [
            PackingRow("青山湖仓库", "80376351F1", "80376351F1", 180, 60, 3, 17.5, 54, 31, 18, "原箱-整数部分"),
            PackingRow("青山湖仓库", "80376351F1", "80376351F1", 55, 55, 1, 17.0, 54, 30, 17, "余数改箱"),
            PackingRow("青山湖仓库", "801905N", "801905", 20, 20, 1, 9.0, 25, 34, 24, "不足一箱改箱"),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "filled.xlsx"
            fill_amazon_packaging_file(REF, rows, out)
            from openpyxl import load_workbook

            wb = load_workbook(out, data_only=True)
            ws = wb["包装箱包装信息"]
            self.assertEqual(ws.cell(3, 13).value, 5)
            # 801905 row6: only last box 20 if order is F1*4 + 801905
            # Our expand order: F1 3 + F1 1 + 801905 → boxes 1-3=60, 4=55, 5=20
            self.assertEqual(ws.cell(6, 17).value, 20)  # 801905 in box 5
            self.assertEqual(ws.cell(7, 13).value, 60)
            self.assertEqual(ws.cell(7, 16).value, 55)
            wb.close()


if __name__ == "__main__":
    unittest.main()
