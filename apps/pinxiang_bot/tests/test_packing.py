#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""拼箱算法单测。"""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from packing import (  # noqa: E402
    LineItem,
    TEMPLATE1_SHEET,
    compute_packing,
    load_packing_result_workbook,
    process_shipment_file,
    write_packing_workbook,
)

SAMPLE_SHIPMENT = Path("/Users/kerden/Desktop/不分仓拼箱/参考表格/发货单-938438362569883648.xlsx")
TEST_SHIPMENT = Path("/Users/kerden/Desktop/测试数据/发货单-943520755484897280.xlsx")
TEST_EXPECTED = Path(
    "/Users/kerden/Desktop/测试数据/SP260805013 SP260805012 SP260805011 SP260805008 拼箱数据.xlsx"
)


def _item(**kwargs) -> LineItem:
    base = dict(
        shipment_sn="SP1",
        warehouse="青山湖仓库",
        sku="X",
        msku="X",
        product_name="",
        qty=1,
        units_per_box=10,
        box_weight_kg=10,
        length_cm=40,
        width_cm=30,
        height_cm=20,
    )
    base.update(kwargs)
    return LineItem(**base)


class PackingLogicTests(unittest.TestCase):
    def test_partial_b_solo_when_weight_ok(self):
        # 20/40, 单箱 17.5 → 单件 0.4375*20=8.75 ∈ [1,18] → B 单装
        result = compute_packing(
            [
                _item(
                    sku="801905N",
                    msku="801905",
                    qty=20,
                    units_per_box=40,
                    box_weight_kg=17.5,
                    length_cm=44,
                    width_cm=33.5,
                    height_cm=24,
                )
            ]
        )
        self.assertEqual(len(result.rows), 1)
        row = result.rows[0]
        self.assertEqual(row.qty, 20)
        self.assertEqual(row.units_per_box, 20)
        self.assertEqual(row.box_count, 1)
        self.assertAlmostEqual(row.box_weight_kg, 8.75, places=2)
        self.assertEqual(row.remark, "不足一箱改箱")

    def test_gt_one_prefer_all_in_one_when_weight_ok(self):
        # 47/40，整票约 17.7kg ∈ [1,18] → 优先拼 1 箱，不拆整数+余数
        result = compute_packing(
            [
                _item(
                    sku="801621",
                    qty=47,
                    units_per_box=40,
                    box_weight_kg=15.1,
                    length_cm=44,
                    width_cm=33.5,
                    height_cm=24,
                )
            ]
        )
        self.assertEqual(len(result.rows), 1)
        row = result.rows[0]
        self.assertEqual(row.remark, "整票拼1箱")
        self.assertEqual(row.qty, 47)
        self.assertEqual(row.units_per_box, 47)
        self.assertEqual(row.box_count, 1)
        self.assertAlmostEqual(row.box_weight_kg, 47 * 15.1 / 40, places=2)

    def test_non_integer_d_split(self):
        # 整票 235*17.5/60 ≈ 68.5kg >18 → 才走 D：整数原箱 + 余数
        result = compute_packing(
            [
                _item(
                    sku="80376351F1",
                    msku="80376351F1",
                    qty=235,
                    units_per_box=60,
                    box_weight_kg=17.5,
                    length_cm=54,
                    width_cm=31,
                    height_cm=18,
                )
            ]
        )
        by = {r.remark: r for r in result.rows}
        self.assertIn("原箱-整数部分", by)
        self.assertIn("余数改箱", by)
        self.assertEqual(by["原箱-整数部分"].qty, 180)
        self.assertEqual(by["原箱-整数部分"].box_count, 3)
        self.assertEqual(by["余数改箱"].qty, 55)

    def test_integer_excluded_from_sheet(self):
        result = compute_packing([_item(sku="A", qty=100, units_per_box=20, box_weight_kg=10)])
        self.assertEqual(len(result.rows), 0)
        self.assertEqual(len(result.all_rows), 1)
        self.assertEqual(result.all_rows[0].remark, "原箱")

    def test_multi_sku_combine_when_light(self):
        # 两 SKU 都极轻 → A 合箱
        result = compute_packing(
            [
                _item(sku="L1", msku="L1", qty=1, units_per_box=100, box_weight_kg=5, warehouse="W1"),
                _item(sku="L2", msku="L2", qty=1, units_per_box=100, box_weight_kg=5, warehouse="W1"),
            ]
        )
        # 单件重 0.05kg，合箱 0.1kg 仍 <1，但应合在一组
        self.assertEqual(len(result.rows), 2)
        self.assertTrue(all(r.remark == "多SKU合箱" for r in result.rows))
        self.assertEqual(result.rows[0].box_group_id, result.rows[1].box_group_id)
        self.assertAlmostEqual(result.rows[0].box_weight_kg, result.rows[1].box_weight_kg)

    def test_different_warehouse_not_combined(self):
        result = compute_packing(
            [
                _item(sku="L1", qty=1, units_per_box=100, box_weight_kg=5, warehouse="W1"),
                _item(sku="L2", qty=1, units_per_box=100, box_weight_kg=5, warehouse="W2"),
            ]
        )
        self.assertEqual(len(result.rows), 2)
        self.assertNotEqual(result.rows[0].box_group_id, result.rows[1].box_group_id)
        self.assertTrue(all(r.remark == "不足一箱改箱" for r in result.rows))

    def test_borrow_when_remainder_too_light(self):
        # 整票 201×0.1=20.1kg >18 → 不走 C；拆 2 原箱 + 余数 1 件(0.1kg) → 借箱
        result = compute_packing(
            [
                _item(
                    sku="B1",
                    qty=201,
                    units_per_box=100,
                    box_weight_kg=10,
                    length_cm=40,
                    width_cm=30,
                    height_cm=20,
                )
            ]
        )
        remarks = {r.remark for r in result.rows}
        self.assertIn("借箱合并", remarks)
        # 整数部分被借走 1 箱后剩 1 箱，不出现 0 箱行
        self.assertFalse(any(r.remark == "原箱-整数部分" and r.box_count < 1 for r in result.rows))
        by = {r.remark: r for r in result.rows}
        self.assertEqual(by["原箱-整数部分"].box_count, 1)
        self.assertEqual(by["借箱合并"].qty, 101)

    def test_small_volume_uses_20_cube(self):
        # 很小体积 → 20³
        result = compute_packing(
            [
                _item(
                    sku="S",
                    qty=1,
                    units_per_box=1000,
                    box_weight_kg=10,
                    length_cm=10,
                    width_cm=10,
                    height_cm=10,
                )
            ]
        )
        row = result.rows[0]
        # 单件体积 1，箱体积 1 < 10000
        self.assertEqual(row.length_cm, 20)
        self.assertEqual(row.width_cm, 20)
        self.assertEqual(row.height_cm, 20)

    def test_result_basename_joins_shipment_sns(self):
        result = compute_packing(
            [
                _item(shipment_sn="SP260722004", sku="A", qty=5, units_per_box=10, box_weight_kg=10),
                _item(shipment_sn="SP260722005", sku="B", qty=5, units_per_box=10, box_weight_kg=10),
                _item(shipment_sn="SP260722004", sku="C", qty=5, units_per_box=10, box_weight_kg=10),
                _item(shipment_sn="SP260722006", sku="D", qty=5, units_per_box=10, box_weight_kg=10),
            ]
        )
        self.assertEqual(
            result.shipment_sns,
            ["SP260722004", "SP260722005", "SP260722006"],
        )
        self.assertEqual(
            result.result_basename(),
            "SP260722004 SP260722005 SP260722006 拼箱数据",
        )

    def test_result_basename_fallback(self):
        result = compute_packing([])
        self.assertEqual(result.result_basename(fallback="发货单-xxx"), "发货单-xxx 拼箱数据")

    @unittest.skipUnless(SAMPLE_SHIPMENT.is_file(), "sample shipment not on this machine")
    def test_sample_shipment_file(self):
        result = process_shipment_file(SAMPLE_SHIPMENT)
        self.assertEqual(len(result.rows), 3)
        skus = {r.sku for r in result.rows}
        self.assertEqual(skus, {"801905N", "80376351F1"})
        with tempfile.TemporaryDirectory() as tmp:
            out = write_packing_workbook(result, Path(tmp) / "out.xlsx")
            self.assertTrue(out.is_file())

    def test_template1_sheet_full_before_partial(self):
        result = compute_packing(
            [
                _item(sku="FULL", qty=100, units_per_box=20, box_weight_kg=10),
                _item(sku="PART", qty=5, units_per_box=20, box_weight_kg=10),
            ]
        )
        from packing import _build_template1_rows

        items = [
            _item(sku="FULL", qty=100, units_per_box=20, box_weight_kg=10, length_cm=40, width_cm=30, height_cm=20),
            _item(sku="PART", qty=5, units_per_box=20, box_weight_kg=10, length_cm=40, width_cm=30, height_cm=20),
        ]
        t1 = _build_template1_rows(items, "美国")
        self.assertEqual([r.sku for r in t1], ["FULL", "PART"])
        self.assertTrue(t1[0].is_full_box)
        self.assertFalse(t1[1].is_full_box)
        self.assertEqual(t1[0].box_count, 5)
        self.assertIsNone(t1[1].units_per_box)

        result.country = "美国"
        result.template1_rows = t1
        with tempfile.TemporaryDirectory() as tmp:
            out = write_packing_workbook(result, Path(tmp) / "out.xlsx")
            loaded = load_packing_result_workbook(out)
            self.assertEqual(len(loaded.template1_rows), 2)
            self.assertEqual(loaded.template1_rows[0].sku, "FULL")
            self.assertTrue(loaded.template1_rows[0].is_full_box)
            self.assertEqual(loaded.template1_rows[1].sku, "PART")
            self.assertFalse(loaded.template1_rows[1].is_full_box)
            from openpyxl import load_workbook

            wb = load_workbook(out, data_only=True)
            self.assertIn(TEMPLATE1_SHEET, wb.sheetnames)
            ws = wb[TEMPLATE1_SHEET]
            self.assertEqual(ws.cell(2, 1).value, "是")
            self.assertEqual(ws.cell(3, 1).value, "否")
            self.assertIsNone(ws.cell(3, 5).value)
            wb.close()

    @unittest.skipUnless(
        TEST_SHIPMENT.is_file() and TEST_EXPECTED.is_file(),
        "desktop test fixtures not on this machine",
    )
    def test_template1_matches_desktop_sample(self):
        result = process_shipment_file(TEST_SHIPMENT)
        self.assertEqual(len(result.template1_rows), 33)
        full = [r for r in result.template1_rows if r.is_full_box]
        partial = [r for r in result.template1_rows if not r.is_full_box]
        self.assertEqual(len(full), 21)
        self.assertEqual(len(partial), 12)
        # 整箱在前、非整箱在后
        self.assertTrue(all(r.is_full_box for r in result.template1_rows[:21]))
        self.assertTrue(all(not r.is_full_box for r in result.template1_rows[21:]))
        from openpyxl import load_workbook

        exp = load_workbook(TEST_EXPECTED, data_only=True)[TEMPLATE1_SHEET]
        got_skus = [(r.is_full_box, r.sku, r.qty) for r in result.template1_rows]
        exp_skus = []
        for row in exp.iter_rows(min_row=2, values_only=True):
            if not row or not row[2]:
                continue
            exp_skus.append((row[0] == "是", str(row[2]), float(row[3])))
        self.assertEqual(got_skus, exp_skus)


if __name__ == "__main__":
    unittest.main()
