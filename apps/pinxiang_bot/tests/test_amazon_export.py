#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from amazon_export import (  # noqa: E402
    create_amazon_workbook,
    _manifest_rows,
    _is_mixed_packing_row,
    _use_imperial,
    CM_TO_INCH,
    KG_TO_LB,
)
from packing import PackingResult, PackingRow, Template1Row, process_shipment_file  # noqa: E402

TPL = APP_DIR / "templates" / (
    "ManifestFileUpload_Template_IncludeCasePack_IncludeExpirationDate_IncludeMLC_MPL.xlsx"
)
TPL2 = APP_DIR / "templates" / (
    "ManifestFileUpload_Template_IncludeCasePack_IncludeExpirationDate_IncludeMLC_MPL2.xlsx"
)
TEST_SHIPMENT = Path("/Users/kerden/Desktop/测试数据/发货单-943520755484897280.xlsx")
TEST_MANIFEST = Path(
    "/Users/kerden/Desktop/测试数据/"
    "ManifestFileUpload_Template_IncludeCasePack_IncludeExpirationDate_IncludeMLC_MPL(3).xlsx"
)


class AmazonExportTests(unittest.TestCase):
    def test_classify_original_vs_mixed(self):
        self.assertFalse(
            _is_mixed_packing_row(
                PackingRow("W", "A", "A", 288, 36, 8, 14.0, 49, 36, 27, "原箱-整数部分")
            )
        )
        self.assertTrue(
            _is_mixed_packing_row(
                PackingRow("W", "A", "A", 12, 12, 1, 4.8, 20, 20, 41, "余数改箱")
            )
        )

    def test_merge_mixed_sku_qty(self):
        rows = _manifest_rows(
            [
                PackingRow("W", "A", "A", 288, 36, 8, 14.0, 49, 36, 27, "原箱-整数部分"),
                PackingRow("W", "A", "A", 12, 12, 1, 4.8, 20, 20, 41, "余数改箱"),
                PackingRow("W", "A", "A", 6, 6, 1, 2.0, 20, 20, 20, "不足一箱改箱"),
                PackingRow("W", "B", "B", 20, 20, 1, 5.0, 20, 20, 20, "不足一箱改箱"),
            ]
        )
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[0][0].qty, 288)
        self.assertTrue(rows[0][1])
        self.assertEqual(rows[1][0].sku, "A")
        self.assertEqual(rows[1][0].qty, 18)
        self.assertFalse(rows[1][1])
        self.assertEqual(rows[2][0].sku, "B")
        self.assertFalse(rows[2][1])

    @unittest.skipUnless(TPL2.is_file(), "template missing")
    def test_write_workbook_fallback_amazon_rows(self):
        result = PackingResult(
            rows=[],
            all_rows=[
                PackingRow("W", "A", "A", 100, 50, 2, 10.0, 40, 30, 20, "原箱"),
                PackingRow("W", "B", "B", 10, 10, 1, 3.0, 20, 20, 20, "不足一箱改箱"),
                PackingRow("W", "B", "B", 5, 5, 1, 1.5, 20, 20, 20, "不足一箱改箱"),
            ],
        )
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "m.xlsx"
            create_amazon_workbook(template_source=TPL2, output_path=out, result=result)
            from openpyxl import load_workbook

            wb = load_workbook(out, data_only=True)
            ws = wb["Create workflow – template"]
            header = next(
                r for r in range(1, 15) if ws.cell(r, 1).value == "Merchant SKU"
            )
            # original A with dims
            self.assertEqual(ws.cell(header + 1, 1).value, "A")
            self.assertEqual(ws.cell(header + 1, 2).value, 100)
            self.assertEqual(ws.cell(header + 1, 5).value, 50)
            self.assertIsNotNone(ws.cell(header + 1, 7).value)
            # merged B no dims
            self.assertEqual(ws.cell(header + 2, 1).value, "B")
            self.assertEqual(ws.cell(header + 2, 2).value, 15)

    def test_use_imperial_by_country(self):
        self.assertTrue(_use_imperial("美国"))
        self.assertTrue(_use_imperial("US"))
        self.assertTrue(_use_imperial(""))
        self.assertFalse(_use_imperial("加拿大"))
        self.assertFalse(_use_imperial("Canada"))
        self.assertFalse(_use_imperial("CA"))

    @unittest.skipUnless(TPL2.is_file(), "template missing")
    def test_write_workbook_from_template1_sheet(self):
        result = PackingResult(
            country="美国",
            template1_rows=[
                Template1Row(
                    is_full_box=True,
                    country="美国",
                    sku="FULL",
                    qty=100,
                    msku="FULL-M",
                    units_per_box=50,
                    box_count=2,
                    box_weight_kg=10.0,
                    length_cm=40,
                    width_cm=30,
                    height_cm=20,
                ),
                Template1Row(
                    is_full_box=False,
                    country="美国",
                    sku="PART",
                    qty=15,
                    msku="PART-M",
                ),
            ],
        )
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "m.xlsx"
            create_amazon_workbook(template_source=TPL2, output_path=out, result=result)
            from openpyxl import load_workbook

            wb = load_workbook(out, data_only=True)
            ws = wb["Create workflow – template"]
            header = next(
                r for r in range(1, 15) if ws.cell(r, 1).value == "Merchant SKU"
            )
            self.assertEqual(ws.cell(header + 1, 1).value, "FULL-M")
            self.assertEqual(ws.cell(header + 1, 2).value, 100)
            self.assertEqual(ws.cell(header + 1, 5).value, 50)
            self.assertEqual(ws.cell(header + 1, 7).value, round(40 * CM_TO_INCH, 2))
            self.assertEqual(ws.cell(header + 1, 10).value, round(10 * KG_TO_LB, 2))
            self.assertEqual(ws.cell(header + 2, 1).value, "PART-M")
            self.assertEqual(ws.cell(header + 2, 2).value, 15)
            self.assertIn(ws.cell(header + 2, 5).value, (None, ""))

    @unittest.skipUnless(TPL2.is_file(), "template missing")
    def test_canada_keeps_metric_units(self):
        result = PackingResult(
            country="加拿大",
            template1_rows=[
                Template1Row(
                    is_full_box=True,
                    country="加拿大",
                    sku="FULL",
                    qty=100,
                    msku="FULL-M",
                    units_per_box=50,
                    box_count=2,
                    box_weight_kg=10.0,
                    length_cm=40,
                    width_cm=30,
                    height_cm=20,
                ),
            ],
        )
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "m.xlsx"
            create_amazon_workbook(template_source=TPL2, output_path=out, result=result)
            from openpyxl import load_workbook

            wb = load_workbook(out, data_only=True)
            ws = wb["Create workflow – template"]
            header = next(
                r for r in range(1, 15) if ws.cell(r, 1).value == "Merchant SKU"
            )
            self.assertEqual(ws.cell(header + 1, 7).value, 40)
            self.assertEqual(ws.cell(header + 1, 8).value, 30)
            self.assertEqual(ws.cell(header + 1, 9).value, 20)
            self.assertEqual(ws.cell(header + 1, 10).value, 10)

    def test_canada_picks_mpl3_shell(self):
        import pinxiang_config

        us = pinxiang_config.amazon_template_mpl("美国")
        ca = pinxiang_config.amazon_template_mpl("加拿大")
        self.assertTrue(str(us).endswith("MPL.xlsx"))
        self.assertTrue(str(ca).endswith("MPL3.xlsx"))
        self.assertTrue(pinxiang_config.AMAZON_TEMPLATE_MPL3.is_file())
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "ca.xlsx"
            result = PackingResult(
                country="加拿大",
                template1_rows=[
                    Template1Row(
                        is_full_box=True,
                        country="加拿大",
                        sku="A",
                        qty=50,
                        msku="A",
                        units_per_box=25,
                        box_count=2,
                        box_weight_kg=12.0,
                        length_cm=40,
                        width_cm=30,
                        height_cm=20,
                    )
                ],
            )
            create_amazon_workbook(template_source=ca, output_path=out, result=result)
            from openpyxl import load_workbook

            wb = load_workbook(out, data_only=True)
            ws = wb["Create workflow – template"]
            header = next(r for r in range(1, 15) if ws.cell(r, 1).value == "Merchant SKU")
            # 表头应为 cm/kg
            self.assertIn("(cm)", str(ws.cell(header, 7).value or ""))
            self.assertIn("(kg)", str(ws.cell(header, 10).value or ""))
            self.assertEqual(ws.cell(header + 1, 7).value, 40)
            self.assertEqual(ws.cell(header + 1, 10).value, 12)

    @unittest.skipUnless(
        TPL.is_file() and TEST_SHIPMENT.is_file() and TEST_MANIFEST.is_file(),
        "desktop fixtures missing",
    )
    def test_manifest_matches_desktop_sample(self):
        result = process_shipment_file(TEST_SHIPMENT)
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "m.xlsx"
            create_amazon_workbook(template_source=TPL, output_path=out, result=result)
            from openpyxl import load_workbook

            got_wb = load_workbook(out, data_only=True)
            exp_wb = load_workbook(TEST_MANIFEST, data_only=True)
            got = got_wb["Create workflow – template"]
            exp = exp_wb["Create workflow – template"]
            gh = next(r for r in range(1, 15) if got.cell(r, 1).value == "Merchant SKU")
            eh = next(r for r in range(1, 15) if exp.cell(r, 1).value == "Merchant SKU")

            def rows(ws, header):
                out_rows = []
                r = header + 1
                while True:
                    sku = ws.cell(r, 1).value
                    if sku is None or str(sku).strip() == "":
                        break
                    out_rows.append(
                        (
                            str(sku).strip(),
                            float(ws.cell(r, 2).value or 0),
                            ws.cell(r, 5).value,
                            ws.cell(r, 6).value,
                            ws.cell(r, 7).value,
                            ws.cell(r, 8).value,
                            ws.cell(r, 9).value,
                            ws.cell(r, 10).value,
                        )
                    )
                    r += 1
                return out_rows

            got_rows = rows(got, gh)
            exp_rows = rows(exp, eh)
            # 整箱段顺序与样例一致；非整箱仅校验集合（样例手工顺序与 sheet 略有不同）
            self.assertEqual(got_rows[:21], exp_rows[:21])
            self.assertEqual(sorted(got_rows[21:]), sorted(exp_rows[21:]))
            # 与 template1 一一对应
            self.assertEqual(len(got_rows), len(result.template1_rows))
            for i, t1 in enumerate(result.template1_rows):
                self.assertEqual(got_rows[i][0], t1.merchant_sku)
                self.assertEqual(got_rows[i][1], float(t1.qty))
                if t1.is_full_box:
                    self.assertEqual(got_rows[i][2], int(t1.units_per_box or 0))
                else:
                    self.assertIn(got_rows[i][2], (None, ""))


if __name__ == "__main__":
    unittest.main()
