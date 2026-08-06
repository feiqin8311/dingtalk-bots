#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""将拼箱结果写入 Amazon Manifest 上传模版。

规则（流程三）：
- 优先用「用于生成模版1」：整箱(是)填箱规；非整箱(否)仅 Merchant SKU + Quantity
- 无该 sheet 时回退 amazon_rows：原箱填箱规；改箱同 SKU 合并数量且不写尺寸
- 国家：美国 → cm/kg 换 in/lb；加拿大 → 公制直填；其它默认按美国
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Union

from openpyxl import load_workbook
from openpyxl.styles import Border, Side

from packing import PackingResult, PackingRow, Template1Row  # noqa: E402

PathLike = Union[str, Path]

CM_TO_INCH = 0.393701
KG_TO_LB = 2.20462

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

TEMPLATE_SHEET = "Create workflow – template"

# 视为原厂包装（不拼箱）的 remark
_ORIGINAL_CASE_REMARKS = frozenset({"原箱", "原箱-整数部分"})

_CANADA_MARKERS = frozenset({"加拿大", "canada", "ca", "can"})
_US_MARKERS = frozenset({"美国", "美國", "us", "usa", "united states", "united states of america"})


def create_amazon_workbook(
    *,
    template_source: PathLike,
    output_path: PathLike,
    result: PackingResult,
    default_prep_owner: str = "Seller",
    default_labeling_owner: str = "Seller",
) -> Path:
    source = Path(template_source)
    if not source.is_file():
        raise FileNotFoundError(f"Amazon 模版不存在: {source}")

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, out)

    wb = load_workbook(out)
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"Amazon 模版缺少工作表: {TEMPLATE_SHEET}")
    ws = wb[TEMPLATE_SHEET]

    _apply_default_owners(ws, default_prep_owner, default_labeling_owner)

    imperial = _use_imperial(_resolve_country(result))
    header_row = _find_header_row(ws)
    data_row = header_row + 1
    if result.template1_rows:
        for t1 in result.template1_rows:
            _write_template1_data_row(ws, data_row, t1, imperial=imperial)
            data_row += 1
    else:
        for row, with_case_pack in _manifest_rows(result.amazon_rows):
            _write_data_row(
                ws, data_row, row, with_case_pack=with_case_pack, imperial=imperial
            )
            data_row += 1

    wb.save(out)
    return out


def reformat_amazon_workbook(
    *,
    data_source: PathLike,
    template_source: PathLike,
    output_path: PathLike,
    default_prep_owner: str = "Seller",
    default_labeling_owner: str = "Seller",
) -> Path:
    """把已填模板的数据行原样拷到新壳子（模板2 与模板1 数据一致，仅格式不同）。"""
    src = Path(data_source)
    if not src.is_file():
        raise FileNotFoundError(f"Amazon 数据源不存在: {src}")
    shell = Path(template_source)
    if not shell.is_file():
        raise FileNotFoundError(f"Amazon 模版不存在: {shell}")

    src_wb = load_workbook(src, data_only=True)
    try:
        if TEMPLATE_SHEET not in src_wb.sheetnames:
            raise ValueError(f"数据源缺少工作表: {TEMPLATE_SHEET}")
        src_ws = src_wb[TEMPLATE_SHEET]
        src_header = _find_header_row(src_ws)
        rows: list[list] = []
        r = src_header + 1
        while True:
            sku = src_ws.cell(r, 1).value
            if sku is None or str(sku).strip() == "":
                break
            rows.append([src_ws.cell(r, c).value for c in range(1, 11)])
            r += 1
    finally:
        src_wb.close()

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(shell, out)
    wb = load_workbook(out)
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"Amazon 模版缺少工作表: {TEMPLATE_SHEET}")
    ws = wb[TEMPLATE_SHEET]
    _apply_default_owners(ws, default_prep_owner, default_labeling_owner)
    header_row = _find_header_row(ws)
    for i, values in enumerate(rows):
        row_idx = header_row + 1 + i
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
    wb.save(out)
    return out


def _apply_default_owners(ws, prep: str, labeling: str) -> None:
    for row_idx in range(1, 8):
        label = str(ws.cell(row_idx, 1).value or "").strip()
        if label == "Default prep owner":
            ws.cell(row_idx, 2, prep)
        elif label == "Default labeling owner":
            ws.cell(row_idx, 2, labeling)


def _find_header_row(ws) -> int:
    for row in range(1, (ws.max_row or 1) + 1):
        if str(ws.cell(row=row, column=1).value or "").strip() == "Merchant SKU":
            return row
    return 8


def _is_mixed_packing_row(row: PackingRow) -> bool:
    """True = 拼箱（不写尺寸、同 SKU 合并）；False = 不拼箱原厂包装。"""
    remark = (row.remark or "").strip()
    if remark in _ORIGINAL_CASE_REMARKS:
        return False
    if remark and remark not in {"修正版"}:
        # 改箱 / 合箱 / 借箱 / 余数 / 整票拼1箱 等
        return True
    # 无 remark 或修正版：有合箱组、或「整行一箱」形态视为拼箱
    if (row.box_group_id or "").strip():
        return True
    if (
        row.qty > 0
        and abs(row.units_per_box - row.qty) < 1e-9
        and abs(row.box_count - 1.0) < 1e-9
    ):
        return True
    return False


def _sku_key(row: PackingRow) -> str:
    return str(row.msku or row.sku or "").strip()


def _manifest_rows(rows: list[PackingRow]) -> list[tuple[PackingRow, bool]]:
    """生成写入模板的行：(row, with_case_pack)。拼箱同 SKU 合并 qty。"""
    originals: list[PackingRow] = []
    mixed_order: list[str] = []
    mixed_qty: dict[str, float] = {}
    mixed_sample: dict[str, PackingRow] = {}

    for row in rows:
        key = _sku_key(row)
        if not key:
            continue
        if _is_mixed_packing_row(row):
            if key not in mixed_qty:
                mixed_order.append(key)
                mixed_sample[key] = row
                mixed_qty[key] = float(row.qty)
            else:
                mixed_qty[key] += float(row.qty)
        else:
            originals.append(row)

    out: list[tuple[PackingRow, bool]] = []
    for row in originals:
        out.append((row, True))
    for key in mixed_order:
        sample = mixed_sample[key]
        qty = mixed_qty[key]
        merged = PackingRow(
            warehouse=sample.warehouse,
            sku=sample.sku,
            msku=sample.msku or sample.sku,
            qty=qty,
            units_per_box=0.0,
            box_count=0.0,
            box_weight_kg=0.0,
            length_cm=0.0,
            width_cm=0.0,
            height_cm=0.0,
            remark=sample.remark or "拼箱合并",
            box_group_id="",
        )
        out.append((merged, False))
    return out


def _num_cell(value: float):
    if value is None:
        return ""
    if abs(value - round(value)) < 1e-9:
        return int(round(value))
    return value


def _resolve_country(result: PackingResult) -> str:
    if (result.country or "").strip():
        return result.country.strip()
    for row in result.template1_rows or []:
        if (row.country or "").strip():
            return row.country.strip()
    return ""


def _use_imperial(country: str) -> bool:
    """美国/空 → 英制；加拿大 → 公制。"""
    text = (country or "").strip().lower()
    if not text:
        return True
    if text in _CANADA_MARKERS or "加拿大" in (country or "") or "canada" in text:
        return False
    if text in _US_MARKERS or "美国" in (country or "") or "美國" in (country or ""):
        return True
    # 未识别：沿用原默认（转英制）
    return True


def _dim_weight_cells(
    length_cm: float,
    width_cm: float,
    height_cm: float,
    weight_kg: float,
    *,
    imperial: bool,
) -> list:
    if not length_cm and not width_cm and not height_cm and not weight_kg:
        return ["", "", "", ""]
    if imperial:
        return [
            round(length_cm * CM_TO_INCH, 2) if length_cm else "",
            round(width_cm * CM_TO_INCH, 2) if width_cm else "",
            round(height_cm * CM_TO_INCH, 2) if height_cm else "",
            round(weight_kg * KG_TO_LB, 2) if weight_kg else "",
        ]
    return [
        round(length_cm, 2) if length_cm else "",
        round(width_cm, 2) if width_cm else "",
        round(height_cm, 2) if height_cm else "",
        round(weight_kg, 2) if weight_kg else "",
    ]


def _write_data_row(
    ws, row_idx: int, row: PackingRow, *, with_case_pack: bool, imperial: bool = True
) -> None:
    msku = row.msku or row.sku
    if with_case_pack:
        values = [
            msku,
            _num_cell(row.qty),
            "",  # Expiration date
            "",  # Manufacturing lot code
            _num_cell(row.units_per_box),
            _num_cell(row.box_count),
            *_dim_weight_cells(
                row.length_cm,
                row.width_cm,
                row.height_cm,
                row.box_weight_kg,
                imperial=imperial,
            ),
        ]
    else:
        # 拼箱：仅 SKU + 数量，尺寸/箱规留空
        values = [
            msku,
            _num_cell(row.qty),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = THIN_BORDER


def _write_template1_data_row(
    ws, row_idx: int, row: Template1Row, *, imperial: bool = True
) -> None:
    """按「用于生成模版1」写 Manifest 行。"""
    msku = row.merchant_sku
    if row.is_full_box:
        values = [
            msku,
            _num_cell(row.qty),
            "",
            "",
            _num_cell(row.units_per_box or 0),
            _num_cell(row.box_count or 0),
            *_dim_weight_cells(
                float(row.length_cm or 0),
                float(row.width_cm or 0),
                float(row.height_cm or 0),
                float(row.box_weight_kg or 0),
                imperial=imperial,
            ),
        ]
    else:
        values = [msku, _num_cell(row.qty), "", "", "", "", "", "", "", ""]
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = THIN_BORDER
