#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""可选：产品信息单证专用表（本地 / SMB）。未配置时拼箱用发货单装箱字段。"""

from __future__ import annotations

import socket
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Optional, Tuple
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook

from packing import ProductSpec


def load_product_specs(
    path: str,
    *,
    smb_username: str = "",
    smb_password: str = "",
    smb_port: int = 445,
    smb_timeout_sec: int = 30,
    smb_client_name: str = "dingtalk-pinxiang-bot",
) -> Dict[str, ProductSpec]:
    raw = (path or "").strip()
    if not raw:
        return {}

    if _is_smb_path(raw):
        host, share, remote_path = _parse_smb_path(raw)
        data = _read_smb_file(
            host=host,
            share=share,
            remote_path=remote_path,
            username=smb_username,
            password=smb_password,
            port=smb_port,
            timeout_sec=smb_timeout_sec,
            client_name=smb_client_name,
        )
        wb = load_workbook(filename=data, data_only=True, read_only=True)
    else:
        local = Path(raw)
        if not local.is_file():
            raise FileNotFoundError(f"产品信息表不存在: {local}")
        wb = load_workbook(filename=str(local), data_only=True, read_only=True)

    try:
        sheet = wb[wb.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return {}

    headers = [_norm(h) for h in rows[0]]
    # (normalized_header, column_index) — keep duplicates (表里可能有两套 SKU 列)
    header_pairs = [(_header_key(h), i) for i, h in enumerate(headers) if _header_key(h)]

    sku_col = _find_col(header_pairs, ("*sku", "sku", "产品sku", "货品sku"))
    if sku_col is None:
        raise ValueError("产品信息表缺少 SKU 列")

    # 优先单证表真实列名：单箱数量(pcs) / 单箱重量(kg) / 外箱规格长宽高(cm)
    units_col = _find_col(header_pairs, ("单箱数量(pcs)", "单箱数量", "外箱数量", "箱规数量"))
    weight_col = _find_col(header_pairs, ("单箱重量(kg)", "单箱重量", "单箱毛重", "外箱重量", "箱子毛重"))
    spec_col = _find_col(header_pairs, ("外箱规格", "箱规", "外箱尺寸"))
    l_col = _find_col(header_pairs, ("外箱规格长(cm)", "外箱规格长", "外箱长", "箱子长度", "长"))
    w_col = _find_col(header_pairs, ("外箱规格宽(cm)", "外箱规格宽", "外箱宽", "箱子宽度", "宽"))
    h_col = _find_col(header_pairs, ("外箱规格高(cm)", "外箱规格高", "外箱高", "箱子高度", "高"))

    out: Dict[str, ProductSpec] = {}
    for raw_row in rows[1:]:
        if not raw_row:
            continue
        sku = _cell(raw_row, sku_col)
        if not sku:
            continue
        length = width = height = None
        if spec_col is not None:
            length, width, height = _parse_dims(_cell(raw_row, spec_col))
        if l_col is not None:
            length = _float(_cell(raw_row, l_col)) or length
        if w_col is not None:
            width = _float(_cell(raw_row, w_col)) or width
        if h_col is not None:
            height = _float(_cell(raw_row, h_col)) or height

        out[sku] = ProductSpec(
            sku=sku,
            units_per_box=_float(_cell(raw_row, units_col)) if units_col is not None else None,
            box_weight_kg=_float(_cell(raw_row, weight_col)) if weight_col is not None else None,
            length_cm=length,
            width_cm=width,
            height_cm=height,
        )
    return out


def _is_smb_path(path: str) -> bool:
    value = path.strip()
    return value.lower().startswith("smb://") or value.startswith("\\\\") or value.startswith("//")


def _parse_smb_path(path: str) -> Tuple[str, str, str]:
    value = path.strip()
    if value.lower().startswith("smb://"):
        parsed = urlparse(value)
        host = unquote(parsed.hostname or "")
        pieces = [unquote(part) for part in parsed.path.split("/") if part]
    elif value.startswith("\\\\"):
        parts = [part for part in value.strip("\\").split("\\") if part]
        host = parts[0] if parts else ""
        pieces = parts[1:]
    else:
        parts = [unquote(part) for part in value.strip("/").split("/") if part]
        host = parts[0] if parts else ""
        pieces = parts[1:]
    if len(pieces) < 2 or not host:
        raise ValueError(f"SMB path must include host, share, and file path: {path}")
    return host, pieces[0], "/" + "/".join(pieces[1:])


def _read_smb_file(
    *,
    host: str,
    share: str,
    remote_path: str,
    username: str,
    password: str,
    port: int,
    timeout_sec: int,
    client_name: str,
) -> BytesIO:
    if not username or not password:
        raise RuntimeError("读取 SMB 产品信息表需要 SMB_USERNAME / SMB_PASSWORD")
    try:
        from smb.SMBConnection import SMBConnection  # type: ignore
    except Exception as exc:
        raise RuntimeError("需要 pysmb 才能读取 SMB 产品信息表") from exc

    buffer = BytesIO()
    machine_name = client_name or socket.gethostname() or "dingtalk-pinxiang-bot"
    conn = SMBConnection(username, password, machine_name, host, use_ntlm_v2=True, is_direct_tcp=True)
    try:
        if not conn.connect(host, port, timeout=timeout_sec):
            raise RuntimeError(f"SMB connect/auth failed: {host}:{port}")
        conn.retrieveFile(share, remote_path, buffer, timeout=timeout_sec)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    buffer.seek(0)
    return buffer


def _norm(value: Any) -> str:
    return str(value or "").strip()


def _header_key(value: Any) -> str:
    """Normalize header for matching: drop *, newlines, spaces; lower-case."""
    text = _norm(value).replace("\n", "").replace("\r", "").replace(" ", "")
    if text.startswith("*"):
        text = text[1:]
    return text.lower()


def _find_col(header_pairs: list[tuple[str, int]], candidates: tuple[str, ...]) -> Optional[int]:
    """Find column among (norm_header, index) pairs; leftmost wins on ties."""
    norms = [_header_key(c) for c in candidates]

    def _best(matches: list[int]) -> Optional[int]:
        return min(matches) if matches else None

    for name in norms:
        hits = [col for key, col in header_pairs if key == name]
        found = _best(hits)
        if found is not None:
            return found
    for name in norms:
        hits = [
            col
            for key, col in header_pairs
            if key.startswith(name) or (len(key) >= 2 and name.startswith(key))
        ]
        found = _best(hits)
        if found is not None:
            return found
    for name in norms:
        if len(name) < 2:
            continue
        hits = [col for key, col in header_pairs if name in key]
        found = _best(hits)
        if found is not None:
            return found
    return None


def _cell(row: tuple, col: Optional[int]) -> str:
    if col is None or col >= len(row) or row[col] is None:
        return ""
    return str(row[col]).strip()


def _float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace("kg", "").replace("KG", "").strip())
    except ValueError:
        return None


def _parse_dims(text: str) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if not text:
        return None, None, None
    cleaned = text.lower().replace("cm", "").replace("×", "x").replace("*", "x")
    parts = [p.strip() for p in cleaned.split("x") if p.strip()]
    if len(parts) < 3:
        return None, None, None
    try:
        return float(parts[0]), float(parts[1]), float(parts[2])
    except ValueError:
        return None, None, None
