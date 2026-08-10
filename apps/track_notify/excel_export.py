from __future__ import annotations

import re
from dataclasses import dataclass, field, replace
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment

# 物流详情行：以 YYYY-MM-DD 开头视为「有日期」
_DETAIL_DATE_PREFIX = re.compile(r"^\d{4}-\d{2}-\d{2}\b")


# 业务回传列（按产品要求）
HEADERS = (
    "发票号",
    "品牌",
    "国家",
    "FBA编码",
    "物流编号",
    "货代公司",
    "发货时间",
    "预计船期",
    "实际送仓时间",
    "负责人",
    "物流详情",
)
_DETAIL_COL = 11  # 物流详情列（1-based）


@dataclass
class ReportItem:
    """一条写入 Excel / 去重库的结果行（一单多节点时 detail 内换行汇总）。"""

    shipment_key: str
    event_key: str
    message: str
    user_ids: list[str]
    invoice_no: str = ""
    brand: str = ""
    country: str = ""
    fba_code: str = ""
    logistics_no: str = ""
    carrier: str = ""
    shipped_at: str = ""
    eta_date: str = ""
    delivered_at: str = ""
    owners: str = ""
    detail: str = ""  # 物流详情：指定节点（多节点用 \\n）
    # 发送成功后需 mark 的节点 key；空则只 mark event_key
    event_keys: list[str] = field(default_factory=list)

    def excel_row(self) -> list[str]:
        return [
            self.invoice_no,
            self.brand,
            self.country,
            self.fba_code,
            self.logistics_no,
            self.carrier,
            self.shipped_at,
            self.eta_date,
            self.delivered_at,
            self.owners,
            self.detail or self.message,
        ]

    def keys_to_mark(self) -> list[str]:
        return list(self.event_keys) if self.event_keys else [self.event_key]


def detail_line_has_date(line: str) -> bool:
    return bool(_DETAIL_DATE_PREFIX.match((line or "").strip()))


def filter_report_item_for_user(
    item: ReportItem,
    user_id: str,
    *,
    full_detail_user_id: str,
) -> ReportItem | None:
    """
    物流人员：物流详情只保留「有日期」的节点行；全无日期则本行不推给该用户。
    总览接收人（柯鹏翔）：完整详情不动。
    返回 None 表示该用户 Excel 不写此行。
    """
    if (user_id or "").strip() == (full_detail_user_id or "").strip():
        return item
    text = (item.detail or item.message or "").strip()
    if not text:
        return item
    lines = [ln for ln in text.splitlines() if ln.strip()]
    dated = [ln.strip() for ln in lines if detail_line_has_date(ln)]
    if not dated:
        return None
    if len(dated) == len(lines):
        return item
    new_detail = "\n".join(dated)
    return replace(item, detail=new_detail)


def format_shipped_at(value: date | datetime | str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def write_report_xlsx(items: Iterable[ReportItem], path: Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "轨迹回传"
    ws.append(list(HEADERS))
    wrap = Alignment(wrap_text=True, vertical="top")
    for item in items:
        ws.append(item.excel_row())
        ws.cell(row=ws.max_row, column=_DETAIL_COL).alignment = wrap
    for col, width in enumerate(
        [14, 12, 10, 18, 22, 10, 12, 12, 14, 14, 48], start=1
    ):
        ws.column_dimensions[chr(64 + col)].width = width
    wb.save(path)
    return path


def export_filename(*, user_id: str = "", when: datetime | None = None) -> str:
    """例：乔丹丹_轨迹回传_20260805_114336.xlsx（含时分秒，避免同日多次跑互相覆盖）。"""
    from owners import display_name

    when = when or datetime.now()
    stamp = when.strftime("%Y%m%d_%H%M%S")
    name = display_name(user_id) if user_id else "全部"
    # 去掉路径不安全字符即可
    safe = "".join("_" if c in r'\/:*?"<>|' else c for c in name).strip() or "unknown"
    return f"{safe}_轨迹回传_{stamp}.xlsx"
