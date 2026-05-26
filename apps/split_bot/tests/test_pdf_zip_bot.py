import tempfile
import unittest
import zipfile
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from pypdf import PdfReader, PdfWriter

from pdf_zip_bot import (
    RuleParseError,
    build_rules_from_workbook,
    format_rule_preview_markdown,
    format_rule_preview_table,
    parse_rules_table,
    process_pdf_to_zip,
)


class ParseRulesTableTests(unittest.TestCase):
    def test_parses_tabular_text_into_rule_rows(self):
        raw = (
            "宁波德韵工具有限公司\t900913\t1-3\n"
            "宁波德韵工具有限公司\t919019\t5\n"
            "余姚嘉城工具工贸有限公司\t911212\t7-8\n"
        )

        rules = parse_rules_table(raw)

        self.assertEqual(
            [rule.company_name for rule in rules],
            ["宁波德韵工具有限公司", "宁波德韵工具有限公司", "余姚嘉城工具工贸有限公司"],
        )
        self.assertEqual([rule.reference_code for rule in rules], ["900913", "919019", "911212"])
        self.assertEqual([rule.page_spec for rule in rules], ["1-3", "5", "7-8"])

    def test_rejects_malformed_rows_with_line_numbers(self):
        with self.assertRaises(RuleParseError) as ctx:
            parse_rules_table("宁波德韵工具有限公司\t900913\n")

        self.assertIn("line 1", str(ctx.exception))

    def test_builds_rules_from_excel_and_pdf_text(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["供应商", "SKU"])
        sheet.append(["宁波德韵工具有限公司", "900913"])
        sheet.append(["宁波德韵工具有限公司", 919019])
        sheet.append(["仓库发", None])
        buffer = BytesIO()
        workbook.save(buffer)

        rules = build_rules_from_workbook(
            buffer.getvalue(),
            [
                "SKU 900913",
                "SKU 900913",
                "SKU 919019",
                "SKU 919019",
                "SKU 919014",
                "SKU 900913",
            ],
        )

        self.assertEqual(
            [rule.company_name for rule in rules],
            ["宁波德韵工具有限公司", "宁波德韵工具有限公司", "仓库发"],
        )
        self.assertEqual([rule.reference_code for rule in rules], ["900913", "919019", ""])
        self.assertEqual([rule.page_spec for rule in rules], ["1-2,6", "3-4", "5"])

    def test_rejects_excel_without_supplier_and_sku_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["公司名", "编号/备注"])
        buffer = BytesIO()
        workbook.save(buffer)

        rules = build_rules_from_workbook(buffer.getvalue(), ["SKU 900913"])

        self.assertEqual([(rule.company_name, rule.reference_code, rule.page_spec) for rule in rules], [("仓库发", "", "1")])

    def test_assigns_unmatched_pages_to_warehouse_without_failing_missing_sku(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["供应商", "SKU"])
        sheet.append(["宁波德韵工具有限公司", "999999"])
        sheet.append(["仓库发", None])
        buffer = BytesIO()
        workbook.save(buffer)

        rules = build_rules_from_workbook(buffer.getvalue(), ["SKU 900913"])

        self.assertEqual([(rule.company_name, rule.reference_code, rule.page_spec) for rule in rules], [("仓库发", "", "1")])

    def test_matches_sku_as_whole_token_not_substring(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["供应商", "SKU"])
        sheet.append(["宁波德韵工具有限公司", "9009"])
        sheet.append(["仓库发", None])
        buffer = BytesIO()
        workbook.save(buffer)

        rules = build_rules_from_workbook(buffer.getvalue(), ["SKU 900913"])

        self.assertEqual([(rule.company_name, rule.reference_code, rule.page_spec) for rule in rules], [("仓库发", "", "1")])

    def test_formats_preview_table_with_split_page_ranges(self):
        rules = parse_rules_table(
            "宁波德韵工具有限公司\t900913\t1-10,12,15-20\n"
            "宁波德韵工具有限公司\t919019\t22-30\n"
            "仓库发\t\t31-40\n"
        )

        result = format_rule_preview_table(rules)

        self.assertEqual(
            result,
            "\n".join(
                [
                    "供应商\tSKU\t页数",
                    "宁波德韵工具有限公司\t900913\t1-10",
                    "宁波德韵工具有限公司\t900913\t12",
                    "宁波德韵工具有限公司\t900913\t15-20",
                    "宁波德韵工具有限公司\t919019\t22-30",
                    "仓库发\t\t31-40",
                ]
            ),
        )

    def test_formats_preview_markdown_table(self):
        rules = parse_rules_table(
            "宁波德韵工具有限公司\t900913\t1-10,12\n"
            "仓库发\t\t31-40\n"
        )

        result = format_rule_preview_markdown(rules)

        self.assertEqual(
            result,
            "\n".join(
                [
                    "| 供应商 | SKU | 页数 |",
                    "| --- | --- | --- |",
                    "| 宁波德韵工具有限公司 | 900913 | 1-10 |",
                    "| 宁波德韵工具有限公司 | 900913 | 12 |",
                    "| 仓库发 |  | 31-40 |",
                ]
            ),
        )

    def test_uses_split_page_column_when_present(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["供应商", "SKU", "拆分页面"])
        sheet.append(["余姚嘉城工具工贸有限公司", "900920BR", 1])
        sheet.append(["余姚嘉城工具工贸有限公司", 932101, "117-145"])
        sheet.append(["仓库发", None, "2-29"])
        buffer = BytesIO()
        workbook.save(buffer)

        rules = build_rules_from_workbook(buffer.getvalue(), ["SKU 000000"])

        self.assertEqual(
            [(rule.company_name, rule.reference_code, rule.page_spec) for rule in rules],
            [
                ("余姚嘉城工具工贸有限公司", "900920BR", "1"),
                ("余姚嘉城工具工贸有限公司", "932101", "117-145"),
                ("仓库发", "", "2-29"),
            ],
        )

    def test_ignores_empty_split_page_column_and_uses_sku_matching(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["供应商", "SKU", "拆分页面"])
        sheet.append(["杭州斯乐恩实业有限公司", 912801, None])
        sheet.append(["仓库发", None, None])
        buffer = BytesIO()
        workbook.save(buffer)

        rules = build_rules_from_workbook(buffer.getvalue(), ["SKU 912801", "SKU 000000"])

        self.assertEqual(
            [(rule.company_name, rule.reference_code, rule.page_spec) for rule in rules],
            [
                ("杭州斯乐恩实业有限公司", "912801", "1"),
                ("仓库发", "", "2"),
            ],
        )

    def test_split_page_column_disables_pdf_sku_matching(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["供应商", "SKU", "拆分页面"])
        sheet.append(["宁波德韵工具有限公司", "919005", "82-116"])
        buffer = BytesIO()
        workbook.save(buffer)

        rules = build_rules_from_workbook(buffer.getvalue(), [])

        self.assertEqual([(rule.company_name, rule.reference_code, rule.page_spec) for rule in rules], [("宁波德韵工具有限公司", "919005", "82-116")])

    def test_uses_column_positions_even_with_nonstandard_headers(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["随便写", "别名SKU", "范围"])
        sheet.append(["宁波德韵工具有限公司", "919005", "82-116"])
        buffer = BytesIO()
        workbook.save(buffer)

        rules = build_rules_from_workbook(buffer.getvalue(), [])

        self.assertEqual([(rule.company_name, rule.reference_code, rule.page_spec) for rule in rules], [("宁波德韵工具有限公司", "919005", "82-116")])


class ProcessPdfToZipTests(unittest.TestCase):
    def test_groups_by_company_and_creates_one_pdf_per_company(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            source_pdf = tmp_path / "source.pdf"
            self._make_pdf(source_pdf, 10)
            rules = parse_rules_table(
                "宁波德韵工具有限公司\t900913\t1-2\n"
                "宁波德韵工具有限公司\t919019\t4\n"
                "余姚嘉城工具工贸有限公司\t911212\t7-9\n"
            )

            zip_path = process_pdf_to_zip(source_pdf, rules, tmp_path / "out")

            self.assertTrue(zip_path.exists())
            self.assertEqual(zip_path.name, "source.zip")
            with zipfile.ZipFile(zip_path) as zf:
                self.assertEqual(
                    sorted(zf.namelist()),
                    [
                        "source-余姚嘉城工具工贸有限公司-一式两份.pdf",
                        "source-宁波德韵工具有限公司-一式两份.pdf",
                    ],
                )
                with zf.open("source-宁波德韵工具有限公司-一式两份.pdf") as fh:
                    company_a_pdf = PdfReader(fh)
                    self.assertEqual(len(company_a_pdf.pages), 3)
                with zf.open("source-余姚嘉城工具工贸有限公司-一式两份.pdf") as fh:
                    company_b_pdf = PdfReader(fh)
                    self.assertEqual(len(company_b_pdf.pages), 3)

    def test_warehouse_file_name_does_not_include_duplicate_suffix(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            source_pdf = tmp_path / "source.pdf"
            self._make_pdf(source_pdf, 3)
            rules = parse_rules_table("仓库发\t\t1-3\n")

            zip_path = process_pdf_to_zip(source_pdf, rules, tmp_path / "out")

            with zipfile.ZipFile(zip_path) as zf:
                self.assertEqual(zf.namelist(), ["source-仓库发.pdf"])

    def test_rejects_out_of_range_pages(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            source_pdf = tmp_path / "source.pdf"
            self._make_pdf(source_pdf, 5)
            rules = parse_rules_table("宁波德韵工具有限公司\t900913\t1-8\n")

            with self.assertRaises(ValueError) as ctx:
                process_pdf_to_zip(source_pdf, rules, tmp_path / "out")

            self.assertIn("line 1", str(ctx.exception))
            self.assertIn("8", str(ctx.exception))

    def _make_pdf(self, path: Path, pages: int) -> None:
        writer = PdfWriter()
        for _ in range(pages):
            writer.add_blank_page(width=72, height=72)
        with path.open("wb") as fh:
            writer.write(fh)


if __name__ == "__main__":
    unittest.main()

from pdf_zip_bot.messages import MessageFormatError, extract_rules_text_from_message


class MessageParsingTests(unittest.TestCase):
    def test_extracts_rules_from_command_message(self):
        message_text = "/pdfsplit\n```\n宁波德韵工具有限公司\t900913\t1-50\n```"

        result = extract_rules_text_from_message(message_text)

        self.assertEqual(result, "宁波德韵工具有限公司\t900913\t1-50")

    def test_rejects_missing_rule_rows(self):
        with self.assertRaises(MessageFormatError):
            extract_rules_text_from_message("/pdfsplit")
