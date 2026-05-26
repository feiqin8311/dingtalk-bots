from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from pypdf import PdfReader


class RuleParseError(ValueError):
    """Raised when a rule table cannot be parsed safely."""


@dataclass(frozen=True)
class SplitRule:
    company_name: str
    reference_code: str
    page_spec: str
    line_number: int


def parse_rules_table(raw_text: str) -> list[SplitRule]:
    """Parse pasted table text into normalized split rules.

    Accepts tab-separated rows or runs of 2+ spaces as delimiters.
    """
    cleaned = raw_text.replace("\ufeff", "").strip()
    if not cleaned:
        raise RuleParseError("rule table is empty")

    rules: list[SplitRule] = []
    for line_number, raw_line in enumerate(cleaned.splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        parts = _split_columns(line)
        if len(parts) != 3:
            raise RuleParseError(
                f"Malformed rule row at line {line_number}: expected 3 columns, got {len(parts)}"
            )
        company_name, reference_code, page_spec = (part.strip() for part in parts)
        if not company_name or not page_spec:
            raise RuleParseError(f"Malformed rule row at line {line_number}: missing company or page range")
        rules.append(
            SplitRule(
                company_name=company_name,
                reference_code=reference_code,
                page_spec=page_spec,
                line_number=line_number,
            )
        )

    if not rules:
        raise RuleParseError("rule table did not contain any usable rows")
    return rules


def parse_rules_workbook(content: bytes) -> list[SplitRule]:
    workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rules: list[SplitRule] = []

    for row_index, raw_row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = [_normalize_cell(value) for value in raw_row]
        if not any(cells):
            continue
        if _is_header_row(cells):
            continue

        trimmed = list(cells[:3])
        while len(trimmed) < 3:
            trimmed.append("")
        if len([value for value in cells[3:] if value]) > 0:
            raise RuleParseError(f"Malformed rule row at Excel row {row_index}: expected 3 columns")

        company_name, reference_code, page_spec = (value.strip() for value in trimmed)
        if not company_name or not page_spec:
            raise RuleParseError(f"Malformed rule row at Excel row {row_index}: missing company or page range")
        rules.append(
            SplitRule(
                company_name=company_name,
                reference_code=reference_code,
                page_spec=page_spec,
                line_number=row_index,
            )
        )

    if not rules:
        raise RuleParseError("rule workbook did not contain any usable rows")
    return rules


def build_rules_from_workbook(content: bytes, pdf_source: str | Path | Iterable[str]) -> list[SplitRule]:
    workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    header_row, column_indexes = _locate_workbook_layout(sheet)
    if _workbook_has_explicit_page_values(sheet, header_row, column_indexes):
        return _build_rules_from_explicit_pages(sheet, header_row, column_indexes)

    page_texts = _load_page_texts(pdf_source)
    rules: list[SplitRule] = []
    matched_pages: set[int] = set()
    warehouse_supplier = "仓库发"

    for row_index, raw_row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        supplier = _normalize_cell(raw_row[column_indexes["supplier"]]) if column_indexes["supplier"] < len(raw_row) else ""
        sku = _normalize_cell(raw_row[column_indexes["sku"]]) if column_indexes["sku"] < len(raw_row) else ""
        if not supplier and not sku:
            continue
        if supplier and not sku:
            warehouse_supplier = supplier
            continue
        if not supplier:
            raise RuleParseError(f"Malformed rule row at Excel row {row_index}: missing 供应商 or SKU")

        matching_pages = [
            page_number
            for page_number, page_text in enumerate(page_texts, start=1)
            if _page_contains_sku(page_text, sku)
        ]
        if not matching_pages:
            continue
        matched_pages.update(matching_pages)

        rules.append(
            SplitRule(
                company_name=supplier,
                reference_code=sku,
                page_spec=_compress_page_numbers(matching_pages),
                line_number=row_index,
            )
        )

    remaining_pages = [
        page_number
        for page_number in range(1, len(page_texts) + 1)
        if page_number not in matched_pages
    ]
    if remaining_pages:
        rules.append(
            SplitRule(
                company_name=warehouse_supplier,
                reference_code="",
                page_spec=_compress_page_numbers(remaining_pages),
                line_number=header_row,
            )
        )

    if not rules:
        raise RuleParseError("rule workbook did not contain any usable rows")
    return rules


def workbook_uses_explicit_pages(content: bytes) -> bool:
    workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    header_row, column_indexes = _locate_workbook_layout(sheet)
    return _workbook_has_explicit_page_values(sheet, header_row, column_indexes)


def format_rule_preview_table(rules: list[SplitRule]) -> str:
    lines = ["供应商\tSKU\t页数"]
    for rule in rules:
        for segment in [chunk.strip() for chunk in rule.page_spec.split(",") if chunk.strip()]:
            lines.append(f"{rule.company_name}\t{rule.reference_code}\t{segment}")
    return "\n".join(lines)


def format_rule_preview_markdown(rules: list[SplitRule]) -> str:
    lines = [
        "| 供应商 | SKU | 页数 |",
        "| --- | --- | --- |",
    ]
    for rule in rules:
        sku = rule.reference_code or ""
        for segment in [chunk.strip() for chunk in rule.page_spec.split(",") if chunk.strip()]:
            lines.append(f"| {rule.company_name} | {sku} | {segment} |")
    return "\n".join(lines)


def _split_columns(line: str) -> list[str]:
    if "\t" in line:
        return next(csv.reader(io.StringIO(line), delimiter="\t"))
    return [part for part in re.split(r"\s{2,}", line) if part]


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_header_row(cells: list[str]) -> bool:
    if len(cells) < 3:
        return False
    first = cells[0].strip().lower()
    third = cells[2].strip().lower()
    return first in {"公司名", "company", "company name"} and third in {"页数范围", "page range", "pages"}


def _locate_workbook_layout(sheet) -> tuple[int, dict[str, int]]:
    for row_index, raw_row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = [_normalize_cell(value) for value in raw_row]
        if not any(cells):
            continue
        if len(cells) < 2:
            raise RuleParseError("Excel first sheet must use column A as 供应商 and column B as SKU")
        header_row = row_index
        columns = {"supplier": 0, "sku": 1}
        if len(cells) >= 3:
            columns["page_spec"] = 2
        return header_row, columns
    raise RuleParseError("Excel first sheet did not contain any usable rows")


def _build_rules_from_explicit_pages(sheet, header_row: int, column_indexes: dict[str, int]) -> list[SplitRule]:
    rules: list[SplitRule] = []
    for row_index, raw_row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        supplier = _normalize_cell(raw_row[column_indexes["supplier"]]) if column_indexes["supplier"] < len(raw_row) else ""
        sku_index = column_indexes.get("sku")
        sku = _normalize_cell(raw_row[sku_index]) if sku_index is not None and sku_index < len(raw_row) else ""
        page_spec = _normalize_cell(raw_row[column_indexes["page_spec"]]) if column_indexes["page_spec"] < len(raw_row) else ""
        if not supplier and not sku and not page_spec:
            continue
        if not supplier or not page_spec:
            raise RuleParseError(f"Malformed rule row at Excel row {row_index}: missing 供应商 or 拆分页面")
        rules.append(
            SplitRule(
                company_name=supplier,
                reference_code=sku,
                page_spec=page_spec,
                line_number=row_index,
            )
        )
    if not rules:
        raise RuleParseError("rule workbook did not contain any usable rows")
    return rules


def _workbook_has_explicit_page_values(sheet, header_row: int, column_indexes: dict[str, int]) -> bool:
    page_spec_index = column_indexes.get("page_spec")
    if page_spec_index is None:
        return False

    for raw_row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if page_spec_index < len(raw_row) and _normalize_cell(raw_row[page_spec_index]):
            return True
    return False


def _load_page_texts(pdf_source: str | Path | Iterable[str]) -> list[str]:
    if isinstance(pdf_source, (str, Path)):
        reader = PdfReader(str(pdf_source))
        return [(page.extract_text() or "") for page in reader.pages]
    return [str(value or "") for value in pdf_source]


def _compress_page_numbers(page_numbers: list[int]) -> str:
    if not page_numbers:
        raise RuleParseError("cannot compress empty page numbers")

    ranges: list[str] = []
    start = previous = page_numbers[0]
    for value in page_numbers[1:]:
        if value == previous + 1:
            previous = value
            continue
        ranges.append(_format_range(start, previous))
        start = previous = value
    ranges.append(_format_range(start, previous))
    return ",".join(ranges)


def _format_range(start: int, end: int) -> str:
    if start == end:
        return str(start)
    return f"{start}-{end}"


def _page_contains_sku(page_text: str, sku: str) -> bool:
    pattern = rf"(?<![0-9A-Za-z]){re.escape(sku)}(?![0-9A-Za-z])"
    return re.search(pattern, page_text) is not None
