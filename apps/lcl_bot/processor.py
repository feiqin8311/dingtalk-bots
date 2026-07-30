"""
File: processor.py
Created Time: 2025-11-07
Author: KePengxiang (Adapted for DingTalk Bot)
Comment: 美国拼箱处理 - Excel处理逻辑模块
"""

import pandas as pd
import warnings
import os
import shutil
import math
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.cell.cell import MergedCell

from . import config

# 设置pandas显示选项
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.max_colwidth', None)
pd.set_option('display.width', None)


class StyleManager:
    """样式管理器 - 集中管理所有Excel样式"""
    
    # 边框样式
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 字体样式
    TITLE_FONT = Font(size=16, bold=True, color="000000")
    STEP_FONT = Font(size=11, bold=True, color="000000")
    NORMAL_FONT = Font(size=11, color="000000")
    BOLD_FONT = Font(bold=True)
    ITALIC_FONT = Font(italic=True)
    RED_FONT = Font(color="FF0000", bold=True)
    ORANGE_FONT = Font(color="FF8C00", bold=True)
    RED_UNDERLINE_FONT = Font(color="FF0000", bold=True, underline='single')
    RED_LARGE_FONT = Font(color="FF0000", bold=True, size=20)
    
    # 填充样式
    GRAY_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    CREAM_FILL = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    
    # 对齐样式
    LEFT_ALIGNMENT = Alignment(horizontal='left', vertical='center')
    CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
    WRAP_ALIGNMENT = Alignment(horizontal='left', vertical='top', wrap_text=True)
    CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)


class UnitConverter:
    """单位转换器"""
    CM_TO_INCH = 0.393701  # 厘米转英寸
    G_TO_LB = 0.00220462   # 克转磅
    M3_TO_CM3 = 1000000    # 立方米转立方厘米
    KG_TO_G = 1000         # 千克转克


class PackingBoxProcessor:
    """拼箱处理器类 - 用于处理美国分仓拼箱业务"""

    def __init__(self, input_file_path, output_file_path):
        """初始化拼箱处理器"""
        # 文件路径配置
        self.input_file_path = input_file_path
        self.output_file_path = output_file_path
        
        # 重量配置（单位：克）
        self.target_weight = 14750
        self.min_weight = 11500
        self.max_weight = 18000
        self.min_weight_limit = 10000
        self.max_weight_limit = 18000
        self.weight_step = 100
        
        # 箱子尺寸配置（单位：厘米）
        self.initial_box_length = 20
        self.initial_box_width = 25
        self.default_box_length = 30.0
        self.default_box_width = 25.0
        self.min_box_dimension = 15
        self.max_box_dimension = 60
        self.dimension_step = 3
        
        # 其他配置
        self.boxes_per_share = 5
        self.min_boxes_for_split = 5
        self.max_iteration = 100
        
        # 列名配置
        self.REQUIRED_COLUMNS = ['单品长（cm）', '单品宽（cm）', '单品高（cm）', '单份数量', '小组名称']
        self.MERGE_COLUMNS = ['小组名称', '实际箱数', '单箱重量', '单箱体积', '单箱理论长', '单箱理论宽', '单箱理论高']
        self.SPECIAL_GROUPS = ['单独装箱', '拆箱单独装']

    def _is_eligible_for_non_integer_large_box_repacking(self, row) -> bool:
        """判断SKU是否需要尝试按新规则重算箱规。"""
        box_count = row.get("箱数")
        quantity = row.get("发货数量")
        per_box_quantity = row.get("单箱数量")
        box_gross_weight = row.get("箱子毛重（kg）")

        if any(pd.isna(value) for value in (box_count, quantity, per_box_quantity, box_gross_weight)):
            return False

        if box_count <= self.min_boxes_for_split:
            return False

        return not float(box_count).is_integer()

    def _repack_non_integer_large_box_row(self, row):
        """为大于5箱且非整箱的SKU寻找新的整数箱规。"""
        if not self._is_eligible_for_non_integer_large_box_repacking(row):
            return None

        quantity = int(row["发货数量"])
        original_per_box_quantity = float(row["单箱数量"])
        original_box_gross_weight = float(row["箱子毛重（kg）"])
        unit_gross_weight = original_box_gross_weight / original_per_box_quantity

        for candidate_box_count in range(6, quantity + 1):
            if quantity % candidate_box_count != 0:
                continue

            candidate_per_box_quantity = quantity // candidate_box_count
            candidate_box_gross_weight = unit_gross_weight * candidate_per_box_quantity
            if 5 <= candidate_box_gross_weight <= 18:
                updated_row = row.copy()
                updated_row["箱数"] = candidate_box_count
                updated_row["单箱数量"] = candidate_per_box_quantity
                updated_row["箱子毛重（kg）"] = candidate_box_gross_weight
                updated_row["_force_packing"] = True
                return updated_row

        return None

    def _apply_non_integer_large_box_repacking(self, invoice_df):
        """批量应用大于5箱且非整箱SKU的自动重算箱规逻辑。"""
        if "_force_packing" not in invoice_df.columns:
            invoice_df = invoice_df.copy()
            invoice_df["_force_packing"] = False
        repacked_rows = []
        failed_skus = []

        for _, row in invoice_df.iterrows():
            repacked_row = self._repack_non_integer_large_box_row(row)
            if repacked_row is not None:
                repacked_rows.append(repacked_row)
                continue

            if self._is_eligible_for_non_integer_large_box_repacking(row):
                failed_skus.append(
                    f"SKU={row.get('SKU', '')}, 发货数量={row.get('发货数量')}, "
                    f"原单箱数量={row.get('单箱数量')}, 原单箱毛重={row.get('箱子毛重（kg）')}kg"
                )
            unchanged_row = row.copy()
            if "_force_packing" not in unchanged_row.index:
                unchanged_row["_force_packing"] = False
            repacked_rows.append(unchanged_row)

        if failed_skus:
            details = "；".join(failed_skus)
            raise ValueError(
                "以下SKU无法自动重算为整数箱规（要求单箱数量为整数，且单箱毛重在5-18kg）："
                f"{details}。请调整数量后重试，或上传人工拼箱结果。"
            )

        return pd.DataFrame(repacked_rows, columns=invoice_df.columns)

    def _is_special_group_name(self, group_name):
        """判断小组名称是否为特殊组（支持包含匹配）"""
        if pd.isna(group_name):
            return False
        name = str(group_name)
        return any(token in name for token in self.SPECIAL_GROUPS)

    def merge_tail_with_non_packing(self, invoice_df, groups_list):
        """
        当尾组重量不达标时，贪心选取不拼箱SKU整体与尾组合并
        
        Args:
            invoice_df: 原始发货单数据（含不拼箱SKU）
            groups_list: 当前分组结果
        
        Returns:
            new_groups_list: 更新后的分组列表
            updated_invoice_df: 更新后的发货单（被使用的不拼箱SKU已移除）
        """
        if not groups_list:
            print("❌ 没有分组，无法调整")
            return groups_list, invoice_df
        
        last_group_name, last_group_skus, last_group_weight = groups_list[-1]
        
        # 检查是否需要合并
        if self.min_weight_limit <= last_group_weight <= self.max_weight_limit:
            print("✅ 尾组重量已在范围内（10000-18000g），无需合并")
            return groups_list, invoice_df
        
        print(f"🔄 尾组重量 {last_group_weight:.0f}g 不在范围内（10000-18000g），尝试与不拼箱SKU合并...")
        
        # 收集不拼箱SKU，计算每个SKU的总重量
        non_packing_df = invoice_df[invoice_df["是否拼箱"] == "不拼箱"].copy()
        
        if non_packing_df.empty:
            print("⚠️ 没有不拼箱SKU可用于合并，保持原状")
            return groups_list, invoice_df
        
        # 计算每个不拼箱SKU的总重量（整体）
        non_packing_df["总重量g"] = non_packing_df["箱子毛重（kg）"] * non_packing_df["箱数"] * 1000
        # 按总重量从小到大排序，优先选取小的
        non_packing_df = non_packing_df.sort_values("总重量g")
        
        merged_weight = last_group_weight
        merged_skus = list(last_group_skus)
        used_sku_names = []
        
        for _, row in non_packing_df.iterrows():
            sku_total_weight = row["总重量g"]
            
            # 检查合并后是否会超过上限
            if merged_weight + sku_total_weight <= self.max_weight_limit:
                merged_weight += sku_total_weight
                
                # 将该SKU整体（所有箱）作为一个合并单元
                sku_data = {
                    "SKU": row["SKU"],
                    "发货数量": row["发货数量"],
                    "箱数": row["箱数"],
                    "单品毛重（g）": row["单品毛重（g）"],
                    "箱子毛重（kg）": row["箱子毛重（kg）"],
                    "单品长（cm）": row.get("单品长（cm）", 0),
                    "单品宽（cm）": row.get("单品宽（cm）", 0),
                    "单品高（cm）": row.get("单品高（cm）", 0),
                    "单份数量": row["发货数量"],  # 整体作为一份
                    "单份总重量": sku_total_weight,
                    "来源": "不拼箱合并",
                }
                merged_skus.append(sku_data)
                used_sku_names.append(row["SKU"])
                
                print(f"   ✅ 合并 SKU {row['SKU']}（{row['箱数']}箱，{sku_total_weight:.0f}g），累计重量: {merged_weight:.0f}g")
                
                # 检查是否已达标
                if merged_weight >= self.min_weight_limit:
                    print(f"✅ 尾组合并成功，最终重量: {merged_weight:.0f}g")
                    break
        
        # 检查最终是否达标
        if merged_weight < self.min_weight_limit:
            print(f"⚠️ 合并后重量 {merged_weight:.0f}g 仍不达标，保持原状")
            return groups_list, invoice_df
        
        # 更新分组
        new_group_name = f"{last_group_name}(混装)" if used_sku_names else last_group_name
        groups_list[-1] = (new_group_name, merged_skus, merged_weight)
        
        # 从发货单中移除已被使用的不拼箱SKU
        updated_invoice_df = invoice_df[~invoice_df["SKU"].isin(used_sku_names)].copy()
        
        print(f"   已将 {len(used_sku_names)} 个不拼箱SKU移出单独装箱: {', '.join(str(s) for s in used_sku_names)}")
        
        return groups_list, updated_invoice_df


    def group_skus_by_weight(self, df, min_weight, max_weight):
        """根据单份总重量组合SKU，每个小组总重量尽量落在范围内
        
        策略：
        1) 先用“降序 + best-fit”填充每组，尽量接近 max_weight
        2) 再对低于 min_weight 的小组做再平衡（搬运/合并）
        """
        sku_list = df.to_dict('records')
        sku_list.sort(key=lambda x: x['单份总重量'], reverse=True)

        groups = []  # [{'skus': [...], 'weight': ...}]
        while sku_list:
            current = sku_list.pop(0)
            group_skus = [current]
            group_weight = current['单份总重量']

            while True:
                remaining = max_weight - group_weight
                if remaining <= 0:
                    break

                best_idx = None
                best_weight = -1
                for i, sku in enumerate(sku_list):
                    w = sku['单份总重量']
                    if w <= remaining and w > best_weight:
                        best_weight = w
                        best_idx = i
                        if w == remaining:
                            break

                if best_idx is None:
                    break

                chosen = sku_list.pop(best_idx)
                group_skus.append(chosen)
                group_weight += chosen['单份总重量']

            groups.append({'skus': group_skus, 'weight': group_weight})

        # 再平衡：尝试让低于 min_weight 的组达标
        max_rebalance_rounds = 100
        rounds = 0
        while rounds < max_rebalance_rounds:
            rounds += 1
            under_idx = next((i for i, g in enumerate(groups) if g['weight'] < min_weight), None)
            if under_idx is None:
                break

            receiver = groups[under_idx]
            best_move = None
            for j, donor in enumerate(groups):
                if j == under_idx:
                    continue
                for k, sku in enumerate(donor['skus']):
                    w = sku['单份总重量']
                    if receiver['weight'] + w > max_weight:
                        continue
                    if donor['weight'] - w < min_weight:
                        continue
                    new_weight = receiver['weight'] + w
                    key = (0 if new_weight >= min_weight else 1, abs(self.target_weight - new_weight))
                    if best_move is None or key < best_move['key']:
                        best_move = {'donor_idx': j, 'sku_idx': k, 'key': key}

            if best_move:
                donor = groups[best_move['donor_idx']]
                sku = donor['skus'].pop(best_move['sku_idx'])
                donor['weight'] -= sku['单份总重量']
                receiver['skus'].append(sku)
                receiver['weight'] += sku['单份总重量']
                continue

            # 如果无法搬运，则尝试合并整个小组
            best_merge_idx = None
            best_merge_key = None
            for j, other in enumerate(groups):
                if j == under_idx:
                    continue
                combined = receiver['weight'] + other['weight']
                if combined > max_weight:
                    continue
                key = (0 if combined >= min_weight else 1, abs(self.target_weight - combined))
                if best_merge_key is None or key < best_merge_key:
                    best_merge_key = key
                    best_merge_idx = j

            if best_merge_idx is not None:
                other = groups.pop(best_merge_idx)
                receiver['skus'].extend(other['skus'])
                receiver['weight'] += other['weight']
                continue

            # 无法再平衡，退出
            break

        groups_list = []
        for idx, g in enumerate(groups, start=1):
            groups_list.append((f"第{idx}组", g['skus'], g['weight']))

        return groups_list

    def find_optimal_grouping(self, df):
        """寻找最优分组，动态调整范围直到找到合适的分组或达到极限"""
        min_weight, max_weight = self.min_weight, self.max_weight
        best_groups, best_score, best_min, best_max = None, float('inf'), 0, 0
        
        while min_weight >= self.min_weight_limit and max_weight <= self.max_weight_limit:
            groups_list = self.group_skus_by_weight(df, min_weight, max_weight)
            print(groups_list)
            
            out_of_range_count = sum(1 for _, _, w in groups_list if w < min_weight or w > max_weight)
            if out_of_range_count == 0:
                return groups_list, min_weight, max_weight
            
            avg_deviation = sum(abs(w - self.target_weight) for _, _, w in groups_list) / len(groups_list) if groups_list else 0
            score = out_of_range_count * 1000 + avg_deviation
            
            if score < best_score:
                best_score, best_groups, best_min, best_max = score, groups_list, min_weight, max_weight
            
            min_weight -= self.weight_step
            max_weight += self.weight_step
        
        return best_groups, best_min, best_max



    def calculate_volumes(self, merge_summary):
        """计算单品体积、单份体积和单箱体积"""
        if missing_columns := [col for col in self.REQUIRED_COLUMNS if col not in merge_summary.columns]:
            print(f"❌ 缺少必要列，无法计算体积: {missing_columns}")
            return merge_summary
        
        if '总体积（m³）-箱子' in merge_summary.columns and '单箱数量' in merge_summary.columns:
            merge_summary['单品体积'] = merge_summary['总体积（m³）-箱子'] * UnitConverter.M3_TO_CM3 / merge_summary['单箱数量']
        else:
            merge_summary['单品体积'] = merge_summary['单品长（cm）'] * merge_summary['单品宽（cm）'] * merge_summary['单品高（cm）']
            
        merge_summary['单份体积'] = merge_summary['单品体积'] * merge_summary['单份数量']
        
        merge_summary = merge_summary.merge(
            merge_summary.groupby('小组名称')['单份体积'].sum().reset_index().rename(columns={'单份体积': '单箱体积'}),
            on='小组名称', how='left'
        )
        
        print("✅ 体积计算完成：\n   - 单品体积：单品长×宽×高\n   - 单份体积：单品体积×单份数量\n   - 单箱体积：小组内所有单份体积合计")
        return merge_summary

    def calculate_box_dimensions(self, merge_summary):
        """计算单箱理论尺寸"""
        if '单箱体积' not in merge_summary.columns:
            print("❌ 缺少必要列：'单箱体积'，无法计算单箱理论尺寸")
            return merge_summary
        
        print("📐 正在计算单箱理论尺寸...")
        dimension_data = []
        
        for group_name in merge_summary['小组名称'].unique():
            box_volume = merge_summary[merge_summary['小组名称'] == group_name].iloc[0]['单箱体积']
            box_length, box_width = self.initial_box_length, self.initial_box_width
            box_height = box_volume / (box_length * box_width)
            
            for _ in range(self.max_iteration):
                if all(self.min_box_dimension <= dim <= self.max_box_dimension 
                       for dim in [box_length, box_width, box_height]):
                    break
                
                if box_height < self.min_box_dimension:
                    box_length = max(self.min_box_dimension, box_length - self.dimension_step) if box_length > self.min_box_dimension else box_length
                    box_width = max(self.min_box_dimension, box_width - self.dimension_step) if box_length <= self.min_box_dimension + 1 and box_width > self.min_box_dimension else box_width
                elif box_height > self.max_box_dimension:
                    box_length = min(self.max_box_dimension, box_length + self.dimension_step) if box_length < self.max_box_dimension else box_length
                    box_width = min(self.max_box_dimension, box_width + self.dimension_step) if box_length >= self.max_box_dimension and box_width < self.max_box_dimension else box_width
                
                box_height = box_volume / (box_length * box_width)
            
            dimension_data.append({'小组名称': group_name, '单箱理论长': box_length, '单箱理论宽': box_width, '单箱理论高': box_height})
            print(f"   - {group_name}: 长={box_length:.1f}cm, 宽={box_width:.1f}cm, 高={box_height:.1f}cm")
        
        merge_summary = merge_summary.merge(pd.DataFrame(dimension_data), on='小组名称', how='left')
        print("✅ 单箱理论尺寸计算完成")
        return merge_summary

    def _align_and_merge(self, data, merge_summary):
        """对齐列并合并数据"""
        if merge_summary.empty:
            return data
        
        for col in merge_summary.columns:
            if col not in data.columns:
                data[col] = None
        data = data[merge_summary.columns]
        
        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', category=FutureWarning)
            return pd.concat([merge_summary, data], ignore_index=True)

    def _add_special_data_to_summary(self, df, merge_summary, group_name, remark_filter):
        """通用方法：将特殊数据添加到拼箱计算结果中"""
        special_data = df[remark_filter].copy()
        
        if special_data.empty:
            return merge_summary
        
        # 调整箱数和发货数量（仅用于拆箱数据）
        if group_name == '拆箱单独装':
            for index, row in special_data.iterrows():
                original_box_count = row['箱数']
                special_data.at[index, '箱数'] = original_box_count - 1
                if original_box_count > 0:
                    new_quantity = row['发货数量'] - row['发货数量'] / original_box_count
                    special_data.at[index, '发货数量'] = new_quantity
                    print(f"   - SKU {row['SKU']}: 箱数 {original_box_count} → {original_box_count - 1}, 发货数量 {row['发货数量']} → {new_quantity}")
        
        special_data['小组名称'] = group_name
        special_data['单箱重量'] = special_data['箱子毛重（kg）'] * UnitConverter.KG_TO_G
        special_data['单箱体积'] = special_data['总体积（m³）-箱子'] * UnitConverter.M3_TO_CM3
        special_data['单箱理论长'] = special_data['箱子长度（cm）']
        special_data['单箱理论宽'] = special_data['箱子宽度（cm）']
        special_data['单箱理论高'] = special_data['箱子高度（cm）']
        special_data['单份数量'] = special_data['单箱数量']
        special_data['单份总重量'] = special_data['单份数量'] * special_data['单品毛重（g）']
        special_data['单品体积'] = special_data['单品长（cm）'] * special_data['单品宽（cm）'] * special_data['单品高（cm）']
        special_data['单份体积'] = special_data['单品体积'] * special_data['单份数量']
        
        print(f"✅ 已添加 {len(special_data)} 个{group_name}的SKU（使用原数据尺寸和重量）")
        return self._align_and_merge(special_data, merge_summary)

    def add_no_remark_data_to_summary(self, df, merge_summary):
        """将全部数据中备注为空并且不拼箱的SKU添加到拼箱计算结果中"""
        return self._add_special_data_to_summary(
            df, merge_summary, '单独装箱',
            (df['备注'].isna() | (df['备注'] == '')) & (df['是否拼箱'] == '不拼箱')
        )



    def validate_data_consistency(self, original_df, result_df):
        """校验原始数据和拼箱计算结果的发货数量和箱数是否一致"""
        if result_df.empty:
            print("ℹ️  拼箱计算结果为空，跳过数据一致性校验")
            return True
        
        original_total_quantity = original_df['发货数量'].sum()
        original_total_boxes = original_df['箱数'].sum()
        result_total_quantity = result_df['发货数量'].sum()
        result_total_boxes = result_df['箱数'].sum()
        
        quantity_match = abs(original_total_quantity - result_total_quantity) < 0.01
        boxes_match = abs(original_total_boxes - result_total_boxes) < 0.01
        
        if not (quantity_match and boxes_match):
            raise ValueError("拼箱结果与原数据不一致，请检查逻辑！")
        
        return True

    def _apply_cell_style(self, cell, border=True, fill=None, font=None, alignment=None):
        """辅助方法：应用单元格样式"""
        if border:
            cell.border = StyleManager.THIN_BORDER
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment

    def _merge_and_style_cells(self, ws, cell_range, value='', fill=None, alignment=None, font=None):
        """辅助方法：合并单元格并应用样式"""
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(':')[0]]
        if value:
            cell.value = value
        self._apply_cell_style(cell, fill=fill, alignment=alignment, font=font)
        
        # 为合并区域的所有单元格设置边框
        start_cell, end_cell = cell_range.split(':')
        from openpyxl.utils import column_index_from_string, get_column_letter
        start_col = column_index_from_string(start_cell[0])
        start_row = int(start_cell[1:])
        end_col = column_index_from_string(end_cell[0])
        end_row = int(end_cell[1:])
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = StyleManager.THIN_BORDER
                if fill:
                    ws.cell(row=row, column=col).fill = fill

    def safe_save_data(self, merge_summary):
        """安全保存数据，如果主要路径失败则尝试备用路径"""
        with pd.ExcelWriter(self.output_file_path, engine='openpyxl') as writer:
            data_to_save = merge_summary if not merge_summary.empty else pd.DataFrame({"提示": ["没有需要拼箱的SKU"]})
            data_to_save.to_excel(writer, sheet_name='拼箱计算结果', index=False)
            
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # 如果有小组名称列，则进行单元格合并
                if not merge_summary.empty and '小组名称' in merge_summary.columns:
                    header_row = [cell.value for cell in worksheet[1]]
                    merge_col_indices = {col_name: header_row.index(col_name) + 1 
                                        for col_name in self.MERGE_COLUMNS if col_name in header_row}
                    
                    if '小组名称' in merge_col_indices:
                        self._merge_group_cells(worksheet, merge_col_indices)
                
                self._apply_worksheet_formatting(worksheet, merge_summary)

    def _merge_group_cells(self, worksheet, merge_col_indices):
        """合并小组单元格"""
        group_col_idx = merge_col_indices['小组名称']
        current_group = None
        start_row = 2
        
        for row_idx in range(2, worksheet.max_row + 2):
            cell_value = worksheet.cell(row=row_idx, column=group_col_idx).value if row_idx <= worksheet.max_row else None
            
            if cell_value != current_group:
                if (current_group is not None and row_idx - start_row > 1 
                    and not self._is_special_group_name(current_group)):
                    for col_idx in merge_col_indices.values():
                        worksheet.merge_cells(start_row=start_row, start_column=col_idx, 
                                            end_row=row_idx - 1, end_column=col_idx)
                current_group = cell_value
                start_row = row_idx

    def _apply_worksheet_formatting(self, worksheet, merge_summary):
        """应用工作表格式"""
        # 遍历所有单元格添加边框和对齐方式
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                      min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = StyleManager.THIN_BORDER
                
                # 设置对齐方式
                if not merge_summary.empty and '小组名称' in merge_summary.columns:
                    header_row = [c.value for c in worksheet[1]]
                    
                    group_name_col_idx = header_row.index('小组名称') + 1 if '小组名称' in header_row else None
                    if group_name_col_idx:
                        row_group_name = worksheet.cell(row=cell.row, column=group_name_col_idx).value
                    else:
                        row_group_name = None
                    
                    is_merge_column = (cell.column <= len(header_row) 
                                      and header_row[cell.column - 1] in self.MERGE_COLUMNS)
                    is_data_row = cell.row > 1
                    is_special_group = self._is_special_group_name(row_group_name)
                    
                    cell.alignment = (StyleManager.CENTER_ALIGNMENT 
                                    if is_merge_column and is_data_row and not is_special_group 
                                    else StyleManager.LEFT_ALIGNMENT)
                else:
                    cell.alignment = StyleManager.LEFT_ALIGNMENT
        
        # 自动调整列宽
        for column in worksheet.columns:
            column_letter = column[0].column_letter
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            worksheet.column_dimensions[column_letter].width = max(max_length + 2, 10)

    def _unmerge_cell_if_needed(self, ws, cell):
        """辅助方法：解除合并单元格并返回新单元格"""
        if isinstance(cell, MergedCell):
            for merged_range in list(ws.merged_cells.ranges):
                if cell.coordinate in merged_range:
                    ws.unmerge_cells(str(merged_range))
                    break
            return ws.cell(row=cell.row, column=cell.column)
        return cell
    
    def _write_cell_value(self, ws, row, col, value, alignment='center'):
        """辅助方法：写入单元格数据并设置对齐"""
        cell = self._unmerge_cell_if_needed(ws, ws.cell(row=row, column=col))
        cell.value = value
        cell.alignment = Alignment(horizontal=alignment, vertical='center')

    def _normalize_sku_value(self, value):
        """规范化SKU值，避免数字SKU在Excel读回时带 .0 导致无法匹配。"""
        if pd.isna(value):
            return ""
        if isinstance(value, (int, np.integer)):
            return str(int(value))
        if isinstance(value, (float, np.floating)):
            if float(value).is_integer():
                return str(int(value))
            return str(value).strip()
        text = str(value).strip()
        if text.endswith(".0"):
            number_part = text[:-2]
            if number_part.replace("-", "", 1).isdigit():
                return number_part
        return text

    def _infer_actual_box_count(self, row):
        """优先通过发货数量和单份数量反推实际箱数，回退到箱数字段。"""
        quantity = row.get('发货数量')
        units_per_box = row.get('单份数量')
        if pd.notna(quantity) and pd.notna(units_per_box) and units_per_box not in (0, ""):
            try:
                value = float(quantity) / float(units_per_box)
                if value > 0:
                    rounded = int(round(value))
                    if abs(value - rounded) < 1e-6:
                        return rounded
            except (TypeError, ValueError, ZeroDivisionError):
                pass

        box_count = row.get('箱数')
        if pd.notna(box_count):
            try:
                rounded = int(round(float(box_count)))
                if rounded > 0:
                    return rounded
            except (TypeError, ValueError):
                pass
        return 0

    def _add_actual_box_count_column(self, df):
        """为结果表补充实际箱数列。"""
        if df.empty:
            df["实际箱数"] = []
            return df
        result = df.copy()
        result["实际箱数"] = result.apply(self._infer_actual_box_count, axis=1)
        return result

    def _get_output_columns(self, include_warehouse=False):
        """获取拼箱结果表输出列顺序。"""
        output_columns = [
            'SKU', '品名', '发货数量', '箱数', '单品毛重（g）', '箱子毛重（kg）', '单品长（cm）', '单品宽（cm）', '单品高（cm）',
            '是否拼箱', '备注', '单份数量', '单份总重量', '小组名称', '实际箱数', '单箱重量', '重量差', '单品体积', '单份体积',
            '单箱体积', '单箱理论长', '单箱理论宽', '单箱理论高',
        ]
        if include_warehouse:
            output_columns.insert(1, "发货仓库（单据）")
        return output_columns

    def _determine_units_per_box_for_packing(self, row):
        """确定拼箱分组时每份应使用的数量。"""
        if bool(row.get("_force_packing", False)):
            return row.get("单箱数量")
        return row.get("发货数量") / self.boxes_per_share

    def _group_skus_by_compatible_box_count(self, lcl_df):
        """按实际箱数隔离拼箱候选，避免不同箱数计划混入同一组。"""
        if lcl_df.empty:
            return []

        working_df = lcl_df.copy()
        working_df["_actual_box_count_key"] = working_df.apply(self._infer_actual_box_count, axis=1)

        groups = []
        for _, compatible_df in working_df.groupby("_actual_box_count_key", sort=False, dropna=False):
            compatible_df = compatible_df.drop(columns=["_actual_box_count_key"])
            part_groups, _, _ = self.find_optimal_grouping(compatible_df)
            groups.extend(part_groups)

        return self._renumber_groups(groups)

    def _renumber_groups(self, groups):
        """统一重排组名，避免分桶分组后出现重复的第1组。"""
        renumbered = []
        for index, (group_name, skus, group_weight) in enumerate(groups, start=1):
            suffix = "(混装)" if "(混装)" in str(group_name) else ""
            renumbered.append((f"第{index}组{suffix}", skus, group_weight))
        return renumbered

    def insert_data_to_amazon_packaging(self, amazon_packaging_file_path, merge_summary):
        """将拼箱数据插入到Amazon包装信息文件中"""
        print(f"\n📦 开始处理Amazon包装信息...")
        
        # 读取并分组数据
        package_df = pd.read_excel(amazon_packaging_file_path, sheet_name="包装箱包装信息", skiprows=4)
        nan_index = package_df[package_df['SKU'].isna()].index[0]
        first_group = package_df.iloc[:nan_index][['SKU']]
        
        # 转换单位
        merge_summary_converted = merge_summary[['SKU','发货数量','小组名称','单份数量','箱数','单箱重量','单箱理论宽','单箱理论长','单箱理论高']].copy()
        merge_summary_converted['_actual_box_count'] = merge_summary_converted.apply(self._infer_actual_box_count, axis=1)
        for col, factor in [('单箱重量', UnitConverter.G_TO_LB), 
                           ('单箱理论宽', UnitConverter.CM_TO_INCH), 
                           ('单箱理论长', UnitConverter.CM_TO_INCH), 
                           ('单箱理论高', UnitConverter.CM_TO_INCH)]:
            merge_summary_converted[col] *= factor
        
        # 统一SKU列的数据类型为字符串，避免merge时类型不匹配错误
        first_group['SKU'] = first_group['SKU'].apply(self._normalize_sku_value)
        merge_summary_converted['SKU'] = merge_summary_converted['SKU'].apply(self._normalize_sku_value)
        
        # 合并数据并准备展开
        first_group_merged = first_group.merge(merge_summary_converted, on='SKU', how='left')
        result_df = first_group_merged[['SKU','小组名称','单份数量','_actual_box_count']].copy()
        
        # 数据展开和转换
        result_df_expanded = self._expand_columns_by_group(result_df)
        specification_df = first_group_merged[['小组名称','_actual_box_count','单箱重量','单箱理论宽','单箱理论长','单箱理论高']].copy()
        specification_df['单箱重量'] = specification_df['单箱重量'].round(2)
        specification_df['单箱理论宽'] = specification_df['单箱理论宽'].round(2)
        specification_df['单箱理论长'] = specification_df['单箱理论长'].round(2)
        specification_df['单箱理论高'] = specification_df['单箱理论高'].round(2)
        specification_transposed = self._transpose_specification(specification_df)

        # 加载Excel文件并插入数据
        wb = load_workbook(amazon_packaging_file_path)
        ws = wb["包装箱包装信息"]
        start_col = 13
        total_box_count = self._get_total_box_count(merge_summary)
        self._update_amazon_box_headers(ws, start_col, total_box_count)
        self._update_amazon_total_box_count(ws, total_box_count)
        self._update_amazon_packed_quantity(ws)
        self._clear_amazon_dynamic_box_area(ws, start_col, 6, nan_index + 4)
        
        # 插入数据
        columns_to_insert = [col for col in result_df_expanded.columns if col not in ['SKU', '小组名称']]
        for row_idx, (_, row_data) in enumerate(result_df_expanded.iterrows()):
            for col_idx, col_name in enumerate(columns_to_insert):
                self._write_cell_value(ws, 6 + row_idx, start_col + col_idx, row_data[col_name])
        print(f"✅ 已插入 {len(result_df_expanded)} 行数据到M6")
        
        spec_start_row = nan_index + 8
        self._clear_amazon_spec_area(ws, start_col, spec_start_row, len(specification_transposed))
        for row_idx, (_, row_data) in enumerate(specification_transposed.iterrows()):
            for col_idx, value in enumerate(row_data):
                self._write_cell_value(ws, spec_start_row + row_idx, start_col + col_idx, value)
        print(f"✅ 已插入 {len(specification_transposed)} 行specification数据到M{spec_start_row}")
        
        wb.save(amazon_packaging_file_path)
        print(f"✅ 数据已成功插入并保存到：{amazon_packaging_file_path}")

    def _get_total_box_count(self, merge_summary):
        """计算最终Amazon文件需要展示的总箱数。"""
        group_box_counts = (
            merge_summary.dropna(subset=['小组名称'])
            .loc[~merge_summary['小组名称'].astype(str).apply(self._is_special_group_name)]
            .copy()
        )
        if group_box_counts.empty:
            return 0
        group_box_counts['_actual_box_count'] = group_box_counts.apply(self._infer_actual_box_count, axis=1)
        return int(group_box_counts.groupby('小组名称')['_actual_box_count'].max().sum())

    def _update_amazon_total_box_count(self, ws, total_box_count):
        """更新包装箱总数单元格。"""
        self._write_cell_value(ws, 3, 13, total_box_count)

    def _update_amazon_box_headers(self, ws, start_col, total_box_count):
        """更新箱数量列标题和包装箱名称行。"""
        for idx in range(total_box_count):
            col = start_col + idx
            self._write_cell_value(ws, 5, col, f"包装箱 {idx + 1} 数量")
            self._write_cell_value(ws, 17, col, f"P1 - B{idx + 1}")

    def _update_amazon_packed_quantity(self, ws):
        """将装箱数量列补齐为预计数量，保持与正确模板一致。"""
        row = 6
        while row <= ws.max_row:
            sku_value = ws.cell(row=row, column=1).value
            if sku_value in (None, ""):
                row += 1
                continue
            expected_quantity = ws.cell(row=row, column=10).value
            self._write_cell_value(ws, row, 11, expected_quantity)
            row += 1

    def _clear_amazon_dynamic_box_area(self, ws, start_col, start_row, end_row):
        """清空SKU行的动态箱数量区域，避免旧值残留。"""
        if end_row < start_row or ws.max_column < start_col:
            return
        for row in range(start_row, end_row + 1):
            for col in range(start_col, ws.max_column + 1):
                self._write_cell_value(ws, row, col, None)

    def _clear_amazon_spec_area(self, ws, start_col, start_row, row_count):
        """清空规格区域旧值，避免箱数缩小时残留历史数据。"""
        if row_count <= 0 or ws.max_column < start_col:
            return
        end_row = start_row + row_count - 1
        for row in range(start_row, end_row + 1):
            for col in range(start_col, ws.max_column + 1):
                self._write_cell_value(ws, row, col, None)
    
    def _expand_columns_by_group(self, df):
        """根据小组名称动态展开列"""
        if '小组名称' not in df.columns:
            return df
        
        unique_groups = sorted(
            g for g in df['小组名称'].dropna().unique()
            if not self._is_special_group_name(g)
        )
        result_df = df[['SKU', '小组名称']].copy()
        
        for group in unique_groups:
            group_rows = df[df['小组名称'] == group]
            group_box_count = int(group_rows['_actual_box_count'].max()) if '_actual_box_count' in group_rows.columns and not group_rows.empty else 0
            for idx in range(1, group_box_count + 1):
                result_df[f'{group}单份数量{idx}'] = df.apply(
                    lambda row: row['单份数量'] if row['小组名称'] == group else '',
                    axis=1
                )
        
        total_columns = len([col for col in result_df.columns if col not in ['SKU', '小组名称']])
        print(f"✅ 列展开完成，{len(unique_groups)}个小组 × {total_columns}列")
        return result_df
    
    def _transpose_specification(self, df):
        """将specification列转行并按小组展开"""
        if '小组名称' not in df.columns:
            return df
        
        unique_groups = sorted(
            g for g in df['小组名称'].dropna().unique()
            if not self._is_special_group_name(g)
        )
        metric_columns = ['单箱重量', '单箱理论宽', '单箱理论长', '单箱理论高']
        result_data = {}
        
        for group in unique_groups:
            group_rows = df[df['小组名称'] == group]
            group_data = group_rows.iloc[0]
            box_count = int(group_rows['_actual_box_count'].max()) if '_actual_box_count' in group_rows.columns else 0
            for i in range(1, box_count + 1):
                result_data[f'{group}{i}'] = [group_data[metric] for metric in metric_columns]
        
        return pd.DataFrame(result_data, index=metric_columns)

    def create_amazon_template(self, template_file_path, merge_summary, template_source=None):
        """基于模板文件创建Amazon发货模板Excel文件并填入数据"""
        print("📝 正在基于模板创建Amazon发货文件...")
        
        template_source = template_source or config.AMAZON_TEMPLATE_SOURCE_FILE
        if not os.path.exists(template_source):
            raise FileNotFoundError(
                f"Amazon模板不存在，请确认路径: {template_source}"
            )

        os.makedirs(os.path.dirname(template_file_path), exist_ok=True)
        shutil.copy2(template_source, template_file_path)
        print(f"✅ 已复制模板文件: {template_source}")
        
        # 加载复制的文件并填写数据
        wb = load_workbook(template_file_path)
        
        # 填写 Create workflow – template 工作表的数据
        ws = wb["Create workflow – template"]
        if not merge_summary.empty:
            self._fill_example_data(ws, merge_summary)
            print(f"✅ 已在模板中填入数据")
        
        wb.save(template_file_path)
        print(f"✅ Amazon发货文件已保存：{template_file_path}")

    def _find_amazon_workflow_header_row(self, ws):
        """定位Amazon模板表头行，兼容含默认owner行的新模板。"""
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=1).value
            if str(value).strip() == "Merchant SKU":
                return row
        return 6

    def _fill_example_data(self, ws, merge_summary):
        """填充示例数据"""
        data_row = self._find_amazon_workflow_header_row(ws) + 1
        for _, row in merge_summary.iterrows():
            # 基本信息
            ws.cell(row=data_row, column=1, value=row.get('SKU', '')).border = StyleManager.THIN_BORDER
            ws.cell(row=data_row, column=2, value=row.get('发货数量', '')).border = StyleManager.THIN_BORDER
            
            # C-D列留空
            for col in range(3, 5):
                ws.cell(row=data_row, column=col, value='').border = StyleManager.THIN_BORDER
            
            # 判断是否是拼箱数据
            group_name = row.get('小组名称', '')
            is_packing = not self._is_special_group_name(group_name)
            
            if is_packing:
                # 拼箱数据：E-J列都留空
                for col in range(5, 11):
                    ws.cell(row=data_row, column=col, value='').border = StyleManager.THIN_BORDER
            else:
                # 单独装箱或拆箱单独装：填入数据
                ws.cell(row=data_row, column=5, value=row.get('单份数量', '')).border = StyleManager.THIN_BORDER
                ws.cell(row=data_row, column=6, value=row.get('箱数', '')).border = StyleManager.THIN_BORDER

                # 尺寸和重量（转换单位）
                dimensions = [
                    (7, row.get('单箱理论长', 0), UnitConverter.CM_TO_INCH),
                    (8, row.get('单箱理论宽', 0), UnitConverter.CM_TO_INCH),
                    (9, row.get('单箱理论高', 0), UnitConverter.CM_TO_INCH),
                    (10, row.get('单箱重量', 0), UnitConverter.G_TO_LB),
                ]
                
                for col, value, factor in dimensions:
                    converted_value = round(value * factor, 2) if value else ''
                    ws.cell(row=data_row, column=col, value=converted_value).border = StyleManager.THIN_BORDER
            
            data_row += 1
        
        print(f"✅ 已填入 {len(merge_summary)} 条数据")

    def process(self):
        """主处理函数 - 执行完整的拼箱处理流程"""
        print("=" * 50)
        print("开始处理拼箱数据...")
        print("=" * 50)
        
        # 读取并预处理数据
        invoice_df = self._load_and_preprocess_data()
        
        # 按发货仓库（单据）拆分处理，同仓库SKU才拼箱
        if "发货仓库（单据）" in invoice_df.columns:
            merge_parts = []
            for _, warehouse_df in invoice_df.groupby("发货仓库（单据）", dropna=False):
                lcl_df = warehouse_df[warehouse_df["是否拼箱"] == "拼箱"].copy()
                if not lcl_df.empty:
                    part = self._process_packing_data(warehouse_df, lcl_df)
                else:
                    part = self._process_no_packing_data(warehouse_df)
                if not part.empty:
                    warehouse_value = warehouse_df["发货仓库（单据）"].iloc[0]
                    if "发货仓库（单据）" not in part.columns:
                        part["发货仓库（单据）"] = warehouse_value
                    warehouse_label = (
                        str(warehouse_value).strip()
                        if pd.notna(warehouse_value) and str(warehouse_value).strip()
                        else "未填写仓库"
                    )
                    if "小组名称" in part.columns:
                        part["小组名称"] = part["小组名称"].apply(
                            lambda name: f"{warehouse_label}-{name}" if str(name).strip() else warehouse_label
                        )
                merge_parts.append(part)
            merge_summary = pd.concat(merge_parts, ignore_index=True) if merge_parts else pd.DataFrame()
        else:
            # 筛选需要拼箱的数据
            lcl_df = invoice_df[invoice_df["是否拼箱"] == "拼箱"].copy()
            if not lcl_df.empty:
                merge_summary = self._process_packing_data(invoice_df, lcl_df)
            else:
                merge_summary = self._process_no_packing_data(invoice_df)
        
        # 整理输出列
        if "重量差" not in merge_summary.columns:
            merge_summary["重量差"] = ""
        if "品名" not in merge_summary.columns:
            merge_summary["品名"] = ""
        merge_summary = self._add_actual_box_count_column(merge_summary)
        output_columns = self._get_output_columns(include_warehouse="发货仓库（单据）" in merge_summary.columns)
        merge_summary = merge_summary[output_columns]
        
        self.safe_save_data(merge_summary)
        
        print("=" * 50)
        print(f"✅ 拼箱处理完成！结果已保存至: {self.output_file_path}")
        print("=" * 50)
        
        return merge_summary

    def _load_and_preprocess_data(self):
        """加载并预处理数据"""
        # 读取发货单详情
        invoice_df = pd.read_excel(self.input_file_path, sheet_name="发货单详情")
        if "发货仓库（单据）" in invoice_df.columns:
            # 合并单元格向下填充
            invoice_df["发货仓库（单据）"] = invoice_df["发货仓库（单据）"].ffill()
        
        # 新增列
        invoice_df["备注"] = ""
        if "品名" not in invoice_df.columns:
            invoice_df["品名"] = ""
        
        # 拆分包装规格
        dimensions = invoice_df['包装规格'].str.split('x', expand=True)
        # print(dimensions)
        invoice_df['单品长（cm）'] = pd.to_numeric(dimensions[0], errors='coerce')
        invoice_df['单品宽（cm）'] = pd.to_numeric(dimensions[1], errors='coerce')
        invoice_df['单品高（cm）'] = pd.to_numeric(dimensions[2], errors='coerce')
        
        # 更改列名
        invoice_df.rename(columns={'发货量': '发货数量', '单品毛重': '单品毛重（g）'}, inplace=True)
        
        # 读取并合并装箱信息
        pack_info_df = pd.read_excel(self.input_file_path, sheet_name="装箱信息")
        
        # 兼容CBM列名
        if 'CBM（m³）-箱子' in pack_info_df.columns:
            if '总体积（m³）-箱子' in pack_info_df.columns:
                pack_info_df = pack_info_df.drop(columns=['总体积（m³）-箱子'])
            pack_info_df.rename(columns={'CBM（m³）-箱子': '总体积（m³）-箱子'}, inplace=True)
            
        pack_info_df = pack_info_df[['SKU','总重量（kg）-箱子','总体积（m³）-箱子','箱子毛重（kg）', '单箱数量', '箱子长度（cm）', '箱子宽度（cm）', '箱子高度（cm）']]
        invoice_df = invoice_df.merge(pack_info_df, on='SKU', how='left')
        # 重新计算箱数
        invoice_df['箱数'] = invoice_df['发货数量'] / invoice_df['单箱数量']
        invoice_df = self._apply_non_integer_large_box_repacking(invoice_df)
        
        invoice_df = self._split_box_rows(invoice_df)
        invoice_df['是否拼箱'] = invoice_df.apply(self._determine_packing_status, axis=1)

        base_columns = [
            'SKU', '品名', '发货数量', '箱数', '单品毛重（g）', '箱子毛重（kg）', '单品长（cm）', '单品宽（cm）', '单品高（cm）',
            '是否拼箱', '备注', '单箱数量', '总重量（kg）-箱子', '总体积（m³）-箱子', '箱子长度（cm）', '箱子宽度（cm）', '箱子高度（cm）',
        ]
        if "_force_packing" in invoice_df.columns:
            base_columns.append("_force_packing")
        if "发货仓库（单据）" in invoice_df.columns:
            base_columns.insert(1, "发货仓库（单据）")
        invoice_df = invoice_df[base_columns]
        return invoice_df

    def _determine_packing_status(self, row_or_box_count):
        if isinstance(row_or_box_count, pd.Series):
            if bool(row_or_box_count.get("_force_packing", False)):
                return "拼箱"
            box_count = row_or_box_count.get("箱数")
        else:
            box_count = row_or_box_count
        if pd.isna(box_count):
            return "拼箱"
        if float(box_count).is_integer() and box_count >= self.min_boxes_for_split:
            return "不拼箱"
        return "拼箱"

    def _split_box_rows(self, df):
        """处理箱数，不再拆分整数和小数部分，直接保持原样"""
        expanded_rows = []
        for _, row in df.iterrows():
            box_count = row.get('箱数')
            per_box_quantity = row.get('单箱数量')

            row_copy = row.copy()
            # 如果发货数量缺失，用箱数计算
            if pd.isna(row_copy.get('发货数量')) and not pd.isna(box_count) and not pd.isna(per_box_quantity):
                row_copy['发货数量'] = box_count * per_box_quantity
            expanded_rows.append(row_copy)

        return pd.DataFrame(expanded_rows, columns=df.columns)

    def _process_packing_data(self, invoice_df, lcl_df):
        """处理需要拼箱的数据"""
        # 在拼箱的情况下，单品毛重使用新规则：箱子毛重（kg）*1000 /单箱数量
        lcl_df["单品毛重（g）"] = lcl_df["箱子毛重（kg）"] * 1000 / lcl_df["单箱数量"]
        
        # 命中自动重算规则的SKU，直接按新单箱数量作为单份数量；其余保持原5份制逻辑。
        lcl_df["单份数量"] = lcl_df.apply(self._determine_units_per_box_for_packing, axis=1)
        # 计算单份总重量
        lcl_df["单份总重量"] = lcl_df["单份数量"] * lcl_df["单品毛重（g）"]

        print(lcl_df)
        used_groups = self._group_skus_by_compatible_box_count(lcl_df)
    
        # 使用新的尾组合并逻辑：将尾组与不拼箱SKU整体合并
        groups, invoice_df = self.merge_tail_with_non_packing(invoice_df, used_groups)
        groups = self._renumber_groups(groups)
        
        grouped_data = [{**sku, '小组名称': group_name, '单箱重量': group_weight} 
                       for group_name, skus, group_weight in groups for sku in skus]
        
        merge_summary_grouped = pd.DataFrame(grouped_data).sort_values('小组名称')
        merge_summary = self.calculate_volumes(merge_summary_grouped)
        merge_summary = self.calculate_box_dimensions(merge_summary)
        merge_summary = self.add_no_remark_data_to_summary(invoice_df, merge_summary)
        self._validate_hangzhou_kongqi_volume(merge_summary)
        
        # 打印分组信息
        for group_name, skus, group_weight in groups:
            in_range = "✅" if self.min_weight <= group_weight <= self.max_weight else "⚠️"
            print(f"   - {group_name}: {len(skus)} 个SKU, 总重量 {group_weight:.2f} g {in_range}")
        
        if (out_of_range_count := sum(1 for _, _, w in groups if not (self.min_weight <= w <= self.max_weight))) > 0:
            print(f"⚠️  注意：有 {out_of_range_count} 个小组的重量不在{self.min_weight}-{self.max_weight}g的理想范围内")
        
        self.validate_data_consistency(invoice_df, merge_summary)
        
        return self._add_actual_box_count_column(merge_summary)

    def _validate_hangzhou_kongqi_volume(self, merge_summary):
        """校验杭州虚拟仓且品名含开孔器的拼箱组单箱体积上限"""
        if merge_summary.empty:
            return

        required_cols = {"发货仓库（单据）", "品名", "单箱体积", "小组名称"}
        if not required_cols.issubset(set(merge_summary.columns)):
            return

        df = merge_summary.copy()
        df["发货仓库（单据）"] = df["发货仓库（单据）"].astype(str).str.strip()
        df["品名"] = df["品名"].astype(str)

        target_mask = (df["发货仓库（单据）"] == "杭州虚拟仓") & df["品名"].str.contains("开孔器", na=False)
        if "小组名称" in df.columns:
            target_mask &= ~df["小组名称"].astype(str).apply(self._is_special_group_name)

        if not target_mask.any():
            return

        target_groups = df.loc[target_mask, "小组名称"].dropna().unique().tolist()
        if not target_groups:
            return

        volume_series = pd.to_numeric(df.loc[df["小组名称"].isin(target_groups), "单箱体积"], errors="coerce")
        volume_by_group = volume_series.groupby(df.loc[df["小组名称"].isin(target_groups), "小组名称"]).max()

        exceeded = volume_by_group[volume_by_group > 50000]
        if exceeded.empty:
            return

        details = ", ".join(f"{name}: {vol:.0f}" for name, vol in exceeded.items())
        raise ValueError(
            "拼箱结果校验失败：杭州虚拟仓且品名含“开孔器”的拼箱组，单箱体积超过50000。"
            f"超标组: {details}"
        )

    def _process_no_packing_data(self, invoice_df):
        """处理不需要拼箱的数据"""
        merge_summary = self.add_no_remark_data_to_summary(invoice_df, pd.DataFrame())
        return merge_summary
