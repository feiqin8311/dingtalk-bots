# !/usr/bin/env python
# -*- coding: utf-8 -*-
"""
File helper utilities for LCL bot.
Handles Excel processing for customs clearance files.
"""

import os
import logging
from typing import Optional, List, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 配置日志
logger = logging.getLogger("Bot.file_helper")

def get_cell_value(
    file_path: str,
    sheet_name: str,
    cell_coordinate: str,
) -> Optional[str]:
    """
    Get value from a specific cell in an Excel file.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the sheet (e.g., "运单信息").
        cell_coordinate: Cell coordinate (e.g., "M8").
        
    Returns:
        String value of the cell, or None if not found/error.
    """
    try:
        # data_only=True gets the value, not the formula
        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in {file_path}")
            return None
        
        sheet = wb[sheet_name]
        value = sheet[cell_coordinate].value
        return str(value).strip() if value is not None else None
    except Exception as e:
        logger.error(f"Error reading {file_path}: {e}")
        return None


def rename_file_by_cell_value(
    file_path: str,
    sheet_name: str = "运单信息",
    cell_coordinate: str = "M8",
) -> Tuple[bool, str]:
    """
    Rename an Excel file based on a cell value.
    
    Args:
        file_path: Path to the source file.
        sheet_name: Sheet to read from.
        cell_coordinate: Cell to read from (e.g., "M8").
        
    Returns:
        (Success, Message/New Path)
    """
    if not os.path.exists(file_path):
        return False, f"文件不存在: {file_path}"
        
    try:
        # Get new name from cell
        new_name_base = get_cell_value(file_path, sheet_name, cell_coordinate)
        if not new_name_base:
            return False, f"无法获取 {sheet_name}!{cell_coordinate} 的值"
            
        # Helper to clean filename
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            new_name_base = new_name_base.replace(char, '_')
            
        # Construct new path
        dir_name = os.path.dirname(file_path)
        extension = os.path.splitext(file_path)[1]
        
        # Original logic: Rename with just the value (e.g., value.xlsx)
        new_filename = f"{new_name_base}{extension}"
        new_path = os.path.join(dir_name, new_filename)
        
        # Check if same file
        if os.path.abspath(new_path) == os.path.abspath(file_path):
            return True, new_path
        
        # Handle duplicates
        if os.path.exists(new_path):
            base, ext = os.path.splitext(new_filename)
            counter = 1
            while os.path.exists(new_path):
                 new_path = os.path.join(dir_name, f"{base}_{counter}{ext}")
                 counter += 1
        
        os.rename(file_path, new_path)
        return True, new_path
        
    except Exception as e:
        return False, f"重命名异常: {str(e)}"


def _format_dimension_value(value) -> Optional[str]:
    """Normalize dimension cell value to string without trailing .0."""
    if value is None:
        return None
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip("0").rstrip(".") if "." in str(value) else str(value)
    text = str(value).strip()
    return text or None


def update_dimensions_and_trim_columns(
    file_path: str,
    sheet_name: str = "运单信息",
    start_row: int = 12,
) -> Tuple[bool, str]:
    """
    Fill column C with "P*Q*R" (长*宽*高) from row start_row, then delete columns after P.

    Args:
        file_path: Path to the Excel file.
        sheet_name: Sheet to update.
        start_row: Starting row for filling.

    Returns:
        (success, message)
    """
    try:
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            return False, f"未找到工作表: {sheet_name}"
        ws = wb[sheet_name]

        max_row = ws.max_row or start_row
        center_alignment = Alignment(horizontal="center", vertical="center")
        for row in range(start_row, max_row + 1):
            p_val = _format_dimension_value(ws.cell(row=row, column=16).value)
            q_val = _format_dimension_value(ws.cell(row=row, column=17).value)
            r_val = _format_dimension_value(ws.cell(row=row, column=18).value)
            i_val = _format_dimension_value(ws.cell(row=row, column=9).value)
            if i_val is not None:
                ws.cell(row=row, column=10, value=1)
            if not (p_val and q_val and r_val):
                continue
            c_cell = ws.cell(row=row, column=3, value=f"{p_val}*{q_val}*{r_val}")
            c_cell.alignment = center_alignment

        # Delete columns from P (column 16) onwards
        if ws.max_column and ws.max_column >= 16:
            ws.delete_cols(16, ws.max_column - 15)

        wb.save(file_path)
        return True, "尺寸已写入C列并删除P列之后的列"
    except Exception as e:
        return False, f"尺寸处理失败: {e}"


def process_customs_files(directory: str) -> List[str]:
    """
    Batch process all Excel files in a directory:
    Rename them based on M8 in '运单信息', then fill C column with P*Q*R and trim columns after P.
    
    Args:
        directory: Directory containing the downloaded files.
        
    Returns:
        List of result messages.
    """
    results = []
    if not os.path.exists(directory):
        return ["目录不存在"]
        
    logger.info(f"Processing files in {directory}")
    
    # List files first to avoid processing newly renamed files if iterating directly
    files_to_process = []
    for filename in os.listdir(directory):
        # Only process xlsx/xls and ignore temp files (~$)
        if filename.lower().endswith(('.xlsx', '.xls')) and not filename.startswith('~$'):
             files_to_process.append(filename)
             
    if not files_to_process:
        return ["目录中没有找到 Excel 文件"]

    for filename in files_to_process:
        file_path = os.path.join(directory, filename)
        
        # Skip if file no longer exists (e.g. somehow renamed/moved externally)
        if not os.path.exists(file_path):
            continue
            
        size_ok, size_msg = update_dimensions_and_trim_columns(file_path)
        success, msg = rename_file_by_cell_value(
            file_path,
            sheet_name="运单信息",
            cell_coordinate="M8"
        )

        if success:
            new_name = os.path.basename(msg)
            if size_ok:
                results.append(f"✅ {filename} -> {new_name} | {size_msg}")
            else:
                results.append(f"⚠️ {filename} -> {new_name} | {size_msg}")
        else:
            results.append(f"❌ {filename} 失败: {msg}")
            
    return results
