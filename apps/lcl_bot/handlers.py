# !/usr/bin/env python
# -*- coding: utf-8 -*-
"""
消息处理器 - 钉钉机器人消息和工作流处理
"""

import logging
import os
import re
import shutil
import time
from typing import List, Optional

import pandas as pd
from dingtalk_stream import AckMessage
import dingtalk_stream

from . import config
from .lingxing_helper import delete_shipment_list
from .file_helper import process_customs_files
from .processor import PackingBoxProcessor
from .shipment_registry import extract_shipment_info
from .state_manager import WorkflowState, get_state_manager
from .utils import (
    DingTalkAPI,
    cleanup_old_files,
)


class WorkflowBotHandler(dingtalk_stream.ChatbotHandler):
    """工作流机器人处理器 - 处理物流和运营的完整工作流"""
    
    def __init__(self, logger: logging.Logger = None, config_obj=None):
        """初始化处理器。config_obj 可选（logistics monorepo 注入，凭证已由 env 加载）。"""
        super(dingtalk_stream.ChatbotHandler, self).__init__()
        self.logger = logger or logging.getLogger(__name__)
        self._external_config = config_obj
        self.state_manager = get_state_manager()
        client_id = config.CLIENT_ID
        client_secret = config.CLIENT_SECRET
        if config_obj is not None:
            client_id = getattr(config_obj, "client_id", None) or client_id
            client_secret = getattr(config_obj, "client_secret", None) or client_secret
        if not client_id or not client_secret:
            self.logger.warning("lcl_bot missing DingTalk credentials; file send may fail")
        self.dingtalk_api = DingTalkAPI(client_id or "", client_secret or "")

        # 定期清理工作目录（保留模板源文件夹）
        cleanup_old_files(config.EXCEL_FILES_DIR, days=7, exclude=["sample_data"])
    
    async def process(self, callback: dingtalk_stream.CallbackMessage):
        """主消息处理入口"""
        try:
            incoming_message = dingtalk_stream.ChatbotMessage.from_dict(callback.data)
            
            # 获取发送者信息
            sender_id = incoming_message.sender_id
            sender_staff_id = getattr(incoming_message, 'sender_staff_id', None) \
                or getattr(incoming_message, 'senderStaffId', None)
            if not sender_staff_id and isinstance(getattr(callback, 'data', None), dict):
                sender_staff_id = callback.data.get('senderStaffId') or callback.data.get('sender_staff_id')
            sender_user_id = sender_staff_id or self._normalize_recipient_id(sender_id)
            incoming_message.sender_user_id = sender_user_id
            incoming_message.sender_staff_id = sender_staff_id
            if sender_staff_id:
                self.dingtalk_api.cache_openid_userid(sender_id, sender_staff_id)
            sender_nick = incoming_message.sender_nick
            conversation_id = incoming_message.conversation_id
            
            self.logger.info(f"收到消息 - 发送者: {sender_nick}({sender_id}), 会话: {conversation_id}")
            
            # 判断用户角色
            user_role = self._get_user_role(sender_id, sender_user_id, sender_staff_id)
            
            if user_role is None:
                self._send_text_reply(
                    "抱歉，您没有权限使用本机器人。\n请联系管理员配置您的用户权限。",
                    incoming_message
                )
                return AckMessage.STATUS_OK, 'OK'
            
            # 获取消息类型
            msg_type = incoming_message.message_type
            
            # 调试输出 - 查看消息结构
            if msg_type == 'file':
                self.logger.info(f"文件消息调试 - 消息对象属性: {list(vars(incoming_message).keys())}")
                if hasattr(incoming_message, 'content'):
                    self.logger.info(f"文件消息调试 - content内容: {incoming_message.content}")
            
            self.logger.info(f"消息类型: {msg_type}, 用户角色: {user_role}, 当前状态: {self.state_manager.get_status()}")
            
            # 根据消息类型和用户角色分发处理
            if msg_type == 'text':
                await self._handle_text_message(incoming_message, user_role)
            elif msg_type == 'file':
                await self._handle_file_message(incoming_message, user_role, sender_id, conversation_id)
            else:
                self._send_text_reply(f"暂不支持的消息类型: {msg_type}", incoming_message)
            
            return AckMessage.STATUS_OK, 'OK'
            
        except Exception as e:
            self.logger.error(f"处理消息时发生错误: {e}", exc_info=True)
            try:
                self._send_text_reply(f"❌ 处理消息时发生错误: {str(e)}", incoming_message)
            except:
                pass
            self._notify_tech_support(f"处理消息时发生错误: {str(e)}")
            return AckMessage.STATUS_OK, 'ERROR'
    
    def _normalize_recipient_id(self, identifier: Optional[str]) -> Optional[str]:
        """将openId转换为userid（若可能）"""
        if not identifier:
            return None
        if identifier.startswith('$:'):
            return self.dingtalk_api.get_userid_by_openid(identifier)
        return identifier
    
    def _get_user_role(self, *candidate_identifiers: Optional[str]) -> Optional[str]:
        """
        判断用户角色

        Returns:
            'logistics' | 'operation' | None

        未配置 LOGISTICS_USERS/OPERATION_USERS 时：开放物流角色（便于 monorepo 菜单 4）。
        """
        identifiers = []
        for identifier in candidate_identifiers:
            if identifier and identifier not in identifiers:
                identifiers.append(identifier)

        for identifier in identifiers:
            if identifier in config.OPERATION_USERS:
                return "operation"

        for identifier in identifiers:
            if identifier in config.LOGISTICS_USERS:
                return "logistics"

        # 未配置白名单：默认物流（与 pinxiang 一致，由菜单分支控制入口）
        if not config.LOGISTICS_USERS and not config.OPERATION_USERS:
            return "logistics"
        if not config.LOGISTICS_USERS and config.OPERATION_USERS:
            return "logistics"
        return None

    def _is_operation_or_dual(self, incoming_message, user_role: str) -> bool:
        """判断是否是运营或兼任运营的物流账号"""
        if user_role == 'operation':
            return True
        if user_role != 'logistics' or incoming_message is None:
            return False
        candidate_ids = set()
        for attr in ('sender_id', 'sender_user_id', 'sender_staff_id'):
            value = getattr(incoming_message, attr, None)
            if value:
                candidate_ids.add(value)
        normalized = self._normalize_recipient_id(getattr(incoming_message, 'sender_id', None))
        if normalized:
            candidate_ids.add(normalized)
        return any(identifier in config.OPERATION_USERS for identifier in candidate_ids)

    def _is_same_user(self, stored_identifier: Optional[str], incoming_message) -> bool:
        """判断消息发送者是否与存储的identifier一致"""
        if not stored_identifier or incoming_message is None:
            return False
        candidate_ids = set()
        for attr in ('sender_id', 'sender_user_id', 'sender_staff_id'):
            value = getattr(incoming_message, attr, None)
            if value:
                candidate_ids.add(value)
        normalized = self._normalize_recipient_id(getattr(incoming_message, 'sender_id', None))
        if normalized:
            candidate_ids.add(normalized)
        return stored_identifier in candidate_ids
    
    async def _handle_text_message(self, incoming_message, user_role: str):
        """处理文本消息"""
        text_content = incoming_message.text.content.strip()

        # 管理指令
        if text_content.startswith("/跳转"):
            self._handle_force_jump_command(incoming_message, text_content)
            return
        if text_content.startswith("/设置"):
            self._handle_set_command(incoming_message, text_content)
            return
        if text_content in ["/状态", "/status"]:
            self._send_status_message(incoming_message)
            return
        
        # 处理特殊命令
        if text_content in ['帮助', 'help', '?']:
            self._send_help_message(incoming_message, user_role)
        elif text_content in ['状态', 'status']:
            self._send_status_message(incoming_message)
        elif text_content in ['重置', 'reset']:
            self._handle_reset_command(incoming_message, user_role)
        elif text_content in ['确认', 'confirm', '确定']:
            await self._handle_confirmation(incoming_message, user_role)
        elif user_role == 'logistics' and self.state_manager.is_waiting_for_ops_select():
            await self._handle_ops_selection(incoming_message, text_content)
        elif text_content in ['删除', '删除发货单', '是', 'yes', 'YES']:
            await self._handle_delete_decision(incoming_message, user_role, delete=True)
        elif text_content in ['不删除', '否', 'no', 'NO', '保留']:
            await self._handle_delete_decision(incoming_message, user_role, delete=False)
        elif text_content in ['跳过', 'skip', 'SKIP']:
            await self._handle_skip_shipment_numbers(incoming_message, user_role)
        elif text_content in ['获取文件', '下载文件', 'getfile', '文件', '获取', '拼箱结果']:
            await self._handle_get_file(incoming_message, user_role)
        elif text_content in ['模板2', '模板 2', '生成模板2', '重新生成模板2', 'template2', 'template 2']:
            await self._handle_generate_template_v2(incoming_message, user_role)
        elif self.state_manager.is_waiting_for_customs_info():
            await self._handle_customs_info_confirmation(incoming_message, user_role, text_content)
        elif self.state_manager.is_waiting_for_logistics_files() and user_role == 'logistics':
            self._send_text_reply(
                "📥 请上传发货单和报关资料Excel文件。",
                incoming_message
            )
        elif self.state_manager.is_waiting_for_shipment_numbers() and self._is_operation_or_dual(incoming_message, user_role):
            # 处理发货单号输入
            await self._handle_shipment_numbers_input(incoming_message, text_content)
        else:
            # 默认回复
            self._send_text_reply(
                "请上传Excel文件进行处理。\n发送\"帮助\"查看使用说明。",
                incoming_message
            )
    
    async def _handle_file_message(self, incoming_message, user_role: str, 
                                   sender_id: str, conversation_id: str):
        """处理文件消息"""
        try:
            # 获取文件信息 - 钉钉文件消息的内容在 rich_text_content 或 extensions 中
            file_name = None
            download_code = None
            
            # 方法1: 从 rich_text_content 获取（推荐）
            if hasattr(incoming_message, 'rich_text_content') and incoming_message.rich_text_content:
                rich_text = incoming_message.rich_text_content
                self.logger.info(f"rich_text_content: {rich_text}")
                if isinstance(rich_text, dict):
                    file_name = rich_text.get('fileName') or rich_text.get('file_name')
                    download_code = rich_text.get('downloadCode') or rich_text.get('download_code')
            
            # 方法2: 从 extensions 获取（推荐 - 钉钉文件消息标准格式）
            if not download_code and hasattr(incoming_message, 'extensions') and incoming_message.extensions:
                extensions = incoming_message.extensions
                self.logger.info(f"extensions: {extensions}")
                if isinstance(extensions, dict):
                    # 文件信息在 extensions['content'] 里面
                    content = extensions.get('content', {})
                    if content and isinstance(content, dict):
                        file_name = content.get('fileName') or content.get('file_name')
                        download_code = content.get('downloadCode') or content.get('download_code')
                        self.logger.info(f"从 extensions.content 获取: fileName={file_name}, downloadCode={download_code[:30] if download_code else None}...")
            
            # 方法3: 尝试从 text 属性获取（某些版本的SDK）
            if not download_code and hasattr(incoming_message, 'text') and incoming_message.text:
                text_content = incoming_message.text
                self.logger.info(f"text content: {text_content}")
                if hasattr(text_content, 'download_code'):
                    download_code = text_content.download_code
                    file_name = getattr(text_content, 'file_name', '') or getattr(text_content, 'fileName', '')
            
            if not download_code or not file_name:
                self.logger.error(f"无法获取文件信息 - rich_text_content: {getattr(incoming_message, 'rich_text_content', None)}, extensions: {getattr(incoming_message, 'extensions', None)}")
                self._send_text_reply("❌ 无法获取文件信息，请重试或联系技术支持", incoming_message)
                return
            
            self.logger.info(f"收到文件: {file_name}, downloadCode: {download_code}")
            
            # 验证文件类型
            if not file_name.lower().endswith(('.xlsx', '.xls')):
                self._send_text_reply(
                    "❌ 请上传Excel文件（.xlsx 或 .xls格式）",
                    incoming_message
                )
                return

            # 物流确认前上传修正版拼箱结果
            if (user_role == 'logistics'
                and self.state_manager.is_waiting_for_confirmation()
                and self._is_same_user(self.state_manager.get_logistics_user_id(), incoming_message)):
                await self._handle_modified_packing_result_file(
                    incoming_message, sender_id, file_name, download_code
                )
                return

            # 物流补充发货单/报关资料文件
            if user_role == 'logistics' and self.state_manager.is_waiting_for_logistics_files():
                await self._handle_logistics_supporting_file(
                    incoming_message, sender_id,
                    file_name, download_code
                )
                return
            
            # 物流兼任运营：有运营 active 任务时走运营通道；否则仍按物流上传
            if (user_role == 'logistics'
                and self._is_operation_or_dual(incoming_message, user_role)
                and self.state_manager.is_waiting_for_operation(sender_id)):
                self.logger.info("检测到物流账号兼任运营且有运营任务，进入运营文件处理流程")
                await self._handle_operation_file(
                    incoming_message, sender_id,
                    file_name, download_code
                )
                return
            
            # 根据用户角色和当前状态分发处理
            if user_role == 'logistics':
                await self._handle_logistics_file(
                    incoming_message, sender_id, conversation_id,
                    file_name, download_code
                )
            elif user_role == 'operation':
                await self._handle_operation_file(
                    incoming_message, sender_id,
                    file_name, download_code
                )
            
        except Exception as e:
            self.logger.error(f"处理文件消息时发生错误: {e}", exc_info=True)
            self._send_text_reply(f"❌ 处理文件时发生错误: {str(e)}", incoming_message)
    
    async def _handle_logistics_file(self, incoming_message, sender_id: str, 
                                     conversation_id: str, file_name: str, 
                                     download_code: str):
        """处理物流人员上传的发货单"""
        workflow_folder = None
        try:
            # 检查状态
            if not self.state_manager.is_idle():
                current_status = self.state_manager.get_status()
                status_text, status_hint = self._get_status_detail(current_status)
                self._send_text_reply(
                    "⚠️  当前有未完成的流程\n"
                    f"📍 当前进度：{status_text}\n"
                    + (f"➡️ 下一步：{status_hint}\n" if status_hint else "")
                    + "请等待当前流程完成，或发送\"重置\"命令重新开始。",
                    incoming_message
                )
                return
            
            workflow_folder = self._create_workflow_folder()
            
            # 发送处理中提示
            self._send_text_reply("📥 收到发货单，正在处理中，请稍候...", incoming_message)
            
            # 1. 下载文件
            self.logger.info("开始下载发货单文件...")
            robot_code = incoming_message.robot_code
            logistics_file_path = self.dingtalk_api.download_file(
                download_code, file_name, workflow_folder,
                file_info=incoming_message.extensions.get('content', {}) if hasattr(incoming_message, 'extensions') else None,
                robot_code=robot_code
            )
            
            # 菜单 4 不做发票号登记/共享盘建夹，直接拼箱
            registry_code = None
            shared_folder_path = None
            try:
                shipment_info = extract_shipment_info(logistics_file_path)
                self.logger.info(
                    "发货单信息: 店铺=%s, 国家=%s, 运输方式=%s",
                    shipment_info.shop,
                    shipment_info.country,
                    shipment_info.transport_method,
                )
            except Exception as info_exc:
                self.logger.warning("读取发货单摘要失败（忽略）: %s", info_exc)

            # 2. 处理文件（拼箱）
            self.logger.info("开始处理拼箱逻辑...")
            timestamp_suffix = self._extract_timestamp_suffix(logistics_file_path)
            shipping_numbers = self._extract_shipping_numbers(logistics_file_path)
            result_file_name = self._build_result_file_name(logistics_file_path, timestamp_suffix)
            result_file_path = os.path.join(workflow_folder, result_file_name)
            
            processor = PackingBoxProcessor(
                input_file_path=logistics_file_path,
                output_file_path=result_file_path
            )
            
            # 执行处理
            merge_summary = processor.process()
            
            self.logger.info(f"拼箱处理完成，结果文件: {result_file_path}")
            
            # 2.1 生成Amazon发货模板
            # 模版1复制时保持原始模板名称
            template_source_name = os.path.basename(config.AMAZON_TEMPLATE_SOURCE_FILE)
            amazon_template_path = os.path.join(workflow_folder, template_source_name)
            processor.create_amazon_template(amazon_template_path, merge_summary)
            self.logger.info(f"Amazon发货模板生成完成: {amazon_template_path}")
            
            # 3. 上传结果文件
            self.logger.info("开始上传结果文件...")
            media_id, uploaded_file_name = self._upload_stream_file(result_file_path, incoming_message)
            
            # 4. 发送结果文件给物流人员
            session_webhook = incoming_message.session_webhook
            self.dingtalk_api.send_file_message(sender_id, media_id, uploaded_file_name, webhook=session_webhook)
            
            # 4.1 抄送拼箱结果给OTHER_USERS
            self._send_excel_copy_to_others(media_id, uploaded_file_name, "拼箱结果")
            
            # 5. 发送确认提示
            self._send_text_reply(
                "📋 处理结果已发送给您，请查收并确认。\n"
                "回复【确认】➡️ 选择运营并转发\n"
                "上传修正版拼箱结果Excel ➡️ 替换当前结果并重新生成Amazon模板\n"
                "回复【重置】➡️ 放弃本次结果并重新上传发货单",
                incoming_message
            )
            
            # 6. 更新状态
            self.state_manager.set_logistics_uploaded(
                logistics_user_id=sender_id,
                logistics_file_path=logistics_file_path,
                packing_result_path=result_file_path,
                conversation_id=conversation_id,
                amazon_template_path=amazon_template_path,
                shipping_numbers=shipping_numbers,
                workflow_folder_path=workflow_folder,
                registry_code=registry_code,
                shared_folder_path=shared_folder_path,
            )
            
            self.logger.info("物流工作流完成，等待确认")
            
        except Exception as e:
            self.logger.error(f"处理物流文件时发生错误: {e}", exc_info=True)
            self._send_text_reply(
                f"❌ 处理文件时发生错误: {str(e)}\n\n"
                f"请检查文件格式是否正确，或联系技术支持。",
                incoming_message
            )
            self._notify_tech_support(f"处理物流文件失败：{str(e)}")
            # 发生错误时重置状态
            try:
                if workflow_folder:
                    self._cleanup_workflow_files(workflow_folder)
                self._reset_workflow()
            except:
                pass

    async def _handle_modified_packing_result_file(self, incoming_message, sender_id: str,
                                                   file_name: str, download_code: str):
        """处理物流人员在确认前上传的修正版拼箱结果"""
        try:
            workflow_folder = self.state_manager.get_workflow_folder_path()
            if not workflow_folder or not os.path.isdir(workflow_folder):
                workflow_folder = self._create_workflow_folder()
                self.state_manager.update_workflow_folder_path(workflow_folder)

            self._send_text_reply("📥 收到修正版拼箱结果，正在校验并重新生成Amazon模板...", incoming_message)

            robot_code = incoming_message.robot_code
            file_info = incoming_message.extensions.get('content', {}) if hasattr(incoming_message, 'extensions') else None
            modified_file_path = self.dingtalk_api.download_file(
                download_code, file_name, workflow_folder,
                file_info=file_info,
                robot_code=robot_code
            )

            result_path, template_path = self._apply_modified_packing_result(modified_file_path)

            media_id, uploaded_file_name = self._upload_stream_file(result_path, incoming_message)
            self.dingtalk_api.send_file_message(
                sender_id,
                media_id,
                uploaded_file_name,
                webhook=incoming_message.session_webhook
            )

            template_media_id, template_file_name = self._upload_stream_file(template_path, incoming_message)
            self.dingtalk_api.send_file_message(
                sender_id,
                template_media_id,
                template_file_name,
                webhook=incoming_message.session_webhook
            )

            self._send_text_reply(
                "✅ 已使用您上传的修正版拼箱结果，并重新生成 Amazon 模板。\n"
                "请检查后回复【确认】转发给运营，或继续上传新的修正版。",
                incoming_message
            )
            self.logger.info(f"修正版拼箱结果已生效: {result_path}")
        except Exception as e:
            self.logger.error(f"修正版拼箱结果处理失败: {e}", exc_info=True)
            self._send_text_reply(
                f"❌ 修正版拼箱结果校验失败：{str(e)}\n"
                "当前仍使用上一版有效拼箱结果，请修改后重新上传，或回复【确认】继续使用上一版。",
                incoming_message
            )

    def _apply_modified_packing_result(self, modified_file_path: str):
        """校验并应用修正版拼箱结果，返回生效的结果文件和模板路径"""
        merge_summary = self._load_modified_packing_result(modified_file_path)

        workflow_folder = self.state_manager.get_workflow_folder_path()
        if not workflow_folder or not os.path.isdir(workflow_folder):
            workflow_folder = self._create_workflow_folder()
            self.state_manager.update_workflow_folder_path(workflow_folder)

        timestamp = time.strftime("%Y%m%d_%H%M%S")
        base_name, ext = os.path.splitext(os.path.basename(modified_file_path))
        safe_base_name = self._sanitize_filename_part(base_name) or "修正版拼箱结果"
        ext = ext if ext.lower() in ('.xlsx', '.xls') else '.xlsx'
        result_path = os.path.join(workflow_folder, f"{safe_base_name}_manual_{timestamp}{ext}")
        if os.path.abspath(modified_file_path) != os.path.abspath(result_path):
            shutil.copy2(modified_file_path, result_path)

        template_source_name = os.path.basename(config.AMAZON_TEMPLATE_SOURCE_FILE)
        template_base, template_ext = os.path.splitext(template_source_name)
        template_path = os.path.join(
            workflow_folder,
            f"{template_base}_manual_{timestamp}{template_ext or '.xlsx'}"
        )
        temp_processor = PackingBoxProcessor(
            input_file_path=result_path,
            output_file_path=result_path
        )
        temp_processor.create_amazon_template(template_path, merge_summary)

        shared_folder_path = self.state_manager.get_full_state().get('shared_folder_path')
        if shared_folder_path:
            try:
                shutil.copy2(result_path, os.path.join(shared_folder_path, os.path.basename(result_path)))
                shutil.copy2(template_path, os.path.join(shared_folder_path, os.path.basename(template_path)))
            except Exception as copy_exc:
                self.logger.warning(f"复制修正版文件到共享盘失败: {copy_exc}")

        self.state_manager.update_packing_result(result_path, template_path)
        return result_path, template_path

    def _load_modified_packing_result(self, file_path: str) -> pd.DataFrame:
        """读取并校验修正版拼箱结果"""
        sheet_name = '拼箱计算结果'
        try:
            merge_summary = pd.read_excel(file_path, sheet_name=sheet_name)
        except ValueError as exc:
            if sheet_name in str(exc):
                raise ValueError(f"缺少工作表：{sheet_name}") from exc
            raise

        required_columns = [
            'SKU',
            '发货数量',
            '箱数',
            '单份数量',
            '小组名称',
            '单箱重量',
            '单箱理论长',
            '单箱理论宽',
            '单箱理论高',
        ]
        missing_columns = [col for col in required_columns if col not in merge_summary.columns]
        if missing_columns:
            raise ValueError(f"缺少必要列：{', '.join(missing_columns)}")

        merge_summary = merge_summary.copy()
        fill_columns = ['小组名称', '单箱重量', '单箱理论长', '单箱理论宽', '单箱理论高']
        for col in fill_columns:
            merge_summary[col] = merge_summary[col].ffill()

        sku_series = merge_summary['SKU'].dropna().map(lambda value: str(value).strip())
        merge_summary = merge_summary.loc[sku_series[sku_series != ''].index].copy()
        if merge_summary.empty:
            raise ValueError("拼箱计算结果中没有可用的SKU行")

        numeric_columns = ['发货数量', '箱数', '单份数量', '单箱重量', '单箱理论长', '单箱理论宽', '单箱理论高']
        invalid_columns = []
        for col in numeric_columns:
            converted = pd.to_numeric(merge_summary[col], errors='coerce')
            if converted.isna().any():
                invalid_columns.append(col)
            else:
                merge_summary[col] = converted
        if invalid_columns:
            raise ValueError(f"以下列存在非数字或空值：{', '.join(invalid_columns)}")

        return merge_summary

    async def _handle_logistics_supporting_file(self, incoming_message, sender_id: str,
                                                file_name: str, download_code: str):
        """处理物流补充上传的发货单/报关资料文件"""
        try:
            import sys
            from pathlib import Path
            from openpyxl import load_workbook

            stored_logistics_id = self.state_manager.get_logistics_user_id()
            if stored_logistics_id and not self._is_same_user(stored_logistics_id, incoming_message):
                self._send_text_reply(
                    "⚠️ 请由本次流程的物流人员上传文件。",
                    incoming_message
                )
                return

            state = self.state_manager.get_full_state()
            registry_code = state.get("registry_code")
            shared_folder_path = state.get("shared_folder_path")
            workflow_folder = self.state_manager.get_workflow_folder_path()

            if shared_folder_path:
                target_dir = shared_folder_path
            else:
                if not workflow_folder or not os.path.isdir(workflow_folder):
                    workflow_folder = self._create_workflow_folder()
                    self.state_manager.update_workflow_folder_path(workflow_folder)
                target_dir = workflow_folder

            os.makedirs(target_dir, exist_ok=True)

            robot_code = incoming_message.robot_code
            file_info = incoming_message.extensions.get('content', {}) if hasattr(incoming_message, 'extensions') else None
            saved_path = self.dingtalk_api.download_file(
                download_code, file_name, target_dir,
                file_info=file_info,
                robot_code=robot_code
            )

            self.state_manager.add_logistics_received_file(saved_path)

            # 尝试导出报关资料PDF及合同PDF
            try:
                common_path = str(getattr(config, 'COMMON_ROOT', config.PROJECT_ROOT.parent))
                if common_path not in sys.path:
                    sys.path.insert(0, common_path)
                from Common.Utils.excel_to_pdf import excel_to_pdf, excel_sheet_to_pdf

                lowered_name = str(file_name)
                is_customs_excel = "报关资料" in lowered_name
                has_contract_sheet = False
                if is_customs_excel:
                    try:
                        wb = load_workbook(saved_path, read_only=True, data_only=True)
                        if "合同" in wb.sheetnames:
                            has_contract_sheet = True
                        wb.close()
                    except Exception:
                        pass

                if is_customs_excel and shared_folder_path:
                    invoice_no = registry_code or ""
                    pdf_dir_name = f"{invoice_no} 报关资料&发票" if invoice_no else "报关资料&发票"
                    pdf_dir = os.path.join(shared_folder_path, pdf_dir_name)
                    os.makedirs(pdf_dir, exist_ok=True)

                    pdf_name = f"{Path(saved_path).stem}.pdf"
                    pdf_path = os.path.join(pdf_dir, pdf_name)
                    excel_to_pdf(saved_path, pdf_path)

                    if has_contract_sheet:
                        contract_pdf_name = f"{invoice_no} 销售合同.pdf" if invoice_no else "销售合同.pdf"
                        contract_pdf_path = os.path.join(shared_folder_path, contract_pdf_name)
                        excel_sheet_to_pdf(saved_path, "合同", contract_pdf_path)

            except Exception as pdf_exc:
                self.logger.error(f"报关资料PDF导出失败: {pdf_exc}")
            received = len(self.state_manager.get_logistics_received_files())
            expected = self.state_manager.get_logistics_files_expected() or 2

            reply = (
                f"✅ 已保存文件到共享盘：{os.path.basename(saved_path)}\n"
                f"已收到 {received}/{expected} 个文件。"
            )
            if received >= expected:
                reply += "\n文件已收齐，开始生成清关资料，请稍候..."
                self._send_text_reply(reply, incoming_message)
                await self._process_customs_after_logistics_upload(
                    incoming_message,
                    self.state_manager.get_lingxing_shipment_numbers() or [],
                )
                return
            else:
                reply += "\n请继续上传剩余文件。"

            self._send_text_reply(reply, incoming_message)

        except Exception as e:
            self.logger.error(f"保存物流补充文件失败: {e}", exc_info=True)
            self._send_text_reply(
                f"❌ 保存文件失败: {str(e)}",
                incoming_message
            )
    
    async def _handle_confirmation(self, incoming_message, user_role: str):
        """处理物流人员的确认 → 进入选运营"""
        try:
            if user_role != "logistics":
                self._send_text_reply("⚠️  只有物流人员可以执行确认操作", incoming_message)
                return

            if not self.state_manager.is_waiting_for_confirmation():
                self._send_text_reply("⚠️  当前没有需要确认的内容", incoming_message)
                return

            ops_list = list(getattr(config, "OPS_USERS", None) or [])
            if not ops_list and config.OPERATION_USERS:
                ops_list = [{"name": uid, "user_id": uid} for uid in config.OPERATION_USERS]
            if not ops_list:
                self._send_text_reply(
                    "⚠️ 未配置运营人员。请在 .env 设置 PINXIANG_OPS_USERS（与流程三相同，格式：姓名:userId,...）",
                    incoming_message,
                )
                return

            packing_result_path = self.state_manager.get_packing_result_path()
            if not packing_result_path or not os.path.exists(packing_result_path):
                self._send_text_reply("⚠️ 拼箱结果文件丢失，请重新上传发货单。", incoming_message)
                return

            self.state_manager.set_waiting_for_ops_select()
            self._send_text_reply(self._ops_menu_text(prefix="已确认拼箱结果。\n请选择要转发的运营：\n"), incoming_message)
            self.logger.info("lcl waiting for ops selection")
        except Exception as e:
            self.logger.error(f"处理确认时发生错误: {e}", exc_info=True)
            self._send_text_reply(f"❌ 确认失败: {str(e)}", incoming_message)
            self._notify_tech_support(f"处理物流确认失败：{str(e)}")

    def _ops_menu_text(self, *, prefix: str = "") -> str:
        ops_list = list(getattr(config, "OPS_USERS", None) or [])
        if not ops_list and config.OPERATION_USERS:
            ops_list = [{"name": uid, "user_id": uid} for uid in config.OPERATION_USERS]
        lines = [prefix.rstrip(), ""]
        for i, ops in enumerate(ops_list, start=1):
            lines.append(f"{i}. {ops.get('name') or ops.get('user_id')}")
        lines.append("")
        lines.append("回复序号或姓名选择；回复【重置】放弃。")
        return "\n".join(lines).strip()

    def _match_ops_choice(self, text: str) -> Optional[dict]:
        ops_list = list(getattr(config, "OPS_USERS", None) or [])
        if not ops_list and config.OPERATION_USERS:
            ops_list = [{"name": uid, "user_id": uid} for uid in config.OPERATION_USERS]
        normalized = re.sub(r"\s+", "", (text or "").strip()).lower()
        if not normalized:
            return None
        for i, ops in enumerate(ops_list, start=1):
            name = re.sub(r"\s+", "", str(ops.get("name") or "")).lower()
            uid = str(ops.get("user_id") or "")
            if normalized == str(i) or normalized.startswith(f"{i}.") or normalized.startswith(f"{i}、"):
                return ops
            if name and (normalized == name or name in normalized or normalized in name):
                return ops
            if uid and normalized == uid.lower():
                return ops
        return None

    async def _handle_ops_selection(self, incoming_message, text_content: str) -> None:
        """物流选择运营后，只转发给该运营。"""
        try:
            ops = self._match_ops_choice(text_content)
            if ops is None:
                self._send_text_reply(self._ops_menu_text(prefix="请选择运营人员：\n"), incoming_message)
                return
            await self._forward_to_selected_ops(incoming_message, ops)
        except Exception as e:
            self.logger.error(f"选择运营失败: {e}", exc_info=True)
            self._send_text_reply(f"❌ 转发失败: {str(e)}", incoming_message)
            self._notify_tech_support(f"lcl 选择运营转发失败：{str(e)}")

    def _push_lcl_ops_job_files(self, corp_user_id: str, job: dict, *, from_queue: bool = False) -> None:
        """私聊推送拼箱结果+模板给运营（队列提升时无 stream message）。"""
        packing_result_path = job.get("packing_result_path") or ""
        amazon_template_path = job.get("amazon_template_path") or ""
        if not packing_result_path or not os.path.exists(packing_result_path):
            raise FileNotFoundError(f"拼箱结果不存在: {packing_result_path}")
        packing_name = os.path.basename(packing_result_path)
        head = (
            "【分仓拼箱】排队任务开始处理。\n"
            if from_queue
            else "【分仓拼箱】物流已审核通过，请处理。\n"
        )
        self.dingtalk_api.send_text_message(
            corp_user_id,
            head
            + "请下载拼箱结果与 Amazon 模板；已默认发送模板1，如需新版格式回复【模板2】。\n"
            + "也可上传卖家后台包装信息表，机器人将自动填写并回传。\n"
            + f"拼箱结果：{packing_name}",
        )
        media_id, file_name = self.dingtalk_api.upload_file(packing_result_path)
        self.dingtalk_api.send_file_message(corp_user_id, media_id, file_name)
        if amazon_template_path and os.path.exists(amazon_template_path):
            t_media, t_name = self.dingtalk_api.upload_file(amazon_template_path)
            self.dingtalk_api.send_file_message(corp_user_id, t_media, t_name)

    def _reset_workflow(self, ops_user_id: Optional[str] = None) -> None:
        """结束流程；若该运营有排队任务则自动提升并推送。"""
        oid = ops_user_id or self.state_manager.get_operation_user_id()
        self.state_manager.reset()
        if oid:
            self._after_ops_workflow_reset(str(oid))

    def _after_ops_workflow_reset(self, ops_user_id: str) -> None:
        """reset 后若提升了队列下一单，推送文件。"""
        if not ops_user_id:
            return
        job = (self.state_manager.state.get("ops_active") or {}).get(str(ops_user_id))
        if not job or not job.get("needs_push"):
            return
        corp = self._normalize_recipient_id(ops_user_id) or ops_user_id
        try:
            self._push_lcl_ops_job_files(corp, job, from_queue=True)
            job["needs_push"] = False
            active = dict(self.state_manager.state.get("ops_active") or {})
            active[str(ops_user_id)] = job
            self.state_manager.state["ops_active"] = active
            self.state_manager._save_state()
            self.logger.info("lcl pushed queued job to ops=%s", ops_user_id)
        except Exception as exc:
            self.logger.exception("lcl push queued job failed ops=%s: %s", ops_user_id, exc)

    async def _forward_to_selected_ops(self, incoming_message, ops: dict) -> None:
        """将拼箱结果 + Amazon 模板发给指定运营（忙则排队，物流立即释放）。"""
        if not self.state_manager.is_waiting_for_ops_select():
            self._send_text_reply("⚠️ 当前不在选运营阶段。", incoming_message)
            return

        packing_result_path = self.state_manager.get_packing_result_path()
        if not packing_result_path or not os.path.exists(packing_result_path):
            self._send_text_reply("⚠️ 拼箱结果文件丢失，请重新上传发货单。", incoming_message)
            return

        shared_folder_path = self.state_manager.state.get("shared_folder_path")
        if shared_folder_path and os.path.exists(packing_result_path):
            try:
                shared_result_path = os.path.join(shared_folder_path, os.path.basename(packing_result_path))
                shutil.copy2(packing_result_path, shared_result_path)
            except Exception as copy_exc:
                self.logger.warning(f"复制拼箱结果到共享盘失败: {copy_exc}")

        amazon_template_path = self.state_manager.get_amazon_template_path()
        if amazon_template_path and os.path.exists(amazon_template_path) and shared_folder_path:
            try:
                shared_template_path = os.path.join(
                    shared_folder_path, os.path.basename(amazon_template_path)
                )
                shutil.copy2(amazon_template_path, shared_template_path)
            except Exception as copy_exc:
                self.logger.warning(f"复制Amazon模板到共享盘失败: {copy_exc}")
        elif not amazon_template_path or not os.path.exists(amazon_template_path or ""):
            self.logger.warning("未找到Amazon发货模板文件，仅发送拼箱结果")

        ops_id = ops.get("user_id") or ""
        ops_name = ops.get("name") or ops_id
        corp_user_id = self._normalize_recipient_id(ops_id) or ops_id
        if not corp_user_id:
            self._send_text_reply(f"❌ 无法解析运营【{ops_name}】的 userId", incoming_message)
            return

        job = self.state_manager.snapshot_current_job()
        packing_name = os.path.basename(packing_result_path)
        forward_keys = self.state_manager.shipment_keys_from_job(job)
        active = (self.state_manager.state.get("ops_active") or {}).get(str(ops_id))

        # 与运营当前单重复：以运营数据为准，不覆盖、不入队
        if active and self.state_manager.shipment_keys_overlap(
            forward_keys, self.state_manager.shipment_keys_from_job(active)
        ):
            self.state_manager.drop_queue_jobs_overlapping(ops_id, forward_keys)
            self.state_manager.release_logistics_session()
            active_name = os.path.basename(active.get("packing_result_path") or packing_name)
            self._send_text_reply(
                f"✅ 运营【{ops_name}】已有本单拼箱数据（以运营上传为准），未重复转发。\n"
                f"运营当前：{active_name}\n"
                f"物流文件：{packing_name}\n\n"
                "您可继续上传下一单。",
                incoming_message,
            )
            try:
                self.dingtalk_api.send_text_message(
                    corp_user_id,
                    f"【分仓拼箱】物流再次转发本单，已忽略（保留您当前/已上传的拼箱数据）。\n"
                    f"物流文件：{packing_name}",
                )
            except Exception as exc:
                self.logger.warning("notify ops skip-dup forward failed: %s", exc)
            self.logger.info("lcl skip duplicate forward ops=%s packing=%s", ops_id, packing_name)
            return

        # 队列里同单先清掉
        if forward_keys:
            self.state_manager.drop_queue_jobs_overlapping(ops_id, forward_keys)

        # 运营忙碌：入队，物流仍释放
        if self.state_manager.ops_is_busy(ops_id):
            pos = self.state_manager.enqueue_ops_job(ops_id, job)
            self.state_manager.release_logistics_session()
            self._send_text_reply(
                f"✅ 已转发给运营【{ops_name}】（排队中）。\n"
                f"本单排队第 {pos} 位，对方完成后会自动推送。\n"
                f"拼箱结果：{packing_name}\n\n"
                "您可继续上传下一单。",
                incoming_message,
            )
            try:
                self.dingtalk_api.send_text_message(
                    corp_user_id,
                    f"【分仓拼箱】新任务已加入队列（第 {pos} 位）。\n"
                    f"文件：{packing_name}\n"
                    "请先完成当前单，完成后系统会自动推送下一单。",
                )
            except Exception as exc:
                self.logger.warning("notify ops queue failed: %s", exc)
            self.logger.info("lcl queued for ops=%s(%s) pos=%s", ops_name, ops_id, pos)
            return

        # 空闲：立刻推送
        try:
            media_id, file_name = self._upload_stream_file(packing_result_path, incoming_message)
            template_media_id = None
            template_file_name = None
            if amazon_template_path and os.path.exists(amazon_template_path):
                template_media_id, template_file_name = self._upload_stream_file(
                    amazon_template_path, incoming_message
                )
            self.dingtalk_api.send_file_message(corp_user_id, media_id, file_name)
            if template_media_id and template_file_name:
                self.dingtalk_api.send_file_message(
                    corp_user_id, template_media_id, template_file_name
                )
            instruction_text = (
                "【分仓拼箱】物流已审核通过，请处理。\n"
                "请下载拼箱结果与 Amazon 模板；已默认发送模板1，如需新版格式回复【模板2】。\n"
                "也可上传卖家后台包装信息表，机器人将自动填写并回传。"
            )
            self.dingtalk_api.send_text_message(corp_user_id, instruction_text)
        except Exception as e:
            self.logger.error(f"发送给运营 {ops_name}({ops_id}) 失败: {e}")
            self._send_text_reply(f"❌ 转发给运营【{ops_name}】失败: {e}", incoming_message)
            return

        self._send_excel_copy_to_others(media_id, file_name, "拼箱结果（物流已确认）")
        if template_media_id and template_file_name:
            self._send_excel_copy_to_others(template_media_id, template_file_name, "Amazon发货模板")

        self.state_manager.activate_ops_job(ops_id, job)
        self.state_manager.release_logistics_session()
        self._send_text_reply(
            f"✅ 已转发给运营【{ops_name}】。\n"
            "已发送拼箱结果与 Amazon 模板1。\n"
            "您可继续上传下一单。",
            incoming_message,
        )
        self.logger.info("lcl forwarded to ops=%s(%s)", ops_name, ops_id)

    async def _handle_delete_decision(self, incoming_message, user_role: str, delete: bool):
        """处理运营确认是否删除发货单"""
        if not self.state_manager.is_waiting_for_delete_confirmation():
            self._send_text_reply("⚠️ 当前没有待确认的删除操作。", incoming_message)
            return

        # 特殊场景：同一人兼任物流与运营
        is_dual_role = False
        if user_role == 'logistics':
            sender_id = incoming_message.sender_id
            sender_user_id = getattr(incoming_message, 'sender_user_id', None)
            sender_staff_id = getattr(incoming_message, 'sender_staff_id', None)
            # 检查是否同时也是运营人员
            for identifier in [sender_id, sender_user_id, sender_staff_id]:
                if identifier and identifier in config.OPERATION_USERS:
                    is_dual_role = True
                    self.logger.info("检测到物流账号兼任运营，允许执行删除确认")
                    break
        
        if user_role != 'operation' and not is_dual_role:
            self._send_text_reply("⚠️ 只有运营人员可以确认是否删除发货单。", incoming_message)
            return

        stored_operation_id = self.state_manager.get_operation_user_id()
        if stored_operation_id and not self._is_same_user(stored_operation_id, incoming_message):
            self._send_text_reply("⚠️ 请由上传Amazon文件的运营人员完成确认。", incoming_message)
            return

        shipment_nos = self._parse_shipment_nos(self.state_manager.get_shipping_numbers())
        if delete:
            if not shipment_nos:
                self._send_text_reply("⚠️ 未找到发货单号，无法调用删除接口，流程将直接结束。", incoming_message)
            else:
                try:
                    api_resp = await delete_shipment_list(shipment_nos)
                    # 简化消息展示，不再返回完整 JSON
                    if api_resp.get('code') == 0:
                        status_msg = "✅ 删除成功"
                    else:
                        status_msg = f"❌ 删除失败：{api_resp.get('message', '未知错误')}"
                    
                    self._send_text_reply(
                        f"📦 发货单号：{', '.join(shipment_nos)}\n"
                        f"✉️ 结果：{status_msg}",
                        incoming_message
                    )
                except Exception as exc:
                    self.logger.error(f"删除发货单失败: {exc}", exc_info=True)
                    self._send_text_reply(
                        f"❌ 删除失败：{exc}\n"
                        "可回复【删除】重试，或回复【不删除】结束流程。",
                        incoming_message
                    )
                    self._notify_tech_support(f"删除发货单失败：{exc}")
                    return
        else:
            self._send_text_reply("ℹ️ 已选择不删除发货单，流程将结束。", incoming_message)

        logistics_user_id = self.state_manager.get_logistics_user_id()
        if logistics_user_id:
            try:
                normalized_logistics_id = self._normalize_recipient_id(logistics_user_id)
                if normalized_logistics_id:
                    summary_line = "✅ 运营人员已完成Amazon包装信息处理"
                    if delete and shipment_nos:
                        summary_line += f"\n已提交删除发货单请求：{', '.join(shipment_nos)}"
                    elif delete and not shipment_nos:
                        summary_line += "\n未找到发货单号，未执行删除"
                    else:
                        summary_line += "\n未删除发货单"
                    self.dingtalk_api.send_text_message(normalized_logistics_id, summary_line)
                else:
                    self.logger.warning(f"无法获取物流人员 {logistics_user_id} 的userid，未推送完成通知")
            except Exception as exc:
                self.logger.error(f"通知物流人员失败: {exc}")

        # 进入询问领星发货单号流程
        self.state_manager.set_waiting_for_shipment_numbers()
        self._send_text_reply(
            "📋 请输入需要处理的领星发货单号（每行一个）：\n\n"
            "格式示例：\n"
            "SP260119001\n"
            "SP260119002\n"
            "SP260119003\n\n"
            "输入完成后发送，或回复【跳过】结束流程。",
            incoming_message
        )
        self.logger.info("等待运营输入领星发货单号")

    async def _handle_skip_shipment_numbers(self, incoming_message, user_role: str):
        """处理跳过发货单号输入"""
        if not self.state_manager.is_waiting_for_shipment_numbers():
            self._send_text_reply("⚠️ 当前没有待输入的发货单号。", incoming_message)
            return
        
        if not self._is_operation_or_dual(incoming_message, user_role):
            self._send_text_reply("⚠️ 只有运营人员可以跳过此步骤。", incoming_message)
            return
        
        self._send_text_reply("ℹ️ 已跳过领星发货单处理，流程结束。", incoming_message)
        
        # 通知物流人员
        logistics_user_id = self.state_manager.get_logistics_user_id()
        if logistics_user_id:
            try:
                normalized_id = self._normalize_recipient_id(logistics_user_id)
                if normalized_id:
                    self.dingtalk_api.send_text_message(
                        normalized_id,
                        "✅ 整个工作流程已完成，可以开始新的流程。"
                    )
            except Exception as exc:
                self.logger.error(f"通知物流人员失败: {exc}")
        
        self._reset_workflow()
        self.logger.info("流程结束，状态已重置")

    async def _handle_shipment_numbers_input(self, incoming_message, text_content: str):
        """处理运营输入的发货单号列表"""
        # 解析发货单号（每行一个）
        lines = [line.strip() for line in text_content.strip().split('\n') if line.strip()]
        shipment_numbers = []
        for line in lines:
            # 验证格式：SP开头 + 数字
            if re.match(r'^SP\d+$', line.upper()):
                shipment_numbers.append(line.upper())
            else:
                self._send_text_reply(
                    f"❌ 发货单号格式错误: {line}\n\n"
                    "请输入正确格式的发货单号（如 SP260119001）",
                    incoming_message
                )
                return
        
        if not shipment_numbers:
            self._send_text_reply("❌ 未识别到有效的发货单号，请重新输入。", incoming_message)
            return
        
        # 保存发货单号列表
        self.state_manager.set_lingxing_shipment_numbers(shipment_numbers)
        
        self._send_text_reply(
            f"📋 收到 {len(shipment_numbers)} 个发货单号：\n"
            f"{chr(10).join(shipment_numbers)}\n\n"
            "🔄 正在启动领星自动化处理，请稍候...",
            incoming_message
        )
        
        # 执行领星自动化处理
        await self._process_lingxing_shipments(incoming_message, shipment_numbers)

    async def _handle_customs_info_confirmation(self, incoming_message, user_role: str, text_content: str):
        """处理物流补全清关资料后的确认"""
        if not self.state_manager.is_waiting_for_customs_info():
            return

        logistics_user_id = self.state_manager.get_logistics_user_id()
        if user_role != 'logistics' or not self._is_same_user(logistics_user_id, incoming_message):
            self._send_text_reply(
                "⚠️ 正在等待物流人员补全清关资料。\n"
                "请物流人员在补全后回复：已补全 SPxxxx",
                incoming_message
            )
            return

        keyword_hit = any(word in text_content for word in ["已补全", "补全", "已完成", "完成"])
        order_matches = re.findall(r"SP\\d+", text_content.upper())
        pending = self.state_manager.get_pending_customs_shipments() or []
        if not pending:
            self._send_text_reply("⚠️ 当前没有待补全清关的发货单。", incoming_message)
            return

        if not keyword_hit and not order_matches:
            self._send_text_reply(
                f"⚠️ 请确认清关资料已补全后回复：已补全 {pending[0]}",
                incoming_message
            )
            return

        order_no = order_matches[0] if order_matches else pending[0]
        if order_no not in pending:
            self._send_text_reply(
                f"⚠️ 发货单号 {order_no} 不在待补全清关列表中。\n"
                f"当前待补全：{', '.join(pending)}",
                incoming_message
            )
            return

        # 仅允许从队首继续
        if order_no != pending[0]:
            self._send_text_reply(
                f"⚠️ 请按顺序处理，当前待补全发货单为：{pending[0]}",
                incoming_message
            )
            return

        self.state_manager.clear_customs_waiting()

        self._send_text_reply(
            f"✅ 已收到补全通知，继续处理发货单：{', '.join(pending)}",
            incoming_message
        )
        await self._process_customs_after_logistics_upload(
            incoming_message,
            pending,
        )

    async def _process_lingxing_shipments(self, incoming_message, shipment_numbers: list,
                                          partial_results: Optional[list] = None,
                                          lock_stock_partial: Optional[dict] = None):
        """执行领星发货单自动化处理"""
        import sys
        sys.path.insert(0, str(getattr(config, 'COMMON_ROOT', config.PROJECT_ROOT.parent)))
        try:
            from Common.web import BrowserDriver
            from Common.web.lingxing import LingXingPortal
            from Common.api.lingxing_client import LingXingClient
            from Common.api import config as api_config
        except ImportError as exc:
            self.logger.error(f"导入 Common 库失败: {exc}")
            self._send_text_reply(
                f"❌ 自动化模块加载失败: {exc}\n请联系技术支持。",
                incoming_message
            )
            self._reset_workflow()
            return
        
        results = list(partial_results or [])
        state = self.state_manager.get_full_state()
        registry_code = state.get("registry_code")
        workflow_folder_path = state.get("workflow_folder_path")
        shared_folder_path = state.get("shared_folder_path")
        if not workflow_folder_path:
            workflow_folder_path = self._create_workflow_folder()
        if not registry_code:
            self._send_text_reply(
                "❌ 未获取到发票号，无法填写物流商单号/查询单号。\n"
                "请确认物流已上传并完成登记后再重试。",
                incoming_message
            )
            self._reset_workflow()
            return

        logistics_no = registry_code
        tracking_no = registry_code

        expected_transport_mode = "海派"
        expected_logistics_channel = "海派"
        expected_logistics_provider = "杭州平谊国际货运代理有限公司"

        remark_template = "资料已上传，箱唛一式两份，{total_box_num}箱，{date}，平谊提货"
        remark_client = None
        if api_config.LINGXING_API_KEY and api_config.LINGXING_TOKEN_URL:
            remark_client = LingXingClient(
                host=api_config.LINGXING_API_HOST,
                app_id=api_config.LINGXING_API_KEY,
                app_secret=api_config.LINGXING_API_SECRET,
                token_url=api_config.LINGXING_TOKEN_URL,
                token_key=api_config.LINGXING_TOKEN_REQUEST_KEY,
                ssl_verify=api_config.LINGXING_SSL_VERIFY,
            )
        lock_stock_success = list((lock_stock_partial or {}).get("success", []))
        lock_stock_failed = list((lock_stock_partial or {}).get("failed", []))
        lock_stock_seen = set(lock_stock_success + lock_stock_failed)

        def _record_lock_stock(target_list: list, order_no: str):
            if order_no in lock_stock_seen:
                return
            target_list.append(order_no)
            lock_stock_seen.add(order_no)
        try:
            driver = BrowserDriver()
            with driver.session() as page:
                portal = LingXingPortal(page)

                # 检查是否需要登录
                portal.goto_home()
                if portal.is_login_page():
                    self._send_text_reply(
                        "⚠️ 领星需要登录，请先在浏览器中登录领星后重试。",
                        incoming_message
                    )
                    self._reset_workflow()
                    return
                
                # 遍历处理每个发货单
                for idx, order_no in enumerate(shipment_numbers):
                    self.logger.info(f"开始处理发货单: {order_no}")
                    # TODO: 需要获取物流商单号和查询单号
                    success, msg = portal.process_shipment(
                        order_no=order_no,
                        logistics_no=logistics_no,
                        tracking_no=tracking_no,
                        expected_transport_mode=expected_transport_mode,
                        expected_logistics_channel=expected_logistics_channel,
                        expected_logistics_provider=expected_logistics_provider,
                        download_dir=workflow_folder_path,
                    )
                    if success and remark_client:
                        try:
                            remark_result = await remark_client.update_shipment_remark_with_template(
                                order_no,
                                template=remark_template,
                            )
                            self.logger.info(
                                "备注更新完成: %s old=%s | new=%s",
                                order_no,
                                remark_result.get("old_remark"),
                                remark_result.get("new_remark"),
                            )
                        except Exception as exc:
                            self.logger.error(f"{order_no}: 备注更新失败: {exc}")
                        try:
                            lock_resp = await remark_client.lock_shipment_stock([order_no])
                            if lock_resp.get("code") in (0, "0"):
                                _record_lock_stock(lock_stock_success, order_no)
                                self.logger.info("库存分配成功: %s", order_no)
                            else:
                                _record_lock_stock(lock_stock_failed, order_no)
                                self.logger.warning(
                                    "库存分配失败: %s code=%s msg=%s details=%s",
                                    order_no,
                                    lock_resp.get("code"),
                                    lock_resp.get("message"),
                                    lock_resp.get("error_details"),
                                )
                        except Exception as exc:
                            _record_lock_stock(lock_stock_failed, order_no)
                            self.logger.error(f"{order_no}: 库存分配异常: {exc}")
                    results.append((order_no, success, msg))
                    self.logger.info(f"处理结果: {msg}")

        except Exception as exc:
            self.logger.error(f"领星自动化处理失败: {exc}", exc_info=True)
            self._send_text_reply(
                f"❌ 领星自动化处理失败: {exc}",
                incoming_message
            )
            self._reset_workflow()
            return
        
        # 汇总结果
        success_count = sum(1 for _, success, _ in results if success)
        fail_count = len(results) - success_count
        
        result_lines = []
        for order_no, success, msg in results:
            status_icon = "✅" if success else "❌"
            result_lines.append(f"{status_icon} {msg}")

        summary = (
            f"📊 领星自动化处理完成\n\n"
            f"成功: {success_count} / 失败: {fail_count}\n\n"
            f"详细结果：\n" + "\n".join(result_lines)
        )
        self._send_text_reply(summary, incoming_message)
        
        # 通知物流人员
        wait_for_logistics_files = False
        logistics_user_id = self.state_manager.get_logistics_user_id()
        if logistics_user_id:
            try:
                normalized_id = self._normalize_recipient_id(logistics_user_id)
                if normalized_id:
                    self.dingtalk_api.send_text_message(
                        normalized_id,
                        f"✅ 领星发货单处理完成\n成功: {success_count} / 失败: {fail_count}\n\n"
                        "请按下方提示上传发货单和报关资料Excel文件。"
                    )
                    success_lines = "\n".join(lock_stock_success) if lock_stock_success else "无"
                    failed_lines = "\n".join(lock_stock_failed) if lock_stock_failed else "无"
                    invoice_no = self.state_manager.get_full_state().get("registry_code") or ""
                    stock_message = (
                        "分配库存成功：\n"
                        f"{success_lines}\n"
                        "分配库存失败：\n"
                        f"{failed_lines}\n"
                        f"发票号：{invoice_no}\n"
                        "请去领星检查并导出发货单,自行生成报关资料，将两个Excel上传给我"
                    )
                    self.dingtalk_api.send_text_message(normalized_id, stock_message)
                    wait_for_logistics_files = True
                else:
                    self.logger.warning(f"无法获取物流人员 {logistics_user_id} 的userid，未推送完成通知")
            except Exception as exc:
                self.logger.error(f"通知物流人员失败: {exc}")

        if wait_for_logistics_files:
            self.state_manager.set_waiting_for_logistics_files(expected_count=2)
            self.logger.info("等待物流上传发货单/报关资料Excel文件")
        else:
            self._reset_workflow()
            self.logger.info("领星自动化处理完成，流程结束")

    async def _process_customs_after_logistics_upload(self, incoming_message, shipment_numbers: list):
        """物流上传文件后生成清关资料并下载处理"""
        import sys
        sys.path.insert(0, str(getattr(config, 'COMMON_ROOT', config.PROJECT_ROOT.parent)))

        if not shipment_numbers:
            self._send_text_reply("⚠️ 未找到领星发货单号，无法生成清关资料。", incoming_message)
            return

        try:
            from Common.web import BrowserDriver
            from Common.web.lingxing import LingXingPortal
        except ImportError as exc:
            self.logger.error(f"导入 Common 库失败: {exc}")
            self._send_text_reply(
                f"❌ 自动化模块加载失败: {exc}\n请联系技术支持。",
                incoming_message
            )
            return

        state = self.state_manager.get_full_state()
        registry_code = state.get("registry_code")
        shared_folder_path = state.get("shared_folder_path")
        workflow_folder_path = state.get("workflow_folder_path")
        if not workflow_folder_path:
            workflow_folder_path = self._create_workflow_folder()
            self.state_manager.update_workflow_folder_path(workflow_folder_path)

        if shared_folder_path and registry_code:
            customs_download_dir = os.path.join(shared_folder_path, f"{registry_code} 报关资料&发票")
        else:
            customs_download_dir = os.path.join(workflow_folder_path, "customs_downloads")
        os.makedirs(customs_download_dir, exist_ok=True)

        results = []
        customs_added_orders = []
        try:
            driver = BrowserDriver()
            with driver.session() as page:
                portal = LingXingPortal(page)
                portal.goto_home()
                if portal.is_login_page():
                    self._send_text_reply(
                        "⚠️ 领星需要登录，请先在浏览器中登录领星后重试。",
                        incoming_message
                    )
                    return

                for idx, order_no in enumerate(shipment_numbers):
                    ok, msg = portal.add_customs_clearance(order_no)
                    if not ok:
                        if self._is_customs_missing_message(msg):
                            pending = shipment_numbers[idx:]
                            self.state_manager.set_waiting_for_customs_info(
                                pending_shipments=pending,
                                reason=msg,
                            )
                            self._notify_logistics_customs_missing(order_no, msg)
                            self._send_text_reply(
                                f"⚠️ {msg}\n已通知物流人员补全清关资料，补全后将继续处理。",
                                incoming_message
                            )
                            return
                        results.append((order_no, False, msg))
                        continue
                    customs_added_orders.append(order_no)
                    results.append((order_no, True, f"{order_no}: 清关资料已提交"))

                if customs_added_orders:
                    try:
                        download_results = portal.download_customs_files_batch(
                            customs_added_orders,
                            download_dir=customs_download_dir,
                        )
                    except Exception as exc:
                        self.logger.error(f"清关文件批量下载失败: {exc}", exc_info=True)
                        download_results = [(False, f"清关文件批量下载异常: {exc}")]
                else:
                    download_results = []

        except Exception as exc:
            self.logger.error(f"生成清关资料失败: {exc}", exc_info=True)
            self._send_text_reply(
                f"❌ 生成清关资料失败: {exc}",
                incoming_message
            )
            return

        customs_file_results = process_customs_files(customs_download_dir)

        # 压缩清关资料文件夹（仅共享盘场景）
        if shared_folder_path and registry_code:
            try:
                from Common.Utils.file_archive import zip_folder
                zip_dir = os.path.join(shared_folder_path, f"{registry_code} 报关资料&发票")
                zip_path = os.path.join(shared_folder_path, f"{registry_code} 报关资料&发票.zip")
                if os.path.isdir(zip_dir):
                    zip_folder(zip_dir, zip_path)
                    try:
                        from Common.api import send_group_text
                        from Common.api.lingxing_client import LingXingClient
                        from openpyxl import load_workbook

                        # Find customs Excel in zip_dir
                        excel_files = [
                            os.path.join(zip_dir, name)
                            for name in os.listdir(zip_dir)
                            if name.lower().endswith((".xlsx", ".xls")) and not name.startswith("~$")
                        ]
                        customs_excel = None
                        if excel_files:
                            preferred = [p for p in excel_files if "报关资料" in os.path.basename(p)]
                            candidates = preferred or excel_files
                            customs_excel = max(candidates, key=lambda p: os.path.getmtime(p))

                        total_box_num = ""
                        if customs_excel:
                            try:
                                wb = load_workbook(customs_excel, data_only=True)
                                if "明细单" in wb.sheetnames:
                                    ws = wb["明细单"]
                                    for row in range(1, (ws.max_row or 1) + 1):
                                        cell_val = ws.cell(row=row, column=1).value
                                        if cell_val and "合计" in str(cell_val):
                                            total_box_num = ws.cell(row=row, column=8).value
                                            break
                                wb.close()
                            except Exception:
                                total_box_num = ""

                        date_text = LingXingClient.get_next_saturday()
                        box_text = str(total_box_num).strip() if total_box_num not in (None, "") else "未知"
                        notify_text = f"{registry_code} {box_text}箱 {date_text} 平谊提货"

                        at_mobiles = []
                        warehouse_mobiles = getattr(config, "AGL_WAREHOUSE_AT_MOBILES", [])
                        doc_mobiles = getattr(config, "AGL_DOC_AT_MOBILES", [])
                        at_mobiles.extend([m for m in warehouse_mobiles if m])
                        at_mobiles.extend([m for m in doc_mobiles if m])
                        send_group_text(notify_text, at_mobiles=at_mobiles, at_all=False)
                    except Exception as exc:
                        self.logger.error(f"清关资料群通知失败: {exc}")
            except Exception as exc:
                self.logger.error(f"清关资料压缩失败: {exc}")

        # 汇总结果
        success_count = sum(1 for _, success, _ in results if success)
        fail_count = len(results) - success_count

        result_lines = []
        for _, success, msg in results:
            status_icon = "✅" if success else "❌"
            result_lines.append(f"{status_icon} {msg}")

        dl_success = sum(1 for ok, _ in download_results) if download_results is not None else 0
        dl_fail = len(download_results) - dl_success if download_results is not None else 0
        dl_lines = []
        for ok, msg in download_results:
            status_icon = "✅" if ok else "❌"
            dl_lines.append(f"{status_icon} {msg}")
        if not dl_lines:
            dl_lines.append("ℹ️ 无可下载的清关文件（未成功添加清关资料）")

        customs_file_summary = "\n".join(customs_file_results) if customs_file_results else "无"

        summary = (
            f"📊 清关资料处理完成\n\n"
            f"成功: {success_count} / 失败: {fail_count}\n\n"
            f"详细结果：\n" + "\n".join(result_lines) + "\n\n"
            "📥 清关文件下载结果\n"
            f"成功: {dl_success} / 失败: {dl_fail}\n\n"
            + "\n".join(dl_lines)
            + "\n\n🧾 清关文件重命名结果（按运单信息!M8）\n"
            + customs_file_summary
        )
        self._send_text_reply(summary, incoming_message)

        self._reset_workflow()
        self.logger.info("清关资料处理完成，流程结束")

    
    def _looks_like_lcl_packing_result(self, file_path: str) -> bool:
        try:
            xl = pd.ExcelFile(file_path)
            return "拼箱计算结果" in xl.sheet_names
        except Exception:
            return False

    async def _handle_ops_manual_packing_upload(
        self, incoming_message, sender_id: str, packing_file_path: str
    ) -> None:
        """运营直接上传拼箱计算结果：无物流转发也可建单；有任务则更新数据。"""
        try:
            self._send_text_reply("📥 收到拼箱数据，正在校验并生成 Amazon 模板…", incoming_message)
            result_path, template_path = self._apply_modified_packing_result(packing_file_path)

            media_id, uploaded_file_name = self._upload_stream_file(result_path, incoming_message)
            self.dingtalk_api.send_file_message(
                sender_id, media_id, uploaded_file_name, webhook=incoming_message.session_webhook
            )
            template_media_id, template_file_name = self._upload_stream_file(template_path, incoming_message)
            self.dingtalk_api.send_file_message(
                sender_id, template_media_id, template_file_name, webhook=incoming_message.session_webhook
            )

            sns_from_name = re.findall(r"\bSP[0-9A-Za-z]+\b", os.path.basename(result_path), flags=re.I)
            shipping_numbers = list(self.state_manager.get_shipping_numbers() or [])
            for sn in sns_from_name:
                u = sn.upper()
                if u not in {str(x).upper() for x in shipping_numbers}:
                    shipping_numbers.append(u)
            job = {
                "logistics_file_path": None,
                "packing_result_path": result_path,
                "amazon_template_path": template_path,
                "shipping_numbers": shipping_numbers,
                "logistics_user_id": self.state_manager.get_logistics_user_id() or "",
                "conversation_id": getattr(incoming_message, "conversation_id", None),
                "workflow_folder_path": self.state_manager.get_workflow_folder_path(),
                "registry_code": self.state_manager.get_full_state().get("registry_code"),
                "shared_folder_path": self.state_manager.get_full_state().get("shared_folder_path"),
                "status": "WAIT_AMAZON",
                "source": "ops_manual",
            }
            keys = self.state_manager.shipment_keys_from_job(job)
            dropped = self.state_manager.drop_queue_jobs_overlapping(str(sender_id), keys)
            drop_note = ""
            if dropped:
                drop_note = f"\n已丢弃队列中与本单重复的物流转发：{', '.join(dropped)}（以您上传为准）。"

            # 有进行中任务则更新路径（运营优先覆盖物流转发）；否则新建 active
            existing = (self.state_manager.state.get("ops_active") or {}).get(str(sender_id))
            if existing:
                existing = dict(existing)
                existing["packing_result_path"] = result_path
                existing["amazon_template_path"] = template_path
                existing["workflow_folder_path"] = job["workflow_folder_path"]
                existing["shipping_numbers"] = shipping_numbers or existing.get("shipping_numbers") or []
                existing["status"] = "WAIT_AMAZON"
                existing["source"] = "ops_manual"
                self.state_manager.activate_ops_job(str(sender_id), existing)
                self._send_text_reply(
                    "✅ 已用您上传的拼箱数据更新当前任务，并重新生成 Amazon 模板。\n"
                    "与物流转发重复时以您上传为准。"
                    f"{drop_note}\n"
                    "可继续上传 Amazon「包装箱包装信息」文件，或回复【模板2】。",
                    incoming_message,
                )
            else:
                self.state_manager.activate_ops_job(str(sender_id), job)
                self._send_text_reply(
                    "【分仓拼箱】已登记您上传的拼箱数据（无需物流转发）。\n"
                    "已发送拼箱结果与 Amazon 模板1；如需新版格式回复【模板2】。\n"
                    "上传卖家后台包装信息表后，机器人将自动填写并回传。"
                    f"{drop_note}",
                    incoming_message,
                )
            self.logger.info("lcl ops manual packing ops=%s packing=%s", sender_id, result_path)
        except Exception as e:
            self.logger.error(f"运营自助拼箱上传失败: {e}", exc_info=True)
            self._send_text_reply(
                f"❌ 拼箱数据校验失败：{e}\n"
                "请确认文件含「拼箱计算结果」表及必要列后重试。",
                incoming_message,
            )

    async def _handle_operation_file(self, incoming_message, sender_id: str,
                                     file_name: str, download_code: str):
        """处理运营人员上传：拼箱结果(自助) 或 Amazon包装信息"""
        try:
            workflow_folder = self.state_manager.get_workflow_folder_path()
            if not workflow_folder or not os.path.isdir(workflow_folder):
                self.logger.warning("找不到现有流程文件夹，自动创建新的存储目录")
                workflow_folder = self._create_workflow_folder()
                self.state_manager.update_workflow_folder_path(workflow_folder)

            robot_code = incoming_message.robot_code
            file_info = incoming_message.extensions.get('content', {}) if hasattr(incoming_message, 'extensions') else None
            downloaded_path = self.dingtalk_api.download_file(
                download_code, file_name, workflow_folder,
                file_info=file_info,
                robot_code=robot_code
            )

            # 运营直接上传拼箱计算结果
            if self._looks_like_lcl_packing_result(downloaded_path):
                await self._handle_ops_manual_packing_upload(
                    incoming_message, sender_id, downloaded_path
                )
                return

            # 绑定该运营 active job（物流可能已开始下一单）
            self.state_manager.bind_ops_job_to_state(sender_id)
            # 检查状态
            if not self.state_manager.is_waiting_for_operation(sender_id):
                self._send_text_reply(
                    "⚠️  当前没有待处理的流程\n\n"
                    "可直接上传含「拼箱计算结果」的拼箱 Excel 开始任务；\n"
                    "或等待物流转发后再上传 Amazon 包装信息文件。",
                    incoming_message
                )
                return
            
            # 发送处理中提示
            self._send_text_reply("📥 收到Amazon包装信息文件，正在处理中，请稍候...", incoming_message)
            
            amazon_file_path = downloaded_path
            
            # 2. 获取拼箱结果（优先该运营 active job）
            packing_result_path = self.state_manager.get_packing_result_path(sender_id)
            
            if not os.path.exists(packing_result_path):
                raise FileNotFoundError(f"拼箱结果文件不存在: {packing_result_path}")
            
            timestamp_suffix = self._extract_timestamp_suffix(packing_result_path)
            shipping_numbers = self.state_manager.get_shipping_numbers()
            if not shipping_numbers:
                logistics_file_path = self.state_manager.get_logistics_file_path()
                if logistics_file_path and os.path.exists(logistics_file_path):
                    shipping_numbers = self._extract_shipping_numbers(logistics_file_path)
            
            # 3. 读取拼箱结果
            self.logger.info("开始处理Amazon包装信息...")
            merge_summary = pd.read_excel(packing_result_path, sheet_name='拼箱计算结果')
            merge_summary['小组名称'] = merge_summary['小组名称'].ffill()
            merge_summary['单箱重量'] = merge_summary['单箱重量'].ffill()
            merge_summary['单箱理论长'] = merge_summary['单箱理论长'].ffill()
            merge_summary['单箱理论高'] = merge_summary['单箱理论高'].ffill()
            merge_summary['单箱理论宽'] = merge_summary['单箱理论宽'].ffill()
            
            # 4. 创建临时处理器来调用insert方法
            temp_processor = PackingBoxProcessor(
                input_file_path=packing_result_path,
                output_file_path=packing_result_path  # 这里不会用到
            )
            
            # 5. 插入数据到Amazon文件（会直接修改原文件）
            temp_processor.insert_data_to_amazon_packaging(amazon_file_path, merge_summary)
            
            self.logger.info(f"Amazon文件处理完成: {amazon_file_path}")
            
            # 5.1 保存最终结果至专用目录
            # 运营侧希望收到原始命名的模板文件，故沿用上传时的文件名
            original_file_name = file_name
            final_result_path = os.path.join(workflow_folder, original_file_name)
            if os.path.abspath(amazon_file_path) != os.path.abspath(final_result_path):
                shutil.copy2(amazon_file_path, final_result_path)
            else:
                self.logger.info("最终结果已在目标位置，无需复制")
            self.logger.info(f"Amazon最终结果文件已保存: {final_result_path}")
            
            # 6. 上传处理后的文件
            self.logger.info("开始上传处理后的文件...")
            media_id, uploaded_file_name = self._upload_stream_file(final_result_path, incoming_message)
            
            # 7. 发送结果文件给运营人员
            session_webhook = incoming_message.session_webhook
            self.dingtalk_api.send_file_message(sender_id, media_id, uploaded_file_name, webhook=session_webhook)
            
            # 7.1 抄送Amazon最终结果给OTHER_USERS
            self._send_excel_copy_to_others(media_id, uploaded_file_name, "Amazon最终结果")
            
            # 8. 发送完成提示并询问是否删除发货单
            self._send_text_reply(
                "✅ Amazon包装信息处理完成！\n"
                "📋 处理结果已发送给您，请查收。\n\n"
                "是否删除发货单？回复【删除】或【不删除】。",
                incoming_message
            )

            # 9. 更新状态，等待运营确认删除
            self.state_manager.set_waiting_for_delete_confirmation(sender_id)

            self.logger.info("运营工作流完成，等待删除确认")
            
        except Exception as e:
            self.logger.error(f"处理运营文件时发生错误: {e}", exc_info=True)
            self._send_text_reply(
                f"❌ 处理文件时发生错误: {str(e)}\n\n"
                f"请检查文件格式是否正确，或联系技术支持。",
                incoming_message
            )
            self._notify_tech_support(f"处理运营文件失败：{str(e)}")
            # 发生错误时不重置状态，允许运营重新上传
    
    def _send_help_message(self, incoming_message, user_role: str):
        """发送帮助信息"""
        if user_role == 'logistics':
            help_text = (
                "📖 物流人员使用说明\n\n"
                "1️⃣ 上传发货单Excel文件\n"
                "   - 文件必须包含\"发货单详情\"和\"装箱信息\"两个工作表\n"
                "   - 机器人会自动进行拼箱处理\n\n"
                "2️⃣ 确认处理结果\n"
                "   - 收到处理结果后，回复\"确认\"将结果转发给运营\n"
                "   - 如需重新处理，回复\"重置\"后重新上传\n\n"
                "3️⃣ 等待运营处理\n"
                "   - 运营完成后会收到通知\n\n"
                "💡 其他命令：\n"
                "- 发送\"状态\"查看当前流程状态\n"
                "- 发送\"重置\"取消当前流程"
            )
        else:  # operation
            help_text = (
                "📖 运营人员使用说明\n\n"
                "1️⃣ 等待物流确认\n"
                "   - 物流人员上传发货单并确认后\n"
                "   - 您会收到通知消息\n\n"
                "2️⃣ 获取拼箱结果\n"
                "   - 机器人会主动私聊推送拼箱结果\n"
                "   - 如需重新获取，可在机器人会话发送\"文件\"或\"获取文件\"\n"
                "   - 下载并查看拼箱结果\n"
                "   - 默认收到Amazon模板1，如需新版格式可发送\"模板2\"重新生成\n\n"
                "3️⃣ 上传Amazon包装信息文件\n"
                "   - 查看拼箱结果后\n"
                "   - 上传Amazon包装信息Excel文件\n"
                "   - 机器人会自动插入拼箱数据\n\n"
                "4️⃣ 获取最终结果\n"
                "   - 处理完成后会收到最终文件\n\n"
                "5️⃣ 确认是否删除发货单\n"
                "   - 收到最终文件后回复\"删除\"或\"不删除\"\n\n"
                "💡 常用命令：\n"
                "- \"文件\" 或 \"获取文件\" - 获取拼箱结果\n"
                "- \"模板2\" 或 \"生成模板2\" - 重新生成新版Amazon模板\n"
                "- \"状态\" - 查看当前流程状态\n"
                "- \"帮助\" - 查看使用说明"
            )
        
        self._send_text_reply(help_text, incoming_message)
    
    def _send_status_message(self, incoming_message):
        """发送当前状态信息"""
        state = self.state_manager.get_full_state()
        status = state.get('status', 'UNKNOWN')
        
        status_text, status_hint = self._get_status_detail(status)
        
        message = f"📊 当前状态：{status_text}\n\n"
        
        if status_hint:
            message += f"➡️ 下一步：{status_hint}\n\n"
        
        if status != WorkflowState.IDLE:
            if state.get('created_at'):
                message += f"⏰ 流程开始时间：{state['created_at'][:19]}\n"
            if state.get('updated_at'):
                message += f"⏰ 最后更新时间：{state['updated_at'][:19]}\n"

        # 关键变量展示
        message += (
            "\n🔧 关键变量：\n"
            f"发票号：{state.get('registry_code') or '未设置'}\n"
            f"共享盘路径：{state.get('shared_folder_path') or '未设置'}\n"
            f"流程目录：{state.get('workflow_folder_path') or '未设置'}\n"
            f"物流用户：{state.get('logistics_user_id') or '未设置'}\n"
            f"运营用户：{state.get('operation_user_id') or '未设置'}\n"
            f"发货单号(原)：{state.get('shipping_numbers') or '未设置'}\n"
            f"领星发货单号：{', '.join(state.get('lingxing_shipment_numbers') or []) or '未设置'}\n"
            f"待补全清关：{', '.join(state.get('pending_customs_shipments') or []) or '无'}\n"
            f"物流文件已收：{len(state.get('logistics_files_received') or [])}"
            f"/{state.get('logistics_files_expected') or '未设置'}\n"
        )
        
        self._send_text_reply(message, incoming_message)

    def _is_tech_user(self, incoming_message) -> bool:
        """判断是否技术支持用户"""
        tech_users = set(getattr(config, 'TECHNOLOGY_USERS', []) or [])
        if not tech_users:
            return False
        candidate_ids = set()
        for attr in ('sender_id', 'sender_user_id', 'sender_staff_id'):
            value = getattr(incoming_message, attr, None)
            if value:
                candidate_ids.add(value)
        normalized = self._normalize_recipient_id(getattr(incoming_message, 'sender_id', None))
        if normalized:
            candidate_ids.add(normalized)
        return any(uid in tech_users for uid in candidate_ids)

    def _handle_force_jump_command(self, incoming_message, text_content: str):
        """处理/跳转命令"""
        if not self._is_tech_user(incoming_message):
            self._send_text_reply("⚠️ 仅技术支持可使用跳转指令。", incoming_message)
            return
        parts = text_content.split(maxsplit=1)
        if len(parts) < 2:
            self._send_text_reply("⚠️ 用法：/跳转 阶段名（分5仓拼箱/领星编辑/生成清关资料）", incoming_message)
            return
        stage_alias = parts[1].strip()
        try:
            new_status = self.state_manager.force_jump(stage_alias)
        except Exception as exc:
            self._send_text_reply(f"❌ 跳转失败：{exc}", incoming_message)
            return

        required = []
        if new_status == WorkflowState.WAIT_LOGISTICS_FILES:
            required = ["发票号", "共享盘路径", "领星发货单号"]
        elif new_status == WorkflowState.WAIT_DELETE_CONFIRMATION:
            required = ["发货单号(原)"]
        elif new_status == WorkflowState.WAIT_SHIPMENT_NUMBERS:
            required = ["领星发货单号(可选)"]

        req_text = "、".join(required) if required else "无"
        self._send_text_reply(
            f"✅ 已跳转至【{stage_alias}】阶段。\n"
            f"当前状态：{new_status}\n"
            f"依赖项：{req_text}\n"
            "可用 /设置 Key=Value 注入数据。",
            incoming_message
        )

    def _handle_set_command(self, incoming_message, text_content: str):
        """处理/设置命令"""
        if not self._is_tech_user(incoming_message):
            self._send_text_reply("⚠️ 仅技术支持可使用设置指令。", incoming_message)
            return
        payload = text_content[len("/设置"):].strip()
        if not payload:
            self._send_text_reply("⚠️ 用法：/设置 发票号=25L03（多项请换行）", incoming_message)
            return

        updates = {}
        for line in payload.splitlines():
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip()
            if not key:
                continue

            if key in ("发票号",):
                updates["registry_code"] = value
            elif key in ("共享盘路径", "共享盘", "路径"):
                updates["shared_folder_path"] = value
            elif key in ("流程目录", "流程文件夹", "本地目录"):
                updates["workflow_folder_path"] = value
            elif key in ("发货单号", "领星发货单号", "单号", "发货单号列表"):
                numbers = [v for v in re.split(r"[\\s,;]+", value) if v]
                updates["lingxing_shipment_numbers"] = numbers
            elif key in ("原发货单号",):
                updates["shipping_numbers"] = value

        if not updates:
            self._send_text_reply("⚠️ 未识别到可设置的Key。", incoming_message)
            return

        self.state_manager.inject_data(updates)
        summary = "\n".join([f"{k} = {v}" for k, v in updates.items()])
        self._send_text_reply(f"✅ 已设置：\n{summary}", incoming_message)

    def _get_status_detail(self, status: str):
        """返回状态描述和下一步提示"""
        status_detail_map = {
            WorkflowState.IDLE: (
                "🟢 空闲 - 可以开始新的流程",
                "物流上传发货单后将自动进入流程"
            ),
            WorkflowState.LOGISTICS_UPLOADED: (
                "🟡 已生成拼箱结果，等待物流确认",
                "物流可回复\"确认\"后选择运营，或回复\"重置\"重新上传"
            ),
            WorkflowState.WAIT_OPS_SELECT: (
                "🟡 物流已确认，等待选择运营",
                "物流回复序号或姓名选择要转发的运营"
            ),
            WorkflowState.LOGISTICS_CONFIRMED: (
                "🟡 已转发给运营，正等待运营上传Amazon包装信息",
                "运营上传Amazon文件后系统会自动处理"
            ),
            WorkflowState.OPERATION_UPLOADED: (
                "🟡 运营文件已上传，系统正在生成最终结果",
                "请稍候，机器人会在生成完成后推送最终文件"
            ),
            WorkflowState.WAIT_DELETE_CONFIRMATION: (
                "🟡 已发送最终文件，等待运营确认是否删除发货单",
                "运营请回复\"删除\"或\"不删除\""
            ),
            WorkflowState.WAIT_SHIPMENT_NUMBERS: (
                "🟡 等待运营输入领星发货单号",
                "运营请回复发货单号（每行一个）或回复\"跳过\""
            ),
            WorkflowState.WAIT_CUSTOMS_INFO: (
                "🟡 等待物流补全清关资料",
                "物流补全后回复\"已补全 SPxxxx\""
            ),
            WorkflowState.WAIT_LOGISTICS_FILES: (
                "🟡 等待物流上传发货单/报关资料文件",
                "物流上传发货单和报关资料Excel文件"
            ),
        }
        return status_detail_map.get(status, (f"未知状态: {status}", ""))

    def _create_workflow_folder(self) -> str:
        """创建流程专用的Excel文件夹"""
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        base_name = f"workflow_{timestamp}"
        folder_name = base_name
        counter = 1
        folder_path = os.path.join(config.EXCEL_FILES_DIR, folder_name)
        while os.path.exists(folder_path):
            folder_name = f"{base_name}_{counter}"
            folder_path = os.path.join(config.EXCEL_FILES_DIR, folder_name)
            counter += 1
        os.makedirs(folder_path, exist_ok=True)
        self.logger.info(f"📁 已创建流程文件夹: {folder_path}")
        return folder_path

    def _build_result_file_name(self, logistics_file_path: str, timestamp_suffix: str) -> str:
        """根据发货单号生成拼箱结果文件名，若缺失则抛出错误"""
        order_numbers = self._extract_order_numbers(logistics_file_path)
        sanitized_numbers = [self._sanitize_filename_part(num) for num in order_numbers]
        sanitized_numbers = [num for num in sanitized_numbers if num]
        if not sanitized_numbers:
            error_message = (
                "无法从发货单中读取有效的「发货单号」，"
                "请确认 Excel 中的《发货单详情》工作表包含该列并填写完整。"
            )
            self.logger.error(error_message)
            raise ValueError(error_message)

        base_name = " ".join(sanitized_numbers)
        return f"{base_name} 拼箱数据.xlsx"

    def _extract_order_numbers(self, file_path: str) -> List[str]:
        """读取发货单号列并按出现顺序去重"""
        try:
            df = pd.read_excel(file_path, sheet_name="发货单详情")
        except Exception as exc:
            self.logger.warning(f"读取发货单详情失败，使用时间戳命名: {exc}")
            return []
        if "发货单号" not in df.columns:
            self.logger.warning("发货单详情中缺少\"发货单号\"列，使用时间戳命名")
            return []
        order_series = df["发货单号"].dropna().map(lambda v: str(v).strip())
        seen = set()
        order_numbers = []
        for value in order_series:
            if not value or value.lower() == 'nan':
                continue
            if value in seen:
                continue
            seen.add(value)
            order_numbers.append(value)
        return order_numbers

    def _sanitize_filename_part(self, text: str) -> str:
        """移除文件名不允许的字符并压缩空白"""
        cleaned = re.sub(r"[\\/:*?\"<>|]", "", text)
        cleaned = re.sub(r"\s+", " ", cleaned)
        return cleaned.strip()

    def _notify_tech_support(self, detail_message: str):
        """向技术支持用户推送错误通知"""
        tech_users = getattr(config, 'TECHNOLOGY_USERS', [])
        if not tech_users:
            self.logger.warning(f"技术支持用户未配置，无法发送通知: {detail_message}")
            return

        notification = (
            "⚠️ 钉钉拼箱机器人异常提醒\n"
            f"{detail_message}"
        )

        for tech_id in tech_users:
            corp_user_id = self._normalize_recipient_id(tech_id)
            if not corp_user_id:
                self.logger.warning(f"无法解析技术支持用户ID: {tech_id}")
                continue
            try:
                self.dingtalk_api.send_text_message(corp_user_id, notification)
            except Exception as exc:
                self.logger.error(f"发送技术支持通知失败 ({tech_id}): {exc}")

    def _cleanup_workflow_files(self, folder_path: Optional[str] = None):
        """删除本轮流程的工作目录"""
        target_folder = folder_path or self.state_manager.get_workflow_folder_path()
        if not target_folder:
            return
        if os.path.isdir(target_folder):
            try:
                shutil.rmtree(target_folder)
                self.logger.info(f"🧹 已清理流程文件夹: {target_folder}")
            except Exception as exc:
                self.logger.warning(f"清理流程文件夹失败 {target_folder}: {exc}")
    
    def _handle_reset_command(self, incoming_message, user_role: str):
        """处理重置命令：物流只清自己会话，不影响运营进行中/排队任务。"""
        # 只有物流人员可以重置
        if user_role != 'logistics':
            self._send_text_reply(
                "⚠️  只有物流人员可以执行重置操作",
                incoming_message
            )
            return
        
        if self.state_manager.can_logistics_upload() and self.state_manager.logistics_phase() == "IDLE":
            # 无物流进行中；若仅有运营任务在跑，也提示不伤运营
            active_n = len(self.state_manager.state.get("ops_active") or {})
            if active_n:
                self._send_text_reply(
                    "ℹ️  当前没有物流进行中的上传/确认。\n"
                    f"运营侧仍有 {active_n} 单在处理，不会被重置。\n"
                    "可直接上传新发货单。",
                    incoming_message,
                )
            else:
                self._send_text_reply(
                    "ℹ️  当前已是空闲状态，无需重置",
                    incoming_message,
                )
            return

        # 只清物流会话目录（若该目录仍被运营 active 引用则跳过删除）
        folder = self.state_manager.get_workflow_folder_path()
        protected = set()
        for job in (self.state_manager.state.get("ops_active") or {}).values():
            if isinstance(job, dict) and job.get("workflow_folder_path"):
                protected.add(job["workflow_folder_path"])
        for jobs in (self.state_manager.state.get("ops_queue") or {}).values():
            for job in jobs or []:
                if isinstance(job, dict) and job.get("workflow_folder_path"):
                    protected.add(job["workflow_folder_path"])
        if folder and folder not in protected:
            self._cleanup_workflow_files(folder)

        self.state_manager.reset_logistics_only()
        
        self._send_text_reply(
            "✅ 物流侧已重置，可以开始新的流程。\n"
            "运营人员正在处理/排队的任务不受影响。\n"
            "请上传发货单Excel文件开始处理。",
            incoming_message
        )
        
        self.logger.info("物流会话已手动重置（运营流程保留）")
    
    async def _handle_get_file(self, incoming_message, user_role: str):
        """处理获取拼箱结果文件"""
        try:
            # 运营人员可以获取，物流人员在测试时也可以
            if user_role not in ['operation', 'logistics']:
                self._send_text_reply(
                    "⚠️  您没有权限获取拼箱结果文件",
                    incoming_message
                )
                return
            
            # 检查状态
            if not self.state_manager.is_waiting_for_operation():
                self._send_text_reply(
                    "⚠️  当前没有可获取的拼箱结果\n\n"
                    "请等待物流人员上传发货单并确认。",
                    incoming_message
                )
                return
            
            # 获取拼箱结果文件路径
            packing_result_path = self.state_manager.get_packing_result_path()
            
            if not os.path.exists(packing_result_path):
                self._send_text_reply(
                    "❌ 拼箱结果文件不存在，请联系管理员",
                    incoming_message
                )
                return
            
            # 发送处理中提示
            self._send_text_reply("📥 正在准备拼箱结果文件，请稍候...", incoming_message)
            
            # 上传文件
            self.logger.info("开始上传拼箱结果文件...")
            media_id, file_name = self._upload_stream_file(packing_result_path, incoming_message)
            
            # 发送文件给运营人员（使用webhook回复）
            session_webhook = incoming_message.session_webhook
            self.dingtalk_api.send_file_message(
                incoming_message.sender_id, 
                media_id, 
                file_name, 
                webhook=session_webhook
            )
            
            # 抄送拼箱结果给OTHER_USERS
            self._send_excel_copy_to_others(media_id, file_name, "拼箱结果（运营获取）")
            
            # 如存在Amazon模板，同步发送
            template_status_line = "⚠️ 尚未生成 Amazon 发货模板，请联系物流人员重新触发。"
            template_path = self.state_manager.get_amazon_template_path()
            if template_path and os.path.exists(template_path):
                template_media_id, template_file_name = self._upload_stream_file(template_path, incoming_message)
                self.dingtalk_api.send_file_message(
                    incoming_message.sender_id,
                    template_media_id,
                    template_file_name,
                    webhook=session_webhook
                )
                # 抄送Amazon模板给OTHER_USERS
                self._send_excel_copy_to_others(template_media_id, template_file_name, "Amazon发货模板（运营获取）")
                template_status_line = f"📑 Amazon 发货模板已发送：{template_file_name}"
            
            # 发送说明（文本提示）
            self._send_text_reply(
                f"✅ 拼箱结果已发送：{file_name}\n\n"
                f"{template_status_line}\n\n"
                f"📄 请下载查看并准备 Amazon 包装信息文件。\n"
                f"默认使用模板1；如需新版格式，请回复【模板2】重新生成。\n"
                f"准备好后直接在此聊天回复【上传】并发送文件即可开始处理。",
                incoming_message
            )
            
            self.logger.info(f"运营人员 {incoming_message.sender_nick} 已获取拼箱结果文件")
            
        except Exception as e:
            self.logger.error(f"获取文件失败: {e}", exc_info=True)
            self._send_text_reply(f"❌ 获取文件失败: {str(e)}", incoming_message)
            self._notify_tech_support(f"运营获取拼箱结果失败：{str(e)}")

    async def _handle_generate_template_v2(self, incoming_message, user_role: str):
        """按需基于当前拼箱结果生成第二版Amazon模板。"""
        try:
            if not self._is_operation_or_dual(incoming_message, user_role):
                self._send_text_reply("⚠️ 只有运营人员可以生成模板2。", incoming_message)
                return

            current_status = self.state_manager.get_status()
            allowed_statuses = {
                WorkflowState.LOGISTICS_CONFIRMED,
                WorkflowState.WAIT_DELETE_CONFIRMATION,
                WorkflowState.WAIT_SHIPMENT_NUMBERS,
                WorkflowState.WAIT_LOGISTICS_FILES,
            }
            if current_status not in allowed_statuses:
                self._send_text_reply(
                    "⚠️ 当前没有可生成模板2的流程。\n"
                    "请等待物流确认并收到拼箱结果后再发送【模板2】。",
                    incoming_message
                )
                return

            template_source = getattr(config, "AMAZON_TEMPLATE_SOURCE_FILE_V2", None)
            if not template_source or not os.path.exists(template_source):
                self._send_text_reply(
                    f"❌ 模板2文件不存在，请确认路径：{template_source}",
                    incoming_message
                )
                return

            packing_result_path = self.state_manager.get_packing_result_path()
            if not packing_result_path or not os.path.exists(packing_result_path):
                self._send_text_reply("❌ 当前流程缺少拼箱结果文件，无法生成模板2。", incoming_message)
                return

            workflow_folder = self.state_manager.get_workflow_folder_path()
            if not workflow_folder or not os.path.isdir(workflow_folder):
                workflow_folder = self._create_workflow_folder()
                self.state_manager.update_workflow_folder_path(workflow_folder)

            self._send_text_reply("📑 正在根据当前拼箱结果生成模板2，请稍候...", incoming_message)

            merge_summary = pd.read_excel(packing_result_path, sheet_name='拼箱计算结果')
            for col in ['小组名称', '单箱重量', '单箱理论长', '单箱理论高', '单箱理论宽']:
                if col in merge_summary.columns:
                    merge_summary[col] = merge_summary[col].ffill()

            current_template_path = self.state_manager.get_amazon_template_path()
            if current_template_path:
                template_file_name = os.path.basename(current_template_path)
            else:
                template_file_name = os.path.basename(config.AMAZON_TEMPLATE_SOURCE_FILE)
            template_path = os.path.join(workflow_folder, template_file_name)
            temp_processor = PackingBoxProcessor(
                input_file_path=packing_result_path,
                output_file_path=packing_result_path
            )
            temp_processor.create_amazon_template(
                template_path,
                merge_summary,
                template_source=template_source,
            )

            media_id, uploaded_file_name = self._upload_stream_file(template_path, incoming_message)
            self.dingtalk_api.send_file_message(
                incoming_message.sender_id,
                media_id,
                uploaded_file_name,
                webhook=incoming_message.session_webhook,
            )
            self._send_excel_copy_to_others(media_id, uploaded_file_name, "Amazon发货模板2")
            self._send_text_reply(
                f"✅ 模板2已生成并发送：{uploaded_file_name}",
                incoming_message
            )
        except Exception as e:
            self.logger.error(f"生成模板2失败: {e}", exc_info=True)
            self._send_text_reply(f"❌ 生成模板2失败: {str(e)}", incoming_message)
            self._notify_tech_support(f"生成Amazon模板2失败：{str(e)}")
    
    def _extract_shipping_numbers(self, file_path: str) -> Optional[str]:
        """读取发货单详情表中的发货单号，去重后按顺序拼接"""
        sheet_name = '发货单详情'
        column_name = '发货单号'
        try:
            details_df = pd.read_excel(file_path, sheet_name=sheet_name, dtype={column_name: str})
        except Exception as exc:
            self.logger.warning(f"读取发货单号失败（{file_path}）: {exc}")
            return None

        if column_name not in details_df.columns:
            self.logger.warning(f"{sheet_name} 缺少列: {column_name}，无法生成自定义命名")
            return None

        raw_numbers = details_df[column_name].dropna().astype(str).map(lambda x: x.strip())
        unique_numbers: List[str] = []
        for number in raw_numbers:
            if number and number not in unique_numbers:
                unique_numbers.append(number)

        return ' '.join(unique_numbers) if unique_numbers else None

    def _parse_shipment_nos(self, shipping_numbers: Optional[str]) -> List[str]:
        """将发货单号字符串解析为列表"""
        if not shipping_numbers:
            return []
        parts = re.split(r'[\s,;]+', str(shipping_numbers).strip())
        return [part for part in parts if part]

    def _build_amazon_filename(self, shipping_numbers: Optional[str], template_index: int,
                               fallback_identifier: str) -> str:
        """根据发货单号或回退标识构造Amazon文件名"""
        identifier = (shipping_numbers or '').strip()
        if not identifier:
            identifier = str(fallback_identifier).strip()
        identifier = ' '.join(identifier.split())
        return f"{identifier} Amazon模版{template_index}.xlsx"

    def _extract_timestamp_suffix(self, file_path: str) -> str:
        """从文件名中提取时间戳后缀，若失败则回退到当前时间"""
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        candidate = base_name.split('_')[-1] if '_' in base_name else ''
        if candidate.isdigit():
            return candidate
        return str(int(time.time()))
    
    def _upload_stream_file(self, file_path: str, incoming_message):
        """封装上传逻辑，自动携带robot_code"""
        robot_code = getattr(incoming_message, 'robot_code', None)
        return self.dingtalk_api.upload_file(
            file_path,
            channel='openapi',
            robot_code=robot_code
        )
    
    def _send_text_reply(self, text: str, incoming_message):
        """发送文本回复"""
        try:
            self.reply_text(text, incoming_message)
        except Exception as e:
            self.logger.error(f"发送文本回复失败: {e}")

    def _is_customs_missing_message(self, msg: str) -> bool:
        """判断清关缺失信息提示"""
        if not msg:
            return False
        return any(
            key in msg
            for key in ("收货仓库信息不完整", "清关产品信息不完整")
        )

    def _notify_logistics_customs_missing(self, order_no: str, reason: str):
        """通知物流补全清关资料"""
        logistics_user_id = self.state_manager.get_logistics_user_id()
        if not logistics_user_id:
            self.logger.warning("未找到物流人员ID，无法发送清关补全通知")
            return

        normalized_id = self._normalize_recipient_id(logistics_user_id)
        if not normalized_id:
            self.logger.warning(f"无法获取物流人员 {logistics_user_id} 的userid，未推送清关补全通知")
            return

        message = (
            f"⚠️ 清关资料缺失，请补全后回复：已补全 {order_no}\n\n"
            f"发货单号：{order_no}\n"
            f"缺失原因：{reason}"
        )
        try:
            self.dingtalk_api.send_text_message(normalized_id, message)
        except Exception as exc:
            self.logger.error(f"发送清关补全通知失败: {exc}")
    
    def _send_excel_copy_to_others(self, media_id: str, file_name: str, file_description: str = "Excel文件"):
        """
        抄送Excel文件给OTHER_USERS配置的用户
        
        Args:
            media_id: 文件的media_id
            file_name: 文件名
            file_description: 文件描述（用于日志和通知）
        """
        if not config.OTHER_USERS:
            self.logger.info("未配置OTHER_USERS，跳过Excel文件抄送")
            return
        
        self.logger.info(f"开始抄送{file_description}给OTHER_USERS: {config.OTHER_USERS}")
        
        success_count = 0
        for other_user_id in config.OTHER_USERS:
            try:
                # 转换openId为userId（如果需要）
                corp_user_id = self._normalize_recipient_id(other_user_id)
                if not corp_user_id:
                    self.logger.warning(f"无法获取用户 {other_user_id} 的userid，已跳过抄送")
                    continue
                
                # 发送文件
                self.dingtalk_api.send_file_message(
                    corp_user_id,
                    media_id,
                    file_name
                )
                
                # 发送说明文本
                notification_text = f"📋 {file_description}抄送\n\n文件名：{file_name}"
                self.dingtalk_api.send_text_message(corp_user_id, notification_text)
                
                success_count += 1
                self.logger.info(f"✅ 已抄送{file_description}给用户: {other_user_id}")
                
            except Exception as e:
                self.logger.error(f"❌ 抄送{file_description}给用户 {other_user_id} 失败: {e}")
        
        if success_count > 0:
            self.logger.info(f"📤 {file_description}抄送完成：成功 {success_count}/{len(config.OTHER_USERS)}")
        else:
            self.logger.warning(f"⚠️ {file_description}抄送失败：未成功发送给任何OTHER_USERS")
