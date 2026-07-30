# !/usr/bin/env python
# -*- coding: utf-8 -*-
"""
发货单登记模块 - 根据店铺类型自动登记到不同的登记表和文件夹
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from . import config

# Common 包（登记表）；缺失时登记功能不可用，拼箱主流程仍可运行
_workspace_root = getattr(config, "COMMON_ROOT", None) or Path(__file__).resolve().parents[3]
if str(_workspace_root) not in sys.path:
    sys.path.insert(0, str(_workspace_root))

try:
    from Common.Utils.excel_registry import (  # type: ignore
        RegistryConfig,
        create_next_entry,
        ensure_target_folder,
    )
except ImportError:  # pragma: no cover
    RegistryConfig = None  # type: ignore
    create_next_entry = None  # type: ignore
    ensure_target_folder = None  # type: ignore


# ==================== 配置定义 ====================

# 登记表路径（从 config 读取）
REGISTER_EXCEL_PATH = config.REGISTER_EXCEL_PATH
REGISTER_SHEET = config.REGISTER_SHEET

# EZARC 配置
EZARC_CODE_COL = config.EZARC_CODE_COL
EZARC_REMARK_COL = config.EZARC_REMARK_COL
EZARC_FOLDER_BASE = config.EZARC_FOLDER_BASE

# TOLESA 配置
TOLESA_CODE_COL = config.TOLESA_CODE_COL
TOLESA_REMARK_COL = config.TOLESA_REMARK_COL
TOLESA_FOLDER_BASE = config.TOLESA_FOLDER_BASE


@dataclass
class ShipmentInfo:
    """从发货单提取的信息"""
    shop: str               # 店铺名（第一个单词，如 EZARC）
    shop_full: str          # 完整店铺名
    country: str            # 国家
    transport_method: str   # 运输方式
    
    @property
    def folder_name(self) -> str:
        """生成文件夹名称：国家运输方式 + 平谊（国家和运输方式之间无空格）"""
        return f"{self.country}{self.transport_method} 平谊"
    
    @property
    def remark(self) -> str:
        """生成备注/B或G列内容：国家运输方式（无空格）"""
        return f"{self.country}{self.transport_method}"
    
    def get_shop_type(self) -> Optional[str]:
        """判断店铺类型"""
        shop_upper = self.shop.upper()
        if shop_upper.startswith("EZARC"):
            return "EZARC"
        elif shop_upper.startswith("TOLESA"):
            return "TOLESA"
        return None


def extract_shipment_info(excel_path: str, sheet_name: str = "发货单详情") -> ShipmentInfo:
    """从发货单中提取店铺、国家、运输方式信息
    
    Args:
        excel_path: 发货单 Excel 文件路径
        sheet_name: 工作表名称
        
    Returns:
        ShipmentInfo 对象
        
    Raises:
        ValueError: 如果必需的列不存在或数据为空
    """
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    
    required_cols = ["店铺", "国家", "运输方式"]
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"发货单详情缺少必需列: {', '.join(missing_cols)}")
    
    if df.empty:
        raise ValueError("发货单详情工作表为空")
    
    # 取第一行数据
    first_row = df.iloc[0]
    
    shop_full = str(first_row["店铺"]).strip()
    country = str(first_row["国家"]).strip()
    transport = str(first_row["运输方式"]).strip()
    
    if not shop_full or shop_full == "nan":
        raise ValueError("店铺列数据为空")
    if not country or country == "nan":
        raise ValueError("国家列数据为空")
    if not transport or transport == "nan":
        raise ValueError("运输方式列数据为空")
    
    # 提取店铺名的第一个单词（如 "EZARC NA-US" -> "EZARC"）
    shop = shop_full.split()[0] if shop_full else ""
    
    return ShipmentInfo(
        shop=shop,
        shop_full=shop_full,
        country=country,
        transport_method=transport,
    )


def register_shipment(
    shipment_info: ShipmentInfo,
    excel_path: str = REGISTER_EXCEL_PATH,
    sheet_name: str = REGISTER_SHEET,
) -> Tuple[str, str]:
    """根据店铺类型登记发货单并创建文件夹

    Args:
        shipment_info: 发货单信息
        excel_path: 登记表 Excel 路径
        sheet_name: 工作表名称
        
    Returns:
        Tuple of (new_code, folder_path)
        
    Raises:
        ValueError: 如果店铺类型不支持
    """
    if RegistryConfig is None or ensure_target_folder is None:
        raise RuntimeError("Common.Utils.excel_registry 不可用，请配置 LCL_COMMON_ROOT 与登记表路径")
    shop_type = shipment_info.get_shop_type()

    if shop_type == "EZARC":
        cfg = RegistryConfig(
            excel_path=excel_path,
            sheet_name=sheet_name,
            code_col_index=EZARC_CODE_COL,
            remark_col_index=EZARC_REMARK_COL,
            folder_base_dir=EZARC_FOLDER_BASE,
            default_remark=shipment_info.remark,
        )
    elif shop_type == "TOLESA":
        cfg = RegistryConfig(
            excel_path=excel_path,
            sheet_name=sheet_name,
            code_col_index=TOLESA_CODE_COL,
            remark_col_index=TOLESA_REMARK_COL,
            folder_base_dir=TOLESA_FOLDER_BASE,
            default_remark=shipment_info.remark,
        )
    else:
        raise ValueError(f"不支持的店铺类型: {shipment_info.shop}（完整名称: {shipment_info.shop_full}）")
    
    # 写入Excel（B列或G列）- 使用 remark（不含"平谊"）
    wb = load_workbook(cfg.excel_path)
    ws = wb[cfg.sheet_name]
    
    # 找到最后一个编号并递增
    from Common.Utils.excel_registry import _find_last_value, _increment_code
    
    last_row, last_code = _find_last_value(ws, cfg.code_col_index)
    if last_row is None or last_code is None:
        raise ValueError(f"列 {cfg.code_col_index} 未找到任何编号，无法生成新编号。")
    
    new_code = _increment_code(last_code)
    new_row = last_row + 1
    ws.cell(row=new_row, column=cfg.code_col_index, value=new_code)
    ws.cell(row=new_row, column=cfg.remark_col_index, value=shipment_info.remark)
    wb.save(cfg.excel_path)
    
    # 创建文件夹（使用 folder_name，包含"平谊"）
    folder_path = ensure_target_folder(
        code=new_code,
        remark=shipment_info.folder_name,
        base_dir=cfg.folder_base_dir,
        create=True,
    )

    # Create customs & invoice subfolder inside shared folder.
    customs_folder = os.path.join(folder_path, f"{new_code} 报关资料&发票")
    os.makedirs(customs_folder, exist_ok=True)

    return new_code, folder_path


def process_logistics_shipment(
    logistics_file_path: str,
) -> Tuple[str, str, ShipmentInfo]:
    """处理物流发货单：提取信息、登记编号、创建文件夹
    
    Args:
        logistics_file_path: 发货单 Excel 文件路径
        
    Returns:
        Tuple of (new_code, folder_path, shipment_info)
        
    Raises:
        ValueError: 如果无法处理
    """
    # 1. 提取发货单信息
    shipment_info = extract_shipment_info(logistics_file_path)
    
    # 2. 登记并创建文件夹
    new_code, folder_path = register_shipment(shipment_info)
    
    return new_code, folder_path, shipment_info
