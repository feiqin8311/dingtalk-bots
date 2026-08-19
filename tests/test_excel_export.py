from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "apps" / "track_notify"
if str(APP) not in sys.path:
    sys.path.insert(0, str(APP))

from excel_export import (  # noqa: E402
    HEADERS,
    ReportItem,
    detail_line_has_date,
    export_filename,
    filter_report_item,
    format_shipped_at,
    write_report_xlsx,
)
from owners import KEPENGXIANG_USER_ID  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from datetime import datetime  # noqa: E402


class ExcelExportTests(unittest.TestCase):
    def test_export_filename_cn_name_date(self):
        when = datetime(2026, 8, 5, 11, 43, 36)
        self.assertEqual(
            export_filename(user_id="17409662804279906", when=when),
            "乔丹丹_轨迹回传_20260805_114336.xlsx",
        )
        self.assertEqual(
            export_filename(user_id="17331048354297047", when=when),
            "柯鹏翔_轨迹回传_20260805_114336.xlsx",
        )
        self.assertEqual(
            export_filename(user_id="no_owner", when=when),
            "无负责人_轨迹回传_20260805_114336.xlsx",
        )

    def test_headers_order(self):
        self.assertEqual(
            HEADERS,
            (
                "发票号",
                "品牌",
                "国家",
                "FBA编码",
                "物流编号",
                "货代公司",
                "发货时间",
                "预计船期",
                "实际送仓时间",
                "负责人",
                "物流详情",
            ),
        )

    def test_format_shipped_at(self):
        self.assertEqual(format_shipped_at(date(2026, 7, 21)), "2026-07-21")
        self.assertEqual(format_shipped_at(None), "")

    def test_write_report_xlsx(self):
        items = [
            ReportItem(
                shipment_key="FBA1",
                event_key="k1",
                message="FBA1 2026-07-01 离港",
                user_ids=["u1"],
                invoice_no="26YP075",
                brand="YPLUS",
                country="美国",
                fba_code="FBA1",
                logistics_no="",
                carrier="平谊",
                shipped_at="2026-07-21",
                eta_date="2026-08-01",
                delivered_at="",
                owners="芋圆",
                detail="2026-07-01 离港",
            ),
            ReportItem(
                shipment_key="r1",
                event_key="missing_fba",
                message="26LBA22 无FBA编码，无法查询平谊轨迹",
                user_ids=["u1"],
                invoice_no="26LBA22",
                brand="LIBRATON",
                country="美国",
                fba_code="",
                logistics_no="",
                carrier="平谊",
                shipped_at="2026-07-21",
                eta_date="2026-07-30",
                delivered_at="2026-08-02",
                owners="芋圆",
                detail="26LBA22 无FBA编码，无法查询平谊轨迹",
            ),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = write_report_xlsx(items, Path(tmp) / "t.xlsx")
            self.assertTrue(path.is_file())
            wb = load_workbook(path)
            ws = wb.active
            self.assertEqual([c.value for c in ws[1]], list(HEADERS))
            self.assertEqual(ws.max_row, 3)
            self.assertEqual(ws["A2"].value, "26YP075")
            self.assertEqual(ws["C2"].value, "美国")
            self.assertEqual(ws["D2"].value, "FBA1")
            self.assertEqual(ws["G2"].value, "2026-07-21")
            self.assertEqual(ws["H2"].value, "2026-08-01")
            self.assertEqual(ws["I2"].value, None)  # openpyxl empty → None
            self.assertEqual(ws["J2"].value, "芋圆")
            self.assertEqual(ws["K2"].value, "2026-07-01 离港")
            self.assertEqual(ws["H3"].value, "2026-07-30")
            self.assertEqual(ws["I3"].value, "2026-08-02")
            self.assertEqual(ws["K3"].value, "26LBA22 无FBA编码，无法查询平谊轨迹")

    def test_multiline_detail_one_cell(self):
        detail = "2026-07-01 离港\n2026-07-10 清关\n2026-07-15 派送中"
        items = [
            ReportItem(
                shipment_key="FBA1",
                event_key="k1",
                message="FBA1 轨迹3条",
                user_ids=["u1"],
                invoice_no="26YP075",
                brand="YPLUS",
                country="美国",
                fba_code="FBA1",
                logistics_no="",
                carrier="平谊",
                shipped_at="2026-07-21",
                owners="乔丹-Joyce",
                detail=detail,
                event_keys=["k1", "k2", "k3"],
            ),
        ]
        self.assertEqual(items[0].keys_to_mark(), ["k1", "k2", "k3"])
        with tempfile.TemporaryDirectory() as tmp:
            path = write_report_xlsx(items, Path(tmp) / "m.xlsx")
            wb = load_workbook(path)
            ws = wb.active
            self.assertEqual(ws.max_row, 2)
            self.assertEqual(ws["K2"].value, detail)
            self.assertTrue(ws["K2"].alignment.wrap_text)

    def test_detail_line_has_date(self):
        self.assertTrue(detail_line_has_date("2026-07-29 出发地 中国宁波"))
        self.assertFalse(detail_line_has_date("已到达卸货港"))
        self.assertFalse(detail_line_has_date("已提柜"))
        self.assertFalse(detail_line_has_date("货物已送达 FC 场地"))

    def test_filter_mixed_keep_dated_for_all(self):
        detail = (
            "2026-08-13 出发地 中国宁波\n"
            "已到达卸货港\n"
            "已提柜\n"
            "货物已送达 FC 场地"
        )
        item = ReportItem(
            shipment_key="AL0",
            event_key="lz:ningbo",
            message="AL0 节点4条",
            user_ids=["17409662804279906", KEPENGXIANG_USER_ID],
            detail=detail,
            event_keys=["lz:ningbo", "lz:pod", "lz:pickup", "lz:fc"],
        )
        only_dated = "2026-08-13 出发地 中国宁波"
        view = filter_report_item(item)
        assert view is not None
        self.assertEqual(view.detail, only_dated)

    def test_filter_all_undated_omit_everyone(self):
        item = ReportItem(
            shipment_key="AL0",
            event_key="lz:pod",
            message="x",
            user_ids=["u1", KEPENGXIANG_USER_ID],
            detail="已到达卸货港\n已提柜",
            event_keys=["lz:pod", "lz:pickup"],
        )
        self.assertIsNone(filter_report_item(item))

    def test_filter_issue_row_kept(self):
        item = ReportItem(
            shipment_key="r1",
            event_key="no_track",
            message="26EA193 AL0 未查到轨迹",
            user_ids=[KEPENGXIANG_USER_ID],
            detail="26EA193 AL0 未查到轨迹",
        )
        self.assertIs(filter_report_item(item), item)


if __name__ == "__main__":
    unittest.main()
